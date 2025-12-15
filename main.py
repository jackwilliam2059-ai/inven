import tkinter as tk
from tkinter import ttk, messagebox, filedialog, simpledialog
import json
import os
from datetime import datetime
from PIL import Image, ImageTk
import base64
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from tkcalendar import DateEntry
import platform
import socket
import time
import threading
import urllib.request
import urllib.parse
from urllib.error import URLError, HTTPError
import re

# Import our refactored modules
import utils
import os
from dotenv import load_dotenv

# .env 파일 로드
load_dotenv()

# Supabase 모드 확인 (기본: Supabase 사용)
USE_SUPABASE = os.getenv('USE_SUPABASE', 'true').lower() == 'true'
USE_SQLALCHEMY = os.getenv('USE_SQLALCHEMY', 'true').lower() == 'true'  # SQLAlchemy 직접 연결

if USE_SUPABASE:
    if USE_SQLALCHEMY:
        from sqlalchemy_data_manager import SQLAlchemyDataManager as DataManager
        print("🌐 SQLAlchemy 모드로 실행 (PostgreSQL 직접 연결)")
    else:
        from supabase_data_manager import DataManager
        print("🌐 Supabase SDK 모드로 실행")
else:
    from data_manager import DataManager
    print("📁 JSON 모드로 실행")

class InventoryManagementSystem:
    def __init__(self, root):
        self.root = root
        self.root.title("재고 관리 시스템")
        self.root.geometry("1400x800")

        # Initialize DataManager for all data operations
        self.data_manager = DataManager()

        # Set up shortcuts to DataManager properties for backward compatibility
        self.config_file = self.data_manager.config_file
        self.cloud_info = self.data_manager.cloud_info
        self.cloud_path = self.data_manager.cloud_path
        self.cloud_type = self.data_manager.cloud_type
        self.data_file = self.data_manager.data_file
        self.lock_file = self.data_manager.lock_file
        self.users_file = self.data_manager.users_file

        # Shortcuts to data arrays
        self.products = self.data_manager.products
        self.orders = self.data_manager.orders
        self.movements = self.data_manager.movements
        self.inbound_records = self.data_manager.inbound_records
        self.outbound_records = self.data_manager.outbound_records
        self.stores = self.data_manager.stores
        self.field_names = self.data_manager.field_names
        self.settlement_balances = self.data_manager.settlement_balances

        self.image_cache = {}

        # 로컬 이미지 폴더
        self.cloud_image_folder = "cloud_images"
        if not os.path.exists(self.cloud_image_folder):
            os.makedirs(self.cloud_image_folder)

        # 클라우드 스토리지 이미지 폴더 (클라우드 경로가 있을 경우)
        if self.cloud_path and os.path.exists(self.cloud_path):
            cloud_images_path = os.path.join(self.cloud_path, "images")
            if not os.path.exists(cloud_images_path):
                try:
                    os.makedirs(cloud_images_path)
                    print(f"클라우드 이미지 폴더 생성: {cloud_images_path}")
                except:
                    pass

        # 다중 사용자 기능
        self.current_user = self.data_manager.current_user
        self.is_locked = self.data_manager.is_locked
        self.lock_check_interval = 5000  # 5초마다 잠금 확인

        # 사용자 이름 기본값 설정 (setup_user_name에서 덮어씌워질 수 있음)
        self.user_display_name = self.data_manager.user_display_name

        # 데이터 로드 여부 추적
        self.data_loaded_by_user = False  # 사용자가 직접 데이터를 불러왔는지 추적

        # 데이터 로드
        self.data_manager.load_data()
        # Refresh shortcuts after loading
        self._refresh_data_shortcuts()
        
        # Realtime 구독 시작 (Supabase 모드인 경우)
        if USE_SUPABASE and hasattr(self.data_manager, 'start_realtime'):
            self.data_manager.set_on_change_callback(self._on_realtime_change)
            self.data_manager.start_realtime()

        # UI 생성
        self.create_ui()

        # 사용자 이름 설정 (UI 생성 후 실행)
        self.root.after(100, self.setup_user_name)

        # 백업 데이터 확인 및 복원 (사용자 설정 후 실행)
        self.root.after(200, self.check_and_restore_backup)

        # 잠금 체크 시작
        self.start_lock_check()

        # 프로그램 종료 시 정리
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)

    def _refresh_data_shortcuts(self):
        """Refresh shortcuts to DataManager data arrays"""
        self.products = self.data_manager.products
        self.orders = self.data_manager.orders
        self.movements = self.data_manager.movements
        self.inbound_records = self.data_manager.inbound_records
        self.outbound_records = self.data_manager.outbound_records
        self.stores = self.data_manager.stores
        self.field_names = self.data_manager.field_names
        self.settlement_balances = self.data_manager.settlement_balances
    
    def _on_realtime_change(self, table: str, event_type: str):
        """Realtime 변경 이벤트 콜백 (다른 스레드에서 호출됨)"""
        # tkinter는 메인 스레드에서만 UI 업데이트 가능
        # root.after()로 메인 스레드에서 실행되도록 예약
        self.root.after(0, lambda: self._handle_realtime_update(table, event_type))
    
    def _handle_realtime_update(self, table: str, event_type: str):
        """Realtime 업데이트 처리 (메인 스레드에서 실행)"""
        try:
            # 새 상품 감지를 위해 기존 상품 ID 저장
            old_product_ids = set(p['id'] for p in self.products) if hasattr(self, 'products') else set()
            
            # 데이터 참조 갱신
            self._refresh_data_shortcuts()
            
            # 테이블별 UI 새로고침
            if table == 'products':
                self.refresh_products_list()
                self.refresh_stock_list()
                
                # 새로 추가된 상품 중 임시 상품 확인
                new_product_ids = set(p['id'] for p in self.products)
                added_ids = new_product_ids - old_product_ids
                
                for pid in added_ids:
                    product = self.data_manager.get_product_by_id(pid)
                    if product and '발주노트에서 임시 추가됨' in (product.get('memo') or ''):
                        self._show_new_product_popup(product)
                        
            elif table == 'orders':
                self.refresh_orders_list()
                self.refresh_stock_list()
            elif table == 'inbound_records':
                self.refresh_inbound_list()
                self.refresh_stock_list()
                self.refresh_products_list()
            elif table == 'outbound_records':
                self.refresh_outbound_list()
                self.refresh_stock_list()
                self.refresh_products_list()
            elif table == 'inventory_movements':
                self.refresh_stock_list()
            elif table == 'stores':
                # 매장 변경 시 관련 콤보박스 업데이트 필요
                pass
            
            # 상태바에 표시 (있는 경우)
            if hasattr(self, 'status_label'):
                self.status_label.config(text=f"📡 동기화: {table} {event_type}")
                # 3초 후 원래 상태로
                self.root.after(3000, lambda: self.status_label.config(text=""))
                
        except Exception as e:
            print(f"⚠️ UI 업데이트 오류: {e}")
        self.outbound_records = self.data_manager.outbound_records
        self.stores = self.data_manager.stores
        self.field_names = self.data_manager.field_names
        self.settlement_balances = self.data_manager.settlement_balances
        self.user_display_name = self.data_manager.user_display_name
        self.is_locked = self.data_manager.is_locked
    
    def _show_new_product_popup(self, product):
        """발주노트에서 추가된 새 상품 알림 팝업"""
        dialog = tk.Toplevel(self.root)
        dialog.title("🆕 새 상품 추가됨")
        dialog.geometry("450x350")
        dialog.transient(self.root)
        dialog.grab_set()
        utils.center_window(dialog)
        
        # 알림 메시지
        ttk.Label(dialog, text="📦 발주노트에서 새 상품이 추가되었습니다!", 
                  font=("Arial", 12, "bold")).pack(pady=15)
        
        # 상품 정보 표시
        info_frame = ttk.LabelFrame(dialog, text="상품 정보", padding=10)
        info_frame.pack(fill=tk.X, padx=20, pady=10)
        
        ttk.Label(info_frame, text=f"상품명: {product.get('name', '-')}").pack(anchor='w', pady=2)
        ttk.Label(info_frame, text=f"상품코드: {product.get('code', '-') or '-'}").pack(anchor='w', pady=2)
        ttk.Label(info_frame, text=f"색상: {', '.join(product.get('colors', [])) or '-'}").pack(anchor='w', pady=2)
        ttk.Label(info_frame, text=f"사이즈: {', '.join(product.get('sizes', [])) or '-'}").pack(anchor='w', pady=2)
        
        ttk.Label(dialog, text="자세한 정보를 수정하시겠습니까?", 
                  font=("Arial", 10)).pack(pady=15)
        
        btn_frame = ttk.Frame(dialog)
        btn_frame.pack(pady=10)
        
        def edit_product():
            dialog.destroy()
            # 상품관리 탭으로 이동
            self.notebook.select(self.products_frame)
            # 해당 상품 선택
            for item in self.products_tree.get_children():
                values = self.products_tree.item(item, 'values')
                if values and str(values[0]) == str(product['id']):
                    self.products_tree.selection_set(item)
                    self.products_tree.see(item)
                    break
            # 수정 다이얼로그 열기
            self.root.after(100, self.edit_product)
        
        ttk.Button(btn_frame, text="✏️ 수정하기", command=edit_product, width=15).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="나중에", command=dialog.destroy, width=15).pack(side=tk.LEFT, padx=5)
    
    def setup_user_name(self):
        """사용자 이름 설정"""
        # 저장된 사용자 이름 확인
        users_config = self.data_manager.load_users_config()

        if self.current_user not in users_config:
            # 새 사용자 - 이름 입력 받기
            dialog = tk.Toplevel()
            dialog.title("사용자 설정")
            dialog.geometry("400x200")
            dialog.transient(self.root)
            dialog.grab_set()
            utils.center_window(dialog)

            ttk.Label(dialog, text="재고관리 시스템 사용자 등록", font=("Arial", 12, "bold")).pack(pady=20)
            ttk.Label(dialog, text="표시될 이름을 입력하세요:").pack(pady=5)

            name_var = tk.StringVar(value=self.user_display_name)  # 기본값 사용
            name_entry = ttk.Entry(dialog, textvariable=name_var, width=30)
            name_entry.pack(pady=10)
            name_entry.focus()

            def save_name():
                name = name_var.get().strip()
                if name:
                    users_config[self.current_user] = {
                        'display_name': name,
                        'registered_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    }
                    self.data_manager.save_users_config(users_config)
                    self.user_display_name = name
                    self.data_manager.user_display_name = name
                    dialog.destroy()
                else:
                    messagebox.showwarning("경고", "이름을 입력해주세요.")

            def on_dialog_close():
                """다이얼로그가 닫힐 때 기본값으로 저장"""
                # 사용자가 설정하지 않고 닫은 경우 기본값을 저장하여 다음에 다시 묻지 않음
                users_config[self.current_user] = {
                    'display_name': self.user_display_name,
                    'registered_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                }
                self.data_manager.save_users_config(users_config)
                dialog.destroy()

            ttk.Button(dialog, text="확인", command=save_name).pack(pady=10)
            dialog.protocol("WM_DELETE_WINDOW", on_dialog_close)

            dialog.wait_window()
        else:
            # 기존 사용자 - 저장된 이름 사용
            self.user_display_name = users_config[self.current_user]['display_name']
    
    def start_lock_check(self):
        """주기적으로 잠금 상태 확인"""
        def check_and_update():
            if self.is_locked:
                # 자신의 잠금 갱신
                self.data_manager.update_lock()
            else:
                # 다른 사용자가 사용 중인지 확인
                locked, user_name = self.data_manager.check_lock()
                if locked:
                    self.show_lock_warning(user_name)
            
            # 다음 체크 예약
            self.root.after(self.lock_check_interval, check_and_update)
        
        self.root.after(self.lock_check_interval, check_and_update)
    
    def show_lock_warning(self, user_name):
        """다른 사용자 사용 중 경고"""
        # 읽기 전용 모드로 전환
        if not hasattr(self, 'lock_warning_shown'):
            self.lock_warning_shown = True
            messagebox.showwarning("다른 사용자 사용 중", 
                f"{user_name} 님이 현재 작업 중입니다.\n"
                f"데이터를 볼 수는 있지만 수정할 수 없습니다.\n"
                f"작업이 필요하면 잠시 후 다시 시도해주세요.")
            self.root.title(f"재고 관리 시스템 - 읽기 전용 ({user_name} 님 작업 중)")
    
    def try_acquire_lock_for_edit(self):
        """편집을 위한 잠금 시도"""
        locked, user_name = self.data_manager.check_lock()
        if locked:
            messagebox.showwarning("사용 중", 
                f"{user_name} 님이 현재 작업 중입니다.\n"
                f"잠시 후 다시 시도해주세요.")
            return False
        
        return self.data_manager.acquire_lock()
    
    def check_and_restore_backup(self):
        """프로그램 시작 시 최신 백업과 현재 데이터 비교 후 복원"""
        backup_info = self.data_manager.check_and_restore_backup()

        if backup_info is None:
            return

        try:
            if backup_info['type'] == 'no_current':
                # 현재 데이터 파일이 없는 경우
                if messagebox.askyesno("백업 발견",
                    f"자동 백업된 데이터를 발견했습니다.\n"
                    f"백업 파일: {backup_info['backup_name']}\n\n"
                    f"백업 데이터를 불러오시겠습니까?"):

                    with open(backup_info['backup_file'], 'r', encoding='utf-8') as f:
                        data = json.load(f)

                    self.data_manager.restore_from_backup(data)
                    self._refresh_data_shortcuts()
                    self.data_loaded_by_user = True  # 데이터 불러옴 표시

                    # UI가 생성되어 있으면 새로고침
                    if hasattr(self, 'product_tree'):
                        self.refresh_all()

                    messagebox.showinfo("복원 완료", "백업 데이터를 성공적으로 불러왔습니다.")

            elif backup_info['type'] == 'newer_backup':
                # 백업이 더 최신인 경우
                if messagebox.askyesno("최신 백업 발견",
                    f"더 최신의 백업 데이터를 발견했습니다.\n\n"
                    f"현재 데이터:\n"
                    f"  - 저장 시간: {backup_info['current_time']}\n"
                    f"  - 저장자: {backup_info['current_user']}\n\n"
                    f"백업 데이터:\n"
                    f"  - 저장 시간: {backup_info['backup_time']}\n"
                    f"  - 저장자: {backup_info['backup_user']}\n\n"
                    f"백업 데이터로 복원하시겠습니까?"):

                    self.data_manager.restore_from_backup(backup_info['backup_data'])
                    self._refresh_data_shortcuts()
                    self.data_loaded_by_user = True  # 데이터 불러옴 표시

                    # UI가 생성되어 있으면 새로고침
                    if hasattr(self, 'product_tree'):
                        self.refresh_all()

                    messagebox.showinfo("복원 완료", "백업 데이터를 성공적으로 불러왔습니다.")

        except Exception as e:
            print(f"백업 확인 중 오류: {str(e)}")
            messagebox.showwarning("백업 확인 오류",
                f"백업 파일 확인 중 오류가 발생했습니다:\n{str(e)}\n\n"
                f"현재 데이터로 계속 진행합니다.")

    def on_closing(self):
        """프로그램 종료 시"""
        result = messagebox.askyesnocancel(
            "프로그램 종료",
            "작업 내용을 저장하시겠습니까?\n\n"
            "예: 저장하고 종료\n"
            "아니오: 저장하지 않고 종료\n"
            "취소: 종료 취소"
        )
        
        if result is None:  # 취소
            return
        elif result:  # 예 - 저장하고 종료
            try:
                self.data_manager.save_data()
                self.data_manager.auto_backup_data()
            except Exception as e:
                if not messagebox.askyesno("저장 오류", 
                    f"저장 중 오류가 발생했습니다:\n{str(e)}\n\n그래도 종료하시겠습니까?"):
                    return
        # 아니오 - 저장하지 않고 종료

        # Realtime 구독 중지
        if hasattr(self.data_manager, 'stop_realtime'):
            self.data_manager.stop_realtime()
        
        self.data_manager.release_lock()
        self.root.destroy()
    
    def change_cloud_path(self):
        """클라우드 스토리지 경로 변경 다이얼로그"""
        dialog = tk.Toplevel(self.root)
        dialog.title("클라우드 스토리지 설정")
        dialog.geometry("700x550")
        dialog.transient(self.root)
        dialog.grab_set()
        utils.center_window(dialog)
        
        ttk.Label(dialog, text="☁️ 데이터 저장 위치 설정", font=("Arial", 14, "bold")).pack(pady=20)
        
        # 현재 설정 표시
        info_frame = ttk.LabelFrame(dialog, text="현재 설정", padding=10)
        info_frame.pack(fill=tk.BOTH, padx=20, pady=10)
        
        current_path = self.cloud_path if self.cloud_path else "로컬 (프로그램 폴더)"
        current_type = self.cloud_type if self.cloud_type != 'local' else "로컬 저장"
        ttk.Label(info_frame, text=f"저장 타입: {current_type}", wraplength=600).pack(anchor='w', pady=3)
        ttk.Label(info_frame, text=f"저장 위치: {current_path}", wraplength=600).pack(anchor='w', pady=3)
        ttk.Label(info_frame, text=f"데이터 파일: {self.data_file}", wraplength=600).pack(anchor='w', pady=3)
        
        # 옵션 선택
        option_frame = ttk.LabelFrame(dialog, text="저장 위치 변경", padding=10)
        option_frame.pack(fill=tk.BOTH, padx=20, pady=10)
        
        def use_local():
            """로컬 저장 사용"""
            if messagebox.askyesno("확인", "로컬 폴더를 사용하시겠습니까?\n현재 데이터를 로컬로 복사합니다."):
                old_file = self.data_file
                self.cloud_path = ''
                self.cloud_type = 'local'
                self.data_file = "inventory_data.json"
                self.data_manager.save_cloud_path('', 'local')
                
                # 기존 데이터 복사
                if os.path.exists(old_file) and old_file != self.data_file:
                    try:
                        import shutil
                        shutil.copy2(old_file, self.data_file)
                        messagebox.showinfo("완료", "로컬 저장으로 변경되었습니다.")
                    except Exception as e:
                        messagebox.showerror("오류", f"데이터 복사 중 오류:\n{str(e)}")
                
                dialog.destroy()
                self.update_title()
        
        def select_cloud_folder():
            """클라우드 폴더 선택"""
            path = filedialog.askdirectory(
                title="클라우드 폴더 선택 (OneDrive, Google Drive, Dropbox 등)", 
                initialdir=self.cloud_path or os.path.expanduser('~')
            )
            
            if path:
                # 경로로 클라우드 타입 추측
                cloud_type = 'Cloud Storage'
                path_lower = path.lower()
                if 'onedrive' in path_lower:
                    cloud_type = 'OneDrive'
                elif 'google' in path_lower or 'drive' in path_lower:
                    cloud_type = 'Google Drive'
                elif 'dropbox' in path_lower:
                    cloud_type = 'Dropbox'
                elif 'naver' in path_lower or '네이버' in path_lower:
                    cloud_type = 'Naver Cloud'
                elif 'icloud' in path_lower:
                    cloud_type = 'iCloud Drive'
                
                old_file = self.data_file
                self.cloud_path = path
                self.cloud_type = cloud_type
                self.data_file = os.path.join(path, "inventory_data.json")
                self.data_manager.save_cloud_path(path, cloud_type)
                
                # 기존 데이터 복사
                if os.path.exists(old_file) and old_file != self.data_file:
                    try:
                        import shutil
                        shutil.copy2(old_file, self.data_file)
                        messagebox.showinfo("완료", f"{cloud_type} 저장으로 변경되었습니다.\n위치: {path}")
                    except Exception as e:
                        messagebox.showerror("오류", f"데이터 복사 중 오류:\n{str(e)}")
                
                dialog.destroy()
                self.update_title()
        
        def auto_detect():
            """자동 감지"""
            detected = utils.auto_detect_cloud()
            if detected and detected.get('path'):
                cloud_name = detected.get('type', '클라우드')
                if messagebox.askyesno("확인", 
                    f"다음 {cloud_name} 경로를 찾았습니다:\n{detected['path']}\n\n이 경로를 사용하시겠습니까?"):
                    old_file = self.data_file
                    self.cloud_path = detected['path']
                    self.cloud_type = detected['type']
                    self.data_file = os.path.join(detected['path'], "inventory_data.json")
                    self.data_manager.save_cloud_path(detected['path'], detected['type'])
                    
                    # 기존 데이터 복사
                    if os.path.exists(old_file) and old_file != self.data_file:
                        try:
                            import shutil
                            shutil.copy2(old_file, self.data_file)
                        except:
                            pass
                    
                    messagebox.showinfo("완료", f"{cloud_name} 경로가 설정되었습니다.")
                    dialog.destroy()
                    self.update_title()
            else:
                messagebox.showwarning("알림", 
                    "클라우드 스토리지 경로를 자동으로 찾을 수 없습니다.\n"
                    "수동으로 클라우드 폴더를 선택해주세요.")
        
        ttk.Button(option_frame, text="🔍 자동으로 클라우드 찾기", command=auto_detect, width=40).pack(pady=5)
        ttk.Button(option_frame, text="📁 클라우드 폴더 직접 선택", command=select_cloud_folder, width=40).pack(pady=5)
        ttk.Button(option_frame, text="☁️ API 클라우드 공유 설정", command=self.setup_api_cloud, width=40).pack(pady=5)
        ttk.Button(option_frame, text="💻 로컬 폴더 사용 (클라우드 사용 안 함)", command=use_local, width=40).pack(pady=5)
        
        # 설명
        help_frame = ttk.Frame(dialog)
        help_frame.pack(fill=tk.BOTH, padx=20, pady=10)
        
        help_text = """
💡 지원하는 클라우드 스토리지:
[폴더 기반 방식]
• OneDrive (원드라이브)
• Google Drive (구글 드라이브)
• Dropbox (드롭박스)
• Naver Cloud (네이버 클라우드)
• iCloud Drive (아이클라우드)

[API 기반 방식]
• Firebase Realtime Database
• Supabase (PostgreSQL + API)
• AWS S3, Google Cloud Storage
• Azure Blob Storage, Dropbox API 등

✨ 클라우드 사용의 장점:
• 여러 컴퓨터에서 데이터 동기화
• 자동 백업 및 버전 관리
• 데이터 안전성 향상
• API 방식: 클라우드 앱 설치 불필요

⚠️ 주의사항:
• 폴더 방식: 클라우드 앱이 설치되어 동기화되어야 합니다
• API 방식: 유효한 API 자격증명 필요
• 인터넷 연결 시 자동 동기화됩니다
        """
        ttk.Label(help_frame, text=help_text, justify=tk.LEFT, foreground="gray", font=("Arial", 8)).pack(anchor='w')
        
        ttk.Button(dialog, text="닫기", command=dialog.destroy).pack(pady=10)
    
    def setup_api_cloud(self):
        """API 기반 클라우드 공유 설정"""
        dialog = tk.Toplevel(self.root)
        dialog.title("API 클라우드 공유 설정")
        dialog.geometry("700x550")
        dialog.transient(self.root)
        dialog.grab_set()
        utils.center_window(dialog)
        
        ttk.Label(dialog, text="☁️ API 기반 클라우드 공유", font=("Arial", 14, "bold")).pack(pady=20)
        
        # 서비스 선택 프레임
        service_frame = ttk.LabelFrame(dialog, text="클라우드 서비스 선택", padding=10)
        service_frame.pack(fill=tk.BOTH, padx=20, pady=10)
        
        service_var = tk.StringVar(value="firebase")
        
        services = [
            ("Firebase Realtime Database", "firebase", "실시간 동기화, 권한 관리 지원"),
            ("Supabase (PostgreSQL + API)", "supabase", "오픈소스 Firebase 대체, SQL 지원"),
            ("AWS S3", "aws_s3", "높은 신뢰성, 엔터프라이즈급"),
            ("Google Cloud Storage", "gcs", "Google 클라우드 플랫폼"),
            ("Azure Blob Storage", "azure", "Microsoft Azure 통합"),
            ("Dropbox API", "dropbox_api", "Dropbox를 API로 관리"),
            ("OneDrive API", "onedrive_api", "Microsoft OneDrive 통합"),
        ]
        
        for service_name, service_id, description in services:
            frame = ttk.Frame(service_frame)
            frame.pack(fill=tk.X, pady=5)
            ttk.Radiobutton(frame, text=service_name, variable=service_var, value=service_id).pack(side=tk.LEFT)
            ttk.Label(frame, text=description, foreground="gray", font=("Arial", 8)).pack(side=tk.LEFT, padx=20)
        
        # API 키 입력 프레임
        api_frame = ttk.LabelFrame(dialog, text="API 자격증명", padding=10)
        api_frame.pack(fill=tk.BOTH, padx=20, pady=10)
        
        ttk.Label(api_frame, text="API 엔드포인트 또는 프로젝트 ID:", font=("Arial", 9)).pack(anchor='w')
        endpoint_entry = ttk.Entry(api_frame, width=50)
        endpoint_entry.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(api_frame, text="API 키:", font=("Arial", 9)).pack(anchor='w')
        key_entry = ttk.Entry(api_frame, width=50, show="*")
        key_entry.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(api_frame, text="API 시크릿 (선택사항):", font=("Arial", 9)).pack(anchor='w')
        secret_entry = ttk.Entry(api_frame, width=50, show="*")
        secret_entry.pack(fill=tk.X)
        
        # 정보
        info_frame = ttk.Frame(dialog)
        info_frame.pack(fill=tk.BOTH, padx=20, pady=10)
        
        info_text = """
📌 API 클라우드 공유 가이드:

1. Firebase: Google Firebase 콘솔에서 Realtime Database 생성 후 자격증명 복사
2. Supabase: supabase.com에서 프로젝트 생성 후 API URL 및 키 복사
3. AWS S3: AWS IAM 사용자 생성 후 Access Key ID, Secret Access Key 입력
4. Google Cloud Storage: Google Cloud 콘솔에서 JSON 서비스 계정 키 생성
5. Azure: Azure Storage 계정 생성 후 Connection String 입력
6. Dropbox API: app.dropbox.com에서 앱 생성 후 Access Token 생성
7. OneDrive API: Microsoft Azure AD에서 앱 등록 후 자격증명 생성

✨ API 공유의 장점:
• 인터넷만 있으면 어디서나 접근 가능
• 다중 장치 실시간 동기화
• 클라우드 앱 설치 불필요
• 고급 권한 관리 지원
        """
        ttk.Label(info_frame, text=info_text, justify=tk.LEFT, foreground="gray", font=("Arial", 8)).pack(anchor='w')
        
        # 버튼 프레임
        button_frame = ttk.Frame(dialog)
        button_frame.pack(fill=tk.X, padx=20, pady=10)
        
        def save_api_config():
            service = service_var.get()
            endpoint = endpoint_entry.get().strip()
            api_key = key_entry.get().strip()
            api_secret = secret_entry.get().strip()
            
            if not endpoint or not api_key:
                messagebox.showwarning("경고", "엔드포인트와 API 키를 입력해주세요.")
                return
            
            try:
                config = {
                    'path': '',
                    'type': 'api',
                    'api_service': service,
                    'api_endpoint': endpoint,
                    'api_key': api_key,
                    'api_secret': api_secret if api_secret else '',
                    'last_updated': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                }
                
                with open(self.config_file, 'w', encoding='utf-8') as f:
                    json.dump(config, f, ensure_ascii=False, indent=2)
                
                self.cloud_type = 'api'
                self.cloud_path = ''  # API는 경로 없음
                
                messagebox.showinfo("완료", f"API 클라우드 공유가 설정되었습니다.\n서비스: {service}")
                dialog.destroy()
                self.update_title()
            except Exception as e:
                messagebox.showerror("오류", f"설정 저장 중 오류가 발생했습니다:\n{str(e)}")
        
        def open_guide():
            """API 설정 가이드 열기"""
            guide_window = tk.Toplevel(dialog)
            guide_window.title("API 설정 상세 가이드")
            guide_window.geometry("800x600")
            
            guide_text = tk.Text(guide_window, wrap=tk.WORD, font=("Courier", 9))
            guide_text.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
            
            guide_content = """
=== Firebase Realtime Database ===
1. https://console.firebase.google.com 접속
2. 새 프로젝트 생성
3. Realtime Database 생성
4. 규칙 탭에서 {"rules": {".read": true, ".write": true}} 설정
5. 설정 > 서비스 계정 > Database 시크릿에서 키 복사

API 엔드포인트: https://your-project.firebaseio.com
API 키: Database Secret

=== Supabase ===
1. https://supabase.com 접속
2. 새 프로젝트 생성
3. Settings > API에서 자격증명 복사
4. URL과 anon key 복사

API 엔드포인트: https://xxxxx.supabase.co
API 키: anon key

=== AWS S3 ===
1. AWS Console 접속
2. IAM 사용자 생성
3. S3 Access Key ID, Secret Access Key 복사
4. S3 버킷 생성

API 엔드포인트: your-bucket-name
API 키: Access Key ID
API 시크릿: Secret Access Key

=== Google Cloud Storage ===
1. Google Cloud Console 접속
2. 서비스 계정 생성
3. JSON 키 다운로드
4. 프로젝트 ID와 프라이빗 키 복사

API 엔드포인트: project-id
API 키: private-key (전체 복사)

=== Azure Blob Storage ===
1. Azure Portal 접속
2. Storage Account 생성
3. Access keys에서 Connection String 복사

API 엔드포인트: storage-account-name
API 키: Connection String

=== Dropbox API ===
1. https://www.dropbox.com/developers 접속
2. 앱 생성 (Scoped Access, Full Dropbox)
3. 생성된 앱에서 Generate access token

API 엔드포인트: dropbox
API 키: Access Token

=== OneDrive API ===
1. https://portal.azure.com 접속
2. App registrations 에서 새 앱 등록
3. Certificates & secrets에서 Client secret 생성
4. Client ID와 Client secret 복사

API 엔드포인트: Client ID
API 키: Client secret
"""
            
            guide_text.insert(1.0, guide_content)
            guide_text.config(state=tk.DISABLED)
            
            scrollbar = ttk.Scrollbar(guide_window, orient=tk.VERTICAL, command=guide_text.yview)
            scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
            guide_text.config(yscrollcommand=scrollbar.set)
        
        ttk.Button(button_frame, text="📖 상세 설정 가이드", command=open_guide).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="✅ 설정 저장", command=save_api_config).pack(side=tk.RIGHT, padx=5)
        ttk.Button(button_frame, text="❌ 취소", command=dialog.destroy).pack(side=tk.RIGHT, padx=5)
    
    def show_user_info(self):
        """현재 사용자 정보 표시"""
        dialog = tk.Toplevel(self.root)
        dialog.title("사용자 정보")
        dialog.geometry("500x400")
        dialog.transient(self.root)
        dialog.grab_set()
        utils.center_window(dialog)
        
        ttk.Label(dialog, text="👤 사용자 정보", font=("Arial", 14, "bold")).pack(pady=20)
        
        info_frame = ttk.LabelFrame(dialog, text="내 정보", padding=15)
        info_frame.pack(fill=tk.BOTH, padx=20, pady=10)
        
        ttk.Label(info_frame, text=f"표시 이름: {self.user_display_name}", font=("Arial", 10)).pack(anchor='w', pady=5)
        ttk.Label(info_frame, text=f"컴퓨터: {socket.gethostname()}", font=("Arial", 10)).pack(anchor='w', pady=5)
        ttk.Label(info_frame, text=f"사용자 계정: {os.environ.get('USERNAME', 'Unknown')}", font=("Arial", 10)).pack(anchor='w', pady=5)
        
        # 작업 상태
        status_frame = ttk.LabelFrame(dialog, text="작업 상태", padding=15)
        status_frame.pack(fill=tk.BOTH, padx=20, pady=10)
        
        if self.is_locked:
            ttk.Label(status_frame, text="✅ 현재 작업 중 (편집 가능)", foreground="green", font=("Arial", 10, "bold")).pack(anchor='w', pady=5)
        else:
            locked, user_name = self.data_manager.check_lock()
            if locked:
                ttk.Label(status_frame, text=f"⏸️ {user_name} 님이 작업 중 (읽기 전용)", foreground="orange", font=("Arial", 10)).pack(anchor='w', pady=5)
            else:
                ttk.Label(status_frame, text="💤 대기 중 (편집 시 자동 잠금)", foreground="blue", font=("Arial", 10)).pack(anchor='w', pady=5)
        
        # 이름 변경
        def change_name():
            name_dialog = tk.Toplevel(dialog)
            name_dialog.title("이름 변경")
            name_dialog.geometry("350x150")
            name_dialog.transient(dialog)
            name_dialog.grab_set()
            utils.center_window(name_dialog)
            
            ttk.Label(name_dialog, text="새 이름:").pack(pady=10)
            name_var = tk.StringVar(value=self.user_display_name)
            name_entry = ttk.Entry(name_dialog, textvariable=name_var, width=30)
            name_entry.pack(pady=5)
            
            def save_new_name():
                new_name = name_var.get().strip()
                if new_name:
                    users_config = self.data_manager.load_users_config()
                    users_config[self.current_user]['display_name'] = new_name
                    self.data_manager.save_users_config(users_config)
                    self.user_display_name = new_name
                    self.data_manager.user_display_name = new_name
                    self.update_title()
                    messagebox.showinfo("완료", "이름이 변경되었습니다.")
                    name_dialog.destroy()
                    dialog.destroy()
            
            ttk.Button(name_dialog, text="저장", command=save_new_name).pack(pady=10)
        
        ttk.Button(dialog, text="이름 변경", command=change_name).pack(pady=10)
        ttk.Button(dialog, text="닫기", command=dialog.destroy).pack(pady=5)
    
    def show_active_users(self):
        """접속 중인 사용자 표시"""
        dialog = tk.Toplevel(self.root)
        dialog.title("접속 중인 사용자")
        dialog.geometry("600x400")
        dialog.transient(self.root)
        dialog.grab_set()
        utils.center_window(dialog)
        
        ttk.Label(dialog, text="👥 현재 접속 중인 사용자", font=("Arial", 14, "bold")).pack(pady=20)
        
        # 현재 작업 중인 사용자
        locked, user_name = self.data_manager.check_lock()
        if locked or self.is_locked:
            active_frame = ttk.LabelFrame(dialog, text="✅ 작업 중", padding=15)
            active_frame.pack(fill=tk.BOTH, padx=20, pady=10)
            
            if self.is_locked:
                ttk.Label(active_frame, text=f"👤 {self.user_display_name} (나)", 
                         font=("Arial", 11, "bold"), foreground="green").pack(anchor='w', pady=5)
            else:
                ttk.Label(active_frame, text=f"👤 {user_name}", 
                         font=("Arial", 11, "bold"), foreground="orange").pack(anchor='w', pady=5)
        else:
            ttk.Label(dialog, text="현재 작업 중인 사용자가 없습니다.", 
                     foreground="gray").pack(pady=20)
        
        # 등록된 사용자 목록
        users_config = self.data_manager.load_users_config()
        if users_config:
            users_frame = ttk.LabelFrame(dialog, text="등록된 사용자", padding=15)
            users_frame.pack(fill=tk.BOTH, padx=20, pady=10)
            
            for user_id, user_info in users_config.items():
                display_name = user_info['display_name']
                registered_at = user_info.get('registered_at', 'Unknown')
                is_me = user_id == self.current_user
                
                user_text = f"👤 {display_name}"
                if is_me:
                    user_text += " (나)"
                
                ttk.Label(users_frame, text=user_text, font=("Arial", 10)).pack(anchor='w', pady=3)
        
        ttk.Button(dialog, text="새로고침", command=lambda: [dialog.destroy(), self.show_active_users()]).pack(pady=10)
        ttk.Button(dialog, text="닫기", command=dialog.destroy).pack(pady=5)
    
    def show_multiuser_guide(self):
        """다중 사용자 가이드"""
        dialog = tk.Toplevel(self.root)
        dialog.title("다중 사용자 가이드")
        dialog.geometry("700x600")
        dialog.transient(self.root)
        
        ttk.Label(dialog, text="📖 다중 사용자 사용 가이드", font=("Arial", 14, "bold")).pack(pady=20)
        
        # 스크롤 가능한 텍스트
        text_frame = ttk.Frame(dialog)
        text_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)
        
        scrollbar = ttk.Scrollbar(text_frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        text_widget = tk.Text(text_frame, wrap=tk.WORD, yscrollcommand=scrollbar.set, font=("Arial", 10))
        text_widget.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.config(command=text_widget.yview)
        
        guide_text = """
🎯 다중 사용자 기능 소개

이 재고관리 시스템은 여러 사람이 클라우드 공유 폴더를 통해 
함께 사용할 수 있도록 설계되었습니다.

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

📋 작동 방식

1. 자동 잠금 시스템
   • 데이터를 수정하려고 하면 자동으로 잠금 획득
   • 다른 사용자는 읽기 전용 모드로 전환
   • 작업 완료 후 프로그램 종료 시 잠금 자동 해제

2. 실시간 상태 확인
   • 5초마다 다른 사용자의 작업 상태 확인
   • 다른 사용자가 작업 중이면 자동 알림
   • 창 제목에 현재 상태 표시

3. 사용자 식별
   • 컴퓨터별로 고유 식별
   • 표시 이름으로 누가 작업 중인지 확인
   • 사용 기록 자동 저장

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

🚀 설정 방법

1단계: 클라우드 공유 폴더 준비
   • OneDrive, Google Drive 등에 폴더 생성
   • 해당 폴더를 팀원들과 공유
   • 모든 팀원에게 편집 권한 부여

2단계: 각 컴퓨터에서 설정
   • 재고관리 프로그램 설치
   • 클라우드 앱으로 공유 폴더 동기화
   • 프로그램에서 클라우드 폴더 선택

3단계: 사용자 이름 등록
   • 첫 실행 시 이름 입력 (팀원이 알아볼 수 있는 이름)
   • 예: "김철수", "홍길동", "이영희" 등

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

✅ 올바른 사용법

권장 작업 방식:
1. 프로그램 실행
2. 다른 사용자가 작업 중인지 확인
3. 작업 수행 (자동으로 잠금 획득)
4. 작업 완료 후 프로그램 종료 (잠금 해제)

주의사항:
• 작업이 끝나면 프로그램을 종료하세요
• 오랜 시간 켜두면 다른 사용자가 대기해야 합니다
• 긴급 조회는 가능 (읽기 전용)

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

❌ 피해야 할 상황

1. 동시 작업
   ❌ 여러 사람이 동시에 편집 (데이터 손실 위험)
   ✅ 한 사람씩 순서대로 작업

2. 프로그램 켜두기
   ❌ 작업 안 할 때도 프로그램 실행 상태
   ✅ 작업 끝나면 프로그램 종료

3. 강제 종료
   ❌ 작업 관리자로 프로그램 강제 종료
   ✅ 정상 종료 (X 버튼 클릭)

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

🔧 문제 해결

Q: 다른 사용자가 작업 중이라고 나오는데 실제로는 없어요
A: 30초 후 자동으로 잠금 해제됩니다. 
   또는 클라우드 폴더에서 "inventory_lock.json" 파일을 
   삭제하면 즉시 해제됩니다.

Q: 제가 작업 중인데 다른 사람이 수정할 수 있어요
A: 프로그램을 다시 시작하세요. 
   그래도 안 되면 클라우드 동기화 상태를 확인하세요.

Q: 누가 작업 중인지 확인하려면?
A: 메뉴 → 설정 → 👥 접속 중인 사용자

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

💡 팁

• 작업 전 "접속 중인 사용자" 메뉴로 확인
• 이름을 알아보기 쉽게 설정 (실명 추천)
• 긴급 조회는 읽기 전용으로 가능
• 정기적으로 데이터 백업 권장
• 클라우드 동기화 상태 항상 확인

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

이 기능으로 팀 전체가 효율적으로 재고를 관리할 수 있습니다!
        """
        
        text_widget.insert('1.0', guide_text)
        text_widget.config(state='disabled')
        
        ttk.Button(dialog, text="닫기", command=dialog.destroy).pack(pady=10)
    
    def show_user_manual(self):
        """통합 사용 설명서"""
        self.show_help_dialog("📚 재고관리 시스템 사용 설명서", """
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

📚 재고관리 시스템 사용 설명서

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

🎯 프로그램 개요

이 재고관리 시스템은 의류, 패션, 소매업 등에서
상품의 발주, 입고, 출고, 재고를 통합 관리하는
전문 프로그램입니다.

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

📑 주요 기능

1. 📦 상품 관리
   • 상품 등록 및 수정
   • 색상/사이즈별 재고 관리
   • 상품 이미지 첨부
   • 상품 검색 및 필터링

2. 🚚 발주 관리
   • 발주 등록 및 수정
   • 발주 상태 추적 (대기중/부분출고/완료)
   • 미입고 수량 자동 계산
   • 엑셀 발주장 출력

3. 📥 입고 관리
   • 입고 등록 (수동/엑셀)
   • 자동 미입고 수량 감소
   • 날짜별 입고 내역
   • 엑셀 입고장 불러오기

4. 📤 출고 관리
   • 출고 등록 (수동/엑셀)
   • 날짜별 출고 내역
   • 엑셀 출고장 불러오기

5. 📊 재고 현황
   • 실시간 재고 조회
   • 색상/사이즈별 현황
   • 총입고/총출고/현재고
   • 미입고 수량 표시

6. ☁️ 클라우드 연동
   • 자동 백업 및 동기화
   • 여러 컴퓨터에서 사용
   • OneDrive, Google Drive 등 지원

7. 👥 다중 사용자
   • 팀원과 데이터 공유
   • 자동 충돌 방지
   • 사용자별 작업 기록

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

🚀 빠른 시작

1단계: 상품 등록
   상품 관리 탭 → ➕ 상품 추가 → 정보 입력

2단계: 발주 등록
   발주 관리 탭 → ➕ 발주 추가 → 상품/수량 선택

3단계: 입고 처리
   입고 관리 탭 → 📥 입고 등록 → 상품/수량 입력

4단계: 재고 확인
   재고 현황 탭 → 실시간 재고 조회

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

💡 편리한 기능

• 검색: 상품명이나 코드로 빠른 검색
• 더블클릭: 상품 상세 정보 보기
• 우클릭: 수정/삭제 메뉴
• DEL 키: 선택 항목 빠른 삭제
• 엑셀 연동: 대량 데이터 처리

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

📌 각 기능의 상세 가이드는 
   도움말 메뉴에서 확인하세요!

• 📦 상품 관리 가이드
• 🚚 발주 관리 가이드
• 📥 입고 처리 가이드
• 📤 출고 처리 가이드
• ☁️ 클라우드 연동 가이드
• 👥 다중 사용자 가이드

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
        """)
    
    def show_product_guide(self):
        """상품 관리 가이드"""
        self.show_help_dialog("📦 상품 관리 가이드", """
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

📦 상품 관리 가이드

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

🎯 상품 관리란?

판매하는 모든 상품의 기본 정보를 등록하고
관리하는 기능입니다.

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

➕ 상품 추가하기

1. 상품 관리 탭 클릭
2. "➕ 상품 추가" 버튼 클릭
3. 상품 정보 입력:

   📝 상품명: 상품의 이름 (필수)
      예: "반팔 티셔츠 A형"
   
   🏷️ 상품코드: 고유 코드 (선택)
      예: "TS-001", "SKU-1234"
   
   🏢 매입처: 공급업체 이름 (선택)
      예: "○○상사", "□□도매"
   
   🎨 색상: 쉼표로 구분하여 입력
      예: "빨강, 파랑, 검정, 흰색"
      없으면: 빈칸 또는 "-"
   
   📏 사이즈: 쉼표로 구분하여 입력
      예: "S, M, L, XL"
      없으면: "FREE" (기본값)
   
   📷 이미지: 상품 사진 첨부 (선택)
      클릭하여 파일 선택

4. "저장" 버튼 클릭
5. 완료!

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

✏️ 상품 수정하기

방법 1: 우클릭 메뉴
   1. 수정할 상품 선택
   2. 우클릭 → "수정" 선택
   3. 정보 변경 후 저장

방법 2: 더블클릭
   1. 상품 더블클릭 → 상세보기
   2. 정보 확인

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

🗑️ 상품 삭제하기

⚠️ 중요: 색상/사이즈 조합별로 삭제됩니다!

예시:
"티셔츠 - 빨강 - M" 선택 후 삭제
→ 빨강색 M사이즈만 삭제
→ 다른 색상/사이즈는 유지

전체 상품 삭제:
모든 색상/사이즈를 하나씩 삭제하거나
마지막 조합 삭제 시 상품 전체 삭제 확인

삭제 방법:
• 우클릭 → "삭제"
• DEL 키 누르기

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

🔍 상품 검색하기

1. 상단 검색창에 입력
2. 실시간으로 필터링됨
3. 검색 대상:
   • 상품명
   • 상품코드

검색 예시:
"티셔츠" 입력 → 티셔츠 관련 상품만 표시
"TS-001" 입력 → 해당 코드 상품만 표시

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

📊 테이블 정보 보기

컬럼 설명:
• 상품명: 상품의 이름
• 상품코드: 등록된 코드
• 색상: 색상 옵션
• 사이즈: 사이즈 옵션
• 현재고: 현재 보유 수량
• 미입고: 발주했지만 아직 안 들어온 수량

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

🖼️ 상품 이미지 보기

• 상품 선택 시 왼쪽에 이미지 표시
• 이미지 크기: 자동 조정 (썸네일)
• 지원 형식: JPG, PNG, GIF

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

💡 활용 팁

1. 상품코드 체계 만들기
   예: TS-빨강-M (티셔츠-색상-사이즈)
   → 일관된 규칙으로 관리 편리

2. 이미지 준비
   • 정사각형 이미지 권장
   • 파일 크기: 1MB 이하 권장
   • 배경 제거 시 더 깔끔

3. 색상/사이즈 미리 정의
   • 회사 표준 색상 리스트 작성
   • 표준 사이즈표 작성
   → 입력 실수 방지

4. 매입처 정보 활용
   • 공급업체별로 정리 가능
   • 발주 시 참고 자료

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

⚠️ 주의사항

• 삭제한 상품은 복구 불가
• 삭제 전 재고 확인 필수
• 발주 기록이 있는 상품 삭제 시 
  모든 관련 기록도 삭제됨

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
        """)
    
    def show_order_guide(self):
        """발주 관리 가이드"""
        self.show_help_dialog("🚚 발주 관리 가이드", """
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

🚚 발주 관리 가이드

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

🎯 발주 관리란?

공급업체에 상품을 주문하는 것을 기록하고
입고 상태를 추적하는 기능입니다.

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

➕ 발주 추가하기

1. 발주 관리 탭 클릭
2. "➕ 발주 추가" 버튼 클릭
3. 발주 정보 입력:

   📅 발주 날짜: 달력에서 선택
      (오늘 날짜가 기본값)
   
   📦 상품 선택: 
      • 드롭다운에서 선택
      • 또는 텍스트 입력하여 검색
      예: "티셔츠" 입력 → 자동 필터링
   
   🎨 색상 선택:
      상품 선택 시 자동으로 옵션 표시
   
   📏 사이즈 선택:
      상품 선택 시 자동으로 옵션 표시
   
   📊 발주 수량:
      숫자로 입력 (양수만 가능)

4. "저장" 버튼 클릭
5. 완료!

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

✏️ 발주 수정하기

1. 수정할 발주 우클릭
2. "수정" 선택
3. 변경 가능 항목:
   • 발주 날짜
   • 발주 수량
4. 저장

⚠️ 주의: 상품/색상/사이즈는 수정 불가
         (새로 발주 등록 필요)

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

🗑️ 발주 삭제하기

방법 1: 우클릭 → "삭제"
방법 2: DEL 키 누르기

⚠️ 발주 삭제 시:
   • 발주 기록만 삭제
   • 실제 재고는 변경 없음
   • 입고 기록은 유지됨

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

📊 발주 상태 이해하기

발주 상태는 3가지:

1. 🕐 대기중
   • 발주했지만 아직 입고 안 됨
   • 출고수량 = 0

2. 📦 부분출고
   • 일부만 입고됨
   • 0 < 출고수량 < 발주수량

3. ✅ 완료
   • 전부 입고됨
   • 출고수량 = 발주수량

상태는 자동으로 변경됩니다!
(입고 처리 시 자동 업데이트)

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

📅 날짜별 조회

1. 상단 "날짜 선택" 드롭다운 클릭
2. 날짜 선택
   • "전체": 모든 발주 보기
   • 특정 날짜: 해당 날짜만 보기

💡 팁: 최근 날짜가 위쪽에 표시됩니다

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

📄 발주장 엑셀 출력

1. "📄 발주장 출력 (Excel)" 버튼 클릭
2. 저장 위치 선택
3. 파일명 입력 (자동 생성됨)
4. 저장

출력 내용:
• 발주일자, 상품명, 상품코드
• 색상, 사이즈
• 발주수량, 출고수량, 미입고수량
• 상태

💡 용도:
• 공급업체 발주서로 사용
• 인쇄하여 보관
• 이메일로 전송

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

📊 테이블 정보 보기

컬럼 설명:
• 발주일자: 발주한 날짜
• 상품명: 발주 상품
• 상품코드: 상품 코드
• 색상: 발주한 색상
• 사이즈: 발주한 사이즈
• 발주수량: 주문한 총 수량
• 출고수량: 실제 입고된 수량
• 미입고수량: 아직 안 들어온 수량
• 상태: 대기중/부분출고/완료

⚠️ 같은 날짜의 같은 상품/색상/사이즈는
    자동으로 합쳐서 표시됩니다!

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

💡 활용 팁

1. 정기 발주 활용
   • 매주/매월 같은 상품 발주
   • 이전 발주 기록 참고

2. 발주서 보관
   • 엑셀 파일로 출력
   • 날짜별로 폴더 정리

3. 미입고 수량 체크
   • 정기적으로 확인
   • 지연 시 공급업체 연락

4. 발주 패턴 분석
   • 어떤 상품을 자주 발주하는지
   • 발주 주기 파악

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

🔄 발주 → 입고 흐름

1. 발주 등록
   ↓
2. 공급업체에 발주서 전송
   ↓
3. 상품 도착 시 입고 처리
   ↓
4. 미입고 수량 자동 감소
   ↓
5. 발주 상태 자동 업데이트
   (대기중 → 부분출고 → 완료)

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

⚠️ 주의사항

• 발주는 미래 날짜로도 등록 가능
• 과거 날짜로 소급 등록도 가능
• 발주 수량은 나중에 수정 가능
• 입고 처리는 입고 관리 탭에서!

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
        """)
    
    def show_inbound_guide(self):
        """입고 처리 가이드"""
        self.show_help_dialog("📥 입고 처리 가이드", """
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

📥 입고 처리 가이드

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

🎯 입고 처리란?

발주한 상품이 실제로 들어왔을 때 
입고 수량을 기록하는 기능입니다.

입고 처리하면:
✅ 재고 자동 증가
✅ 미입고 수량 자동 감소
✅ 발주 상태 자동 업데이트

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

➕ 입고 등록하기 (수동)

1. 입고 관리 탭 클릭
2. "📥 입고 등록" 버튼 클릭
3. 입고 정보 입력:

   📅 입고 날짜: 
      실제 상품이 들어온 날짜
      (기본값: 오늘)
   
   📦 상품 검색:
      • 드롭다운에서 선택
      • 텍스트 입력하여 검색
      예: "티셔츠" → 자동 필터링
   
   🎨 색상: 자동 로드됨
   📏 사이즈: 자동 로드됨
   
   📊 입고 수량:
      실제 들어온 수량 입력
   
   📝 비고: 
      특이사항 기록 (선택)
      예: "일부 불량", "박스 파손"

4. "저장" 버튼 클릭
5. 완료!

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

📄 입고장 엑셀로 불러오기

대량의 입고 처리를 한번에!

1. 엑셀 입고장 준비
   필수 컬럼 순서:
   ┌──────────────────────────┐
   │ 모델명 │ 컬러 │ 사이즈 │ 입고수량 │
   ├──────────────────────────┤
   │ 티셔츠A│ 빨강 │  M   │   50   │
   │ 티셔츠A│ 파랑 │  L   │   30   │
   │ 바지B │ 검정 │ FREE │   20   │
   └──────────────────────────┘

2. 입고 관리 탭에서
   "📄 엑셀 파일 불러오기" 클릭

3. 준비한 엑셀 파일 선택

4. 자동으로 입고 처리!

💡 엑셀 작성 팁:
• 첫 행은 헤더 (자동 무시됨)
• 모델명은 등록된 상품명과 정확히 일치
• 빈 행은 자동 스킵
• 날짜는 오늘 날짜 자동 적용

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

🗑️ 입고 삭제하기

잘못 입력한 경우:

1. 삭제할 입고 선택
2. DEL 키 누르기
3. 확인

⚠️ 입고 삭제 시:
   • 해당 입고 기록 삭제
   • 재고 자동 감소
   • 미입고 수량 다시 증가

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

📅 날짜별 조회

1. "날짜 필터" 드롭다운 선택
2. 원하는 날짜 선택
   • "전체": 모든 입고 기록
   • 특정 날짜: 해당 날짜만

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

📊 테이블 정보 보기

컬럼 설명:
• 입고일자: 입고 처리한 날짜
• 상품명: 입고된 상품
• 상품코드: 상품 코드
• 색상: 입고된 색상
• 사이즈: 입고된 사이즈
• 입고수량: 들어온 수량
• 비고: 특이사항

⚠️ 같은 날짜의 같은 상품/색상/사이즈는
    자동으로 합쳐서 표시됩니다!

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

🔄 자동 발주 매칭

입고 처리하면 자동으로:

1. 해당 상품/색상/사이즈의 발주 검색
2. 가장 오래된 발주부터 순차 매칭
3. 발주의 출고수량 자동 증가
4. 미입고수량 자동 감소
5. 발주 상태 자동 업데이트

예시:
┌─────────────────────────────┐
│ 발주 상태                      │
├─────────────────────────────┤
│ 발주: 티셔츠-빨강-M 100개      │
│ 현재: 출고 50개, 미입고 50개   │
├─────────────────────────────┤
│ ↓ 입고 처리: 30개            │
├─────────────────────────────┤
│ 결과: 출고 80개, 미입고 20개   │
│ 상태: 대기중 → 부분출고        │
└─────────────────────────────┘

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

💡 활용 팁

1. 실시간 입고 처리
   • 상품 도착하면 즉시 입고 처리
   • 재고 항상 최신 상태 유지

2. 검수 후 입고
   • 불량품 확인 후 입고 처리
   • 비고란에 불량 수량 기록

3. 분할 입고
   • 발주 100개를 50개씩 2번 입고
   • 각각 따로 입고 처리
   • 자동으로 발주에 누적

4. 엑셀 활용
   • 공급업체 납품서를 엑셀로 정리
   • 한번에 대량 입고 처리

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

⚠️ 주의사항

• 입고는 발주 없이도 가능
  (직접 구매, 반품 입고 등)

• 입고 수량 > 발주 수량도 가능
  (추가 서비스, 덤 등)

• 과거 날짜로 소급 입고 가능
  (나중에 발견한 입고 처리)

• 입고 삭제는 신중하게!
  (재고가 감소함)

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

📋 입고 업무 흐름 예시

오전:
1. 공급업체에서 상품 도착
2. 납품서 확인
3. 검수 (수량, 품질 체크)

오후:
4. 프로그램에서 입고 처리
   • 수동 입력 또는
   • 엑셀 일괄 처리
5. 창고 정리
6. 재고 현황 확인

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
        """)
    
    def show_outbound_guide(self):
        """출고 처리 가이드"""
        self.show_help_dialog("📤 출고 처리 가이드", """
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

📤 출고 처리 가이드

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

🎯 출고 처리란?

판매, 배송, 샘플 제공 등으로 
상품이 나갔을 때 기록하는 기능입니다.

출고 처리하면:
✅ 재고 자동 감소
✅ 출고 이력 기록
✅ 실시간 재고 업데이트

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

➕ 출고 등록하기 (수동)

1. 출고 관리 탭 클릭
2. "📤 출고 등록" 버튼 클릭
3. 출고 정보 입력:

   📅 출고 날짜:
      실제 상품이 나간 날짜
      (기본값: 오늘)
   
   📦 상품 검색:
      • 드롭다운에서 선택
      • 텍스트 입력하여 검색
      예: "티셔츠" → 자동 필터링
   
   🎨 색상: 자동 로드됨
   📏 사이즈: 자동 로드됨
   
   📊 출고 수량:
      실제 나간 수량 입력
   
   📝 비고:
      출고 사유 기록 (선택)
      예: "온라인 주문", "매장 판매"
          "샘플 제공", "불량 폐기"

4. "저장" 버튼 클릭
5. 완료!

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

📄 출고장 엑셀로 불러오기

대량의 출고 처리를 한번에!

1. 엑셀 출고장 준비
   필수 컬럼 순서:
   ┌──────────────────────────┐
   │ 모델명 │ 컬러 │ 사이즈 │ 출고수량 │
   ├──────────────────────────┤
   │ 티셔츠A│ 빨강 │  M   │   10   │
   │ 티셔츠A│ 파랑 │  L   │   15   │
   │ 바지B │ 검정 │ FREE │    5   │
   └──────────────────────────┘

2. 출고 관리 탭에서
   "📄 엑셀 파일 불러오기" 클릭

3. 준비한 엑셀 파일 선택

4. 자동으로 출고 처리!

💡 엑셀 작성 팁:
• 첫 행은 헤더 (자동 무시됨)
• 모델명은 등록된 상품명과 정확히 일치
• 날짜는 오늘 날짜 자동 적용
• 온라인 쇼핑몰 주문서를 엑셀로 정리하여 활용

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

🗑️ 출고 삭제하기

잘못 입력한 경우:

1. 삭제할 출고 선택
2. DEL 키 누르기
3. 확인

⚠️ 출고 삭제 시:
   • 해당 출고 기록 삭제
   • 재고 자동 증가 (복구됨)

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

📅 날짜별 조회

1. "날짜 필터" 드롭다운 선택
2. 원하는 날짜 선택
   • "전체": 모든 출고 기록
   • 특정 날짜: 해당 날짜만

💡 용도:
• 일별 판매량 확인
• 월별 출고 통계
• 특정 기간 분석

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

📊 테이블 정보 보기

컬럼 설명:
• 출고일자: 출고 처리한 날짜
• 상품명: 출고된 상품
• 상품코드: 상품 코드
• 색상: 출고된 색상
• 사이즈: 출고된 사이즈
• 출고수량: 나간 수량
• 비고: 출고 사유

⚠️ 같은 날짜의 같은 상품/색상/사이즈는
    자동으로 합쳐서 표시됩니다!

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

💡 활용 팁

1. 실시간 출고 처리
   • 판매 즉시 출고 처리
   • 재고 항상 정확하게 유지

2. 출고 사유 기록
   • 비고란 적극 활용
   • 판매/샘플/불량 구분
   • 나중에 분석 가능

3. 일괄 처리
   • 하루 판매분을 모아서
   • 엑셀로 정리하여 일괄 처리
   • 업무 효율 향상

4. 마감 정산
   • 일일 마감 시 출고 확인
   • 매장 판매 vs 온라인 구분
   • 채널별 분석

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

📋 출고 유형별 처리 방법

🛒 일반 판매
비고: "매장 판매" 또는 "온라인 주문 #123"

🎁 샘플/증정
비고: "샘플 제공 - △△ 바이어"

📦 반품 출고 (고객 → 공급업체)
비고: "불량 반품 - 업체 반송"

❌ 불량 폐기
비고: "불량품 폐기 - 검수팀"

🔄 매장 간 이동
비고: "강남점 → 홍대점 이동"

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

⚠️ 재고 부족 시

출고 수량 > 현재 재고인 경우:

❌ 프로그램이 막지는 않습니다
   (음수 재고 가능)

⚠️ 하지만 주의 필요:
   • 재고 현황에서 음수 표시
   • 실제로는 없는 상품
   • 발주 또는 입고 필요

💡 해결 방법:
   1. 재고 현황 확인
   2. 음수 재고 발견
   3. 즉시 발주 처리
   4. 입고 대기

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

📋 출고 업무 흐름 예시

매장 판매:
1. 고객 결제
2. 상품 포장 및 전달
3. 프로그램에서 출고 처리
4. 재고 확인

온라인 판매:
1. 주문 접수
2. 상품 피킹
3. 포장 및 송장 부착
4. 출고 처리 (주문번호 기록)
5. 배송 업체 인계

일괄 처리:
1. 하루 판매분 수집
2. 엑셀로 정리
3. 일괄 출고 처리
4. 재고 확인 및 발주 검토

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

⚠️ 주의사항

• 출고는 재고를 감소시킵니다
• 실제 판매/출고 시에만 처리
• 잘못된 출고는 즉시 삭제
• 정기적으로 재고 실사 필요
• 음수 재고 발생 시 즉시 확인

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
        """)
    
    def show_cloud_guide(self):
        """클라우드 연동 가이드"""
        self.show_help_dialog("☁️ 클라우드 연동 가이드", r"""
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

☁️ 클라우드 연동 가이드

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

🎯 클라우드 연동이란?

OneDrive, Google Drive, Dropbox 등의
클라우드 스토리지에 데이터를 저장하여:

✅ 자동 백업
✅ 여러 컴퓨터에서 사용
✅ 팀원과 데이터 공유
✅ 데이터 안전 보관

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

☁️ 지원하는 클라우드

✅ OneDrive (원드라이브)
   • Windows 기본 내장
   • 무료 5GB

✅ Google Drive (구글 드라이브)
   • 무료 15GB (가장 큼)
   • Gmail 계정 필요

✅ Dropbox (드롭박스)
   • 가장 빠른 동기화
   • 무료 2GB

✅ Naver Cloud (네이버 클라우드)
   • 한국 서버 (빠름)
   • 무료 30GB

✅ iCloud Drive (아이클라우드)
   • Mac/iPhone 연동
   • 무료 5GB

✅ 기타 모든 클라우드
   • 로컬 동기화 폴더가 있으면 모두 가능

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

🚀 설정 방법

1단계: 클라우드 앱 설치
   • OneDrive: Windows 기본 설치됨
   • Google Drive: drive.google.com/download
   • Dropbox: dropbox.com/install
   • 설치 후 계정 로그인

2단계: 프로그램 설정
   1. 메뉴 → 설정 → ☁️ 클라우드 스토리지 설정
   2. 방법 선택:

   방법A: 자동 찾기 (추천!)
      "🔍 자동으로 클라우드 찾기" 클릭
      → 발견된 클라우드 확인
      → "예" 클릭

   방법B: 수동 선택
      "📁 클라우드 폴더 직접 선택" 클릭
      → 클라우드 폴더 찾기
      → 선택

3단계: 완료!
   창 제목에 클라우드 이름 표시됨
   예: "재고 관리 시스템 - Google Drive"

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

🔄 여러 컴퓨터에서 사용하기

시나리오: 사무실과 집에서 사용

사무실 컴퓨터:
1. 클라우드 앱 설치 및 로그인
2. 재고관리 프로그램 설치
3. 클라우드 폴더 선택
4. 작업 수행 → 자동 저장

집 컴퓨터:
1. 동일한 클라우드 앱 설치
2. 같은 계정으로 로그인
3. 동기화 완료 대기
4. 재고관리 프로그램 설치
5. 동일한 클라우드 폴더 선택
6. 사무실 데이터가 그대로!

⚠️ 주의: 같은 계정 사용 필수!

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

👥 팀원과 공유하기

1. 관리자: 클라우드에 공유 폴더 생성
   예: "회사재고관리" 폴더

2. 관리자: 팀원 초대
   폴더 우클릭 → 공유 → 이메일 추가
   권한: "편집 가능"

3. 팀원: 초대 수락
   이메일 확인 → 초대 수락
   클라우드 앱에서 공유 폴더 확인

4. 팀원: 프로그램 설정
   공유 폴더 선택

5. 완료!
   모든 팀원이 동일한 데이터 사용

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

💾 데이터 저장 위치

클라우드 사용 시:
[클라우드 폴더]/inventory_data.json

예시:
• OneDrive:
  C:/Users/홍길동/OneDrive/inventory_data.json

• Google Drive:
  C:/Users/홍길동/Google Drive/inventory_data.json

• Dropbox:
  C:/Users/홍길동/Dropbox/inventory_data.json

로컬 저장:
프로그램이 설치된 폴더/inventory_data.json

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

🔄 클라우드 전환하기

다른 클라우드로 변경 가능:

1. 메뉴 → 설정 → ☁️ 클라우드 스토리지 설정
2. 새 클라우드 선택
3. 데이터 자동 복사됨
4. 완료!

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

💡 활용 팁

1. 정기 백업도 함께
   • 클라우드 + 수동 백업
   • 메뉴 → 설정 → 📊 데이터 백업
   • USB/외장하드에 추가 저장

2. 공유 폴더 전용 사용
   • 다른 파일과 섞이지 않게
   • "재고관리 전용" 폴더 생성

3. 동기화 상태 확인
   • 작업 표시줄에서 클라우드 아이콘 확인
   • "동기화 중" 또는 "최신 상태"

4. 정기적 점검
   • 주 1회 동기화 상태 확인
   • 클라우드 저장 공간 확인

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

🔧 문제 해결

Q: 동기화가 안 돼요
A: 1. 인터넷 연결 확인
   2. 클라우드 앱 실행 확인
   3. 일시정지 해제
   4. 클라우드 앱 재시작

Q: 저장 공간이 부족해요
A: 1. 이미지 파일 최소화
   2. 오래된 백업 삭제
   3. 클라우드 용량 확장
   4. 다른 클라우드로 전환

Q: 다른 컴퓨터에서 데이터가 안 보여요
A: 1. 같은 계정인지 확인
   2. 동기화 완료 대기
   3. 클라우드 폴더 경로 확인
   4. 프로그램 재시작

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

⚠️ 주의사항

• 인터넷 연결 필요 (동기화 시)
• 오프라인에서도 작업 가능
• 동기화는 인터넷 연결 시 자동
• 공용 컴퓨터 사용 시 로그아웃 필수
• 중요 데이터는 추가 백업 권장

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
        """)
    
    def show_help_dialog(self, title, content):
        """도움말 다이얼로그 표시"""
        dialog = tk.Toplevel(self.root)
        dialog.title(title)
        dialog.geometry("800x700")
        dialog.transient(self.root)
        
        # 제목
        title_label = ttk.Label(dialog, text=title, font=("Arial", 14, "bold"))
        title_label.pack(pady=10)
        
        # 스크롤 가능한 텍스트
        text_frame = ttk.Frame(dialog)
        text_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)
        
        scrollbar = ttk.Scrollbar(text_frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        text_widget = tk.Text(text_frame, wrap=tk.WORD, yscrollcommand=scrollbar.set, 
                             font=("맑은 고딕", 10), padx=10, pady=10)
        text_widget.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.config(command=text_widget.yview)
        
        text_widget.insert('1.0', content)
        text_widget.config(state='disabled')
        
        # 닫기 버튼
        ttk.Button(dialog, text="닫기", command=dialog.destroy).pack(pady=10)
    
    def show_faq(self):
        """자주 묻는 질문"""
        self.show_help_dialog("❓ 자주 묻는 질문 (FAQ)", """
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

❓ 자주 묻는 질문 (FAQ)

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

💾 데이터 관련

Q: 데이터는 어디에 저장되나요?
A: 클라우드 사용 시: 클라우드 폴더
   로컬 사용 시: 프로그램 폴더
   파일명: inventory_data.json

Q: 데이터를 백업하려면?
A: 메뉴 → 설정 → 📊 데이터 백업
   USB, 외장하드에 저장 권장

Q: 잘못 삭제한 데이터 복구 가능한가요?
A: 1. 백업 파일로 복원
   2. 클라우드 버전 기록으로 복원
   3. 클라우드 휴지통 확인

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

📦 상품 관련

Q: 색상/사이즈가 없는 상품은?
A: 색상: 빈칸 또는 "-"
   사이즈: "FREE" (기본값)

Q: 상품을 잘못 삭제했어요
A: 1. 데이터 복원 기능 사용
   2. 또는 다시 등록

Q: 이미지가 안 보여요
A: 1. 이미지 파일 형식 확인 (JPG, PNG)
   2. 파일 크기 확인 (10MB 이하)
   3. 프로그램 재시작

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

🚚 발주/입고 관련

Q: 발주 없이 입고 가능한가요?
A: 네! 가능합니다.
   직접 구매, 반품 입고 등에 사용

Q: 입고 수량이 발주보다 많을 수 있나요?
A: 네! 가능합니다.
   추가 서비스, 덤 등으로 더 올 수 있음

Q: 미입고 수량이 자동으로 안 줄어요
A: 1. 상품/색상/사이즈 정확히 일치 확인
   2. 프로그램 재시작
   3. 발주 상태 확인

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

☁️ 클라우드 관련

Q: 어떤 클라우드를 써야 하나요?
A: 이미 사용 중인 것 추천
   • Windows 사용자 → OneDrive
   • Gmail 사용자 → Google Drive
   • 큰 용량 필요 → Google Drive (15GB)

Q: 클라우드 없이 사용 가능한가요?
A: 네! 로컬 저장 가능
   단, 백업은 수동으로 필요

Q: 동기화가 느려요
A: 1. 인터넷 속도 확인
   2. 다른 파일 동기화 중인지 확인
   3. 클라우드 앱 재시작

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

👥 다중 사용자 관련

Q: 동시에 작업 가능한가요?
A: 한 명만 편집 가능
   다른 사람은 읽기 전용
   (자동 충돌 방지)

Q: 다른 사람이 작업 중인지 확인하려면?
A: 메뉴 → 설정 → 👥 접속 중인 사용자

Q: 작업 중인데 다른 사람이 수정해요
A: 프로그램 재시작 필요
   잠금 시스템 오류일 수 있음

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

📊 재고 관련

Q: 재고가 음수가 됐어요
A: 출고 > 입고인 경우 발생
   1. 재고 현황 확인
   2. 입고 누락 확인
   3. 출고 오류 확인
   4. 조정 필요 시 입고 처리

Q: 실제 재고와 프로그램이 달라요
A: 1. 정기 재고 실사
   2. 차이나는 수량 조정
   3. 입고/출고 기록 재확인

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

🔧 기술 문제

Q: 프로그램이 느려요
A: 1. 데이터 정리 (오래된 기록 백업 후 삭제)
   2. 이미지 최소화
   3. 컴퓨터 재시작

Q: 저장이 안 돼요
A: 1. 다른 사람 작업 중인지 확인
   2. 디스크 공간 확인
   3. 권한 확인
   4. 프로그램 재시작

Q: 엑셀 불러오기가 안 돼요
A: 1. 엑셀 형식 확인 (.xlsx)
   2. 컬럼 순서 확인
   3. 상품명 정확히 일치 확인
   4. 빈 행 제거

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

💡 기타

Q: 여러 매장을 관리할 수 있나요?
A: 비고란에 매장명 기록하여 구분 가능
   또는 매장별로 폴더 분리

Q: 바코드 스캐너 사용 가능한가요?
A: 상품코드에 바코드 번호 입력
   검색 시 바코드 스캔하여 사용

Q: 인쇄 기능이 있나요?
A: 엑셀 출력 후 인쇄 권장
   또는 화면 캡처하여 인쇄

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

📞 추가 지원

더 궁금한 사항은:
• 도움말 메뉴의 각 가이드 참고
• 프로그램 개발자/관리자에게 문의

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
        """)
    
    def show_about(self):
        """프로그램 정보"""
        dialog = tk.Toplevel(self.root)
        dialog.title("프로그램 정보")
        dialog.geometry("500x500")
        dialog.transient(self.root)
        dialog.grab_set()
        utils.center_window(dialog)
        
        ttk.Label(dialog, text="📦 재고관리 시스템", font=("Arial", 16, "bold")).pack(pady=20)
        ttk.Label(dialog, text="버전 2.0", font=("Arial", 12)).pack(pady=5)
        
        info_frame = ttk.Frame(dialog)
        info_frame.pack(fill=tk.BOTH, expand=True, padx=30, pady=20)
        
        info_text = f"""
━━━━━━━━━━━━━━━━━━━━━━

프로그램 정보

━━━━━━━━━━━━━━━━━━━━━

📌 버전: 2.0
📅 업데이트: 2024
🖥️ 플랫폼: Windows, Mac, Linux

━━━━━━━━━━━━━━━━━━━━━

주요 기능

✅ 상품/발주/입고/출고 관리
✅ 실시간 재고 현황
✅ 클라우드 동기화
✅ 다중 사용자 지원
✅ 엑셀 연동

━━━━━━━━━━━━━━━━━━━━━

현재 사용자

👤 {self.user_display_name}
💻 {socket.gethostname()}

━━━━━━━━━━━━━━━━━━━━━

저장 위치

{self.cloud_type if self.cloud_path else '로컬'}
{self.data_file}

━━━━━━━━━━━━━━━━━━━━━

Made with ❤️ for inventory management
        """
        
        ttk.Label(info_frame, text=info_text, justify=tk.LEFT).pack()
        
        ttk.Button(dialog, text="닫기", command=dialog.destroy).pack(pady=20)
        """창 제목 업데이트"""
        title = "재고 관리 시스템"
        
        if self.cloud_path:
            title += f" - {self.cloud_type}"
        
        title += f" ({self.user_display_name})"
        
        if self.is_locked:
            title += " - 작업 중"
        
        self.root.title(title)
        
    def update_title(self):
        """프로그램 제목 업데이트"""
        users_config = self.data_manager.load_users_config()
        user_display_name = users_config.get(self.current_user, {}).get('display_name', '사용자')
        
        if self.cloud_type == 'api':
            # API 기반 클라우드
            try:
                config = json.load(open(self.config_file, 'r', encoding='utf-8')) if os.path.exists(self.config_file) else {}
                api_service = config.get('api_service', 'API')
                title = f"재고 관리 시스템 - [{user_display_name}] - {api_service}"
            except:
                title = f"재고 관리 시스템 - [{user_display_name}] - API 클라우드"
        elif self.cloud_path and os.path.exists(self.cloud_path):
            # 클라우드 폴더 기반
            title = f"재고 관리 시스템 - [{user_display_name}] - {self.cloud_type}"
        else:
            # 로컬
            title = f"재고 관리 시스템 - [{user_display_name}] - 로컬"
        
        self.root.title(title)
    
    def create_ui(self):
        # 메뉴바 생성
        menubar = tk.Menu(self.root)
        self.root.config(menu=menubar)
        
        # 관리 메뉴
        manage_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="관리", menu=manage_menu)
        manage_menu.add_command(label="상품 관리", command=lambda: self.notebook.select(0))
        manage_menu.add_command(label="발주 관리", command=lambda: self.notebook.select(1))
        manage_menu.add_command(label="입고 관리", command=lambda: self.notebook.select(2))
        manage_menu.add_command(label="출고 관리", command=lambda: self.notebook.select(3))
        manage_menu.add_command(label="재고 현황", command=lambda: self.notebook.select(4))
        
        # 출고/정산 관리 메뉴
        settlement_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="💰 출고/정산 관리", menu=settlement_menu)
        settlement_menu.add_command(label="📤 출고장 업로드 (날짜 지정)", command=self.upload_settlement_sheet)
        settlement_menu.add_command(label="💵 잔액 확인/수정", command=self.manage_settlement_balances)

        # 설정 메뉴
        settings_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="⚙️ 설정", menu=settings_menu)
        settings_menu.add_command(label="☁️ 클라우드 스토리지 설정", command=self.change_cloud_path)
        settings_menu.add_separator()
        settings_menu.add_command(label="📝 필드명 변경", command=self.change_field_names)
        settings_menu.add_separator()
        settings_menu.add_command(label="👤 사용자 정보", command=self.show_user_info)
        settings_menu.add_command(label="👥 접속 중인 사용자", command=self.show_active_users)
        settings_menu.add_separator()
        settings_menu.add_command(label="📊 데이터 백업", command=self.backup_data)
        settings_menu.add_command(label="📥 데이터 복원", command=self.restore_data)

        # 도움말 메뉴
        help_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="도움말", menu=help_menu)
        help_menu.add_command(label="📚 사용 설명서", command=self.show_user_manual)
        help_menu.add_separator()
        help_menu.add_command(label="📦 상품 관리 가이드", command=self.show_product_guide)
        help_menu.add_command(label="🚚 발주 관리 가이드", command=self.show_order_guide)
        help_menu.add_command(label="📥 입고 처리 가이드", command=self.show_inbound_guide)
        help_menu.add_command(label="📤 출고 처리 가이드", command=self.show_outbound_guide)
        help_menu.add_separator()
        help_menu.add_command(label="☁️ 클라우드 연동 가이드", command=self.show_cloud_guide)
        help_menu.add_command(label="👥 다중 사용자 가이드", command=self.show_multiuser_guide)
        help_menu.add_separator()
        help_menu.add_command(label="❓ 자주 묻는 질문 (FAQ)", command=self.show_faq)
        help_menu.add_command(label="ℹ️ 프로그램 정보", command=self.show_about)
        
        # 제목 업데이트
        self.update_title()
        
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        self.products_frame = ttk.Frame(self.notebook)
        self.orders_frame = ttk.Frame(self.notebook)
        self.inbound_frame = ttk.Frame(self.notebook)
        self.outbound_frame = ttk.Frame(self.notebook)
        self.stock_frame = ttk.Frame(self.notebook)
        self.stores_frame = ttk.Frame(self.notebook)
        
        self.notebook.add(self.products_frame, text="📦 상품 관리")
        self.notebook.add(self.orders_frame, text="🚚 발주 관리")
        self.notebook.add(self.inbound_frame, text="📥 입고 관리")
        self.notebook.add(self.outbound_frame, text="📤 출고 관리")
        self.notebook.add(self.stock_frame, text="📊 재고 현황")
        self.notebook.add(self.stores_frame, text="🏪 매장 관리")
        
        self.create_products_tab()
        self.create_orders_tab()
        self.create_inbound_tab()
        self.create_outbound_tab()
        self.create_stock_tab()
        self.create_stores_tab()
        
        # 하단 저작권 표시
        copyright_frame = ttk.Frame(self.root)
        copyright_frame.pack(side=tk.BOTTOM, fill=tk.X, pady=5)
        copyright_label = ttk.Label(copyright_frame, text="Made By EACH Inc.", 
                                   font=("Arial", 9), foreground="gray")
        copyright_label.pack()
        
    def backup_data(self):
        """데이터 백업"""
        if not os.path.exists(self.data_file):
            messagebox.showwarning("경고", "백업할 데이터가 없습니다.")
            return
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        default_filename = f"재고데이터_백업_{timestamp}.json"
        
        filename = filedialog.asksaveasfilename(
            title="데이터 백업",
            defaultextension=".json",
            filetypes=[("JSON files", "*.json"), ("All files", "*.*")],
            initialfile=default_filename
        )
        
        if filename:
            try:
                import shutil
                shutil.copy2(self.data_file, filename)
                messagebox.showinfo("완료", f"데이터가 백업되었습니다:\n{filename}")
            except Exception as e:
                messagebox.showerror("오류", f"백업 중 오류가 발생했습니다:\n{str(e)}")
    
    def restore_data(self):
        """데이터 복원"""
        if messagebox.askyesno("확인", "현재 데이터를 덮어쓰시겠습니까?\n이 작업은 되돌릴 수 없습니다."):
            filename = filedialog.askopenfilename(
                title="데이터 복원",
                filetypes=[("JSON files", "*.json"), ("All files", "*.*")]
            )
            
            if filename:
                try:
                    import shutil
                    # 현재 데이터 임시 백업
                    if os.path.exists(self.data_file):
                        temp_backup = self.data_file + ".temp_backup"
                        shutil.copy2(self.data_file, temp_backup)
                    
                    # 복원
                    shutil.copy2(filename, self.data_file)
                    
                    # 데이터 다시 로드
                    self.data_manager.load_data()
                    self.refresh_all()
                    
                    # 임시 백업 삭제
                    if os.path.exists(temp_backup):
                        os.remove(temp_backup)
                    
                    messagebox.showinfo("완료", "데이터가 복원되었습니다.")
                except Exception as e:
                    messagebox.showerror("오류", f"복원 중 오류가 발생했습니다:\n{str(e)}")
                    # 오류 발생 시 임시 백업 복원
                    if os.path.exists(temp_backup):
                        try:
                            shutil.copy2(temp_backup, self.data_file)
                            os.remove(temp_backup)
                        except:
                            pass
    
    def refresh_all(self):
        """모든 목록 새로고침"""
        self.refresh_products_list()
        self.refresh_orders_list()
        self.refresh_inbound_list()
        self.refresh_outbound_list()
        self.refresh_stock_list()
        self.update_inbound_date_filter()
        self.update_outbound_date_filter()
    
    def manual_save(self):
        """수동 저장 - DB 모드에서는 동기화 역할"""
        try:
            # 캐시 무효화 후 DB에서 최신 데이터 로드
            if hasattr(self.data_manager, 'invalidate_all_cache'):
                self.data_manager.invalidate_all_cache()
            self._refresh_data_shortcuts()
            self.refresh_products_list()
            self.refresh_orders_list()
            self.refresh_stock_list()
            messagebox.showinfo("동기화 완료", "✅ DB와 동기화되었습니다.")
        except Exception as e:
            messagebox.showerror("동기화 실패", f"동기화 중 오류가 발생했습니다:\n{str(e)}")
    
    def on_products_tree_double_click(self, event):
        """상품관리탭 더블클릭 시 셀 편집"""
        region = self.products_tree.identify("region", event.x, event.y)
        if region != "cell":
            return
        
        column = self.products_tree.identify_column(event.x)
        col_index = int(column.replace('#', '')) - 1
        columns = self.products_tree['columns']
        col_name = columns[col_index]
        
        item_id = self.products_tree.identify_row(event.y)
        if not item_id:
            return
        
        # 편집 가능한 열만 허용
        editable_columns = ['상품코드', '상품명', '매입처']
        if col_name not in editable_columns:
            self.show_product_detail(event)
            return
        
        values = self.products_tree.item(item_id)['values']
        current_value = values[col_index]
        
        new_value = simpledialog.askstring(f"{col_name} 수정", 
                                           f"새로운 {col_name}:",
                                           initialvalue=str(current_value))
        if new_value is None:
            return
        
        product_name = values[0]
        product_code = values[1]
        
        for p in self.products:
            if p['name'] == product_name and p.get('code', '') == product_code:
                if col_name == '상품코드':
                    p['code'] = new_value
                elif col_name == '상품명':
                    p['name'] = new_value
                elif col_name == '매입처':
                    p['supplier'] = new_value
                
                # DB에 즉시 저장
                try:
                    self.data_manager.update_product_in_db(p['id'], p)
                    print(f"✅ DB 저장 완료: {p['name']}")
                except Exception as e:
                    print(f"❌ DB 저장 오류: {e}")
                break
        
        self._refresh_data_shortcuts()
        self.refresh_products_list()
        messagebox.showinfo("수정 완료", "✅ 변경사항이 저장되었습니다.")
    
    def on_orders_tree_double_click(self, event):
        """발주관리탭 더블클릭 시 수량 편집"""
        region = self.orders_tree.identify("region", event.x, event.y)
        if region != "cell":
            return
        
        column = self.orders_tree.identify_column(event.x)
        col_index = int(column.replace('#', '')) - 1
        columns = self.orders_tree['columns']
        col_name = columns[col_index]
        
        if col_name not in ['발주수량', '출고수량']:
            return
        
        item_id = self.orders_tree.identify_row(event.y)
        if not item_id:
            return
        
        values = self.orders_tree.item(item_id)['values']
        current_value = values[col_index]
        
        new_value = simpledialog.askinteger(f"{col_name} 수정",
                                            f"새로운 {col_name}:",
                                            initialvalue=int(current_value))
        if new_value is None:
            return
        
        order_date = values[0]
        product_name = values[1]
        color = values[3] if values[3] != '-' else ''
        size = values[4]
        store_name = values[5] if values[5] != '-' else ''
        
        store_id = ''
        for s in self.stores:
            if s['name'] == store_name:
                store_id = s.get('id', '')
                break
        
        product = None
        for p in self.products:
            if p['name'] == product_name:
                product_colors = p.get('colors', [''])
                if not color or color in product_colors:
                    product = p
                    break
        
        if not product:
            return
        
        for order in self.orders:
            if (order['product_id'] == product['id'] and
                order['date'] == order_date and
                order.get('color', '') == color and
                order.get('size', 'FREE') == size and
                order.get('store_id', '') == store_id):
                
                if col_name == '발주수량':
                    order['quantity'] = new_value
                elif col_name == '출고수량':
                    order['shipped_quantity'] = new_value
                
                # DB에 즉시 저장
                try:
                    self.data_manager.update_order(order['id'], order)
                    print(f"✅ 발주 DB 저장 완료: ID={order['id']}")
                except Exception as e:
                    print(f"❌ 발주 DB 저장 오류: {e}")
                break
        
        self._refresh_data_shortcuts()
        self.refresh_orders_list()
        messagebox.showinfo("수정 완료", "✅ 변경사항이 저장되었습니다.")
    
    def refresh_with_merge(self):
        """ID 충돌 검사 및 중복 상품 병합 후 새로고침"""
        # 1. ID 충돌 검사 및 자동 수정 (저장 안 함)
        fixed_id_count = self.data_manager.check_and_fix_duplicate_ids()
        
        # 2. 중복 상품 병합
        merged_count = self.merge_duplicate_products()
        
        # 3. 새로고침
        self.refresh_products_list()
        
        # 4. 결과 메시지
        messages = []
        if fixed_id_count > 0:
            messages.append(f"✅ ID 충돌 수정: {fixed_id_count}개")
        if merged_count > 0:
            messages.append(f"✅ 중복 상품 병합: {merged_count}개")
        
        if messages:
            messagebox.showinfo("새로고침 완료", "\n".join(messages))
        else:
            messagebox.showinfo("새로고침 완료", "✅ ID 충돌 및 중복 상품이 없습니다.")
    
    def merge_duplicate_products(self):
        """상품명, 상품코드, 색상이 모두 일치하는 상품 병합"""
        merged_count = 0
        products_to_remove = []
        
        # 상품을 순회하면서 중복 찾기
        for i in range(len(self.products)):
            if self.products[i] in products_to_remove:
                continue
                
            product_a = self.products[i]
            product_name = product_a['name']
            product_code = product_a.get('code', '')
            colors_a_list = sorted(product_a.get('colors', ['']))
            colors_a_str = ','.join(colors_a_list)
            
            # 같은 상품명, 상품코드를 가진 다른 상품 찾기
            for j in range(i + 1, len(self.products)):
                if self.products[j] in products_to_remove:
                    continue
                    
                product_b = self.products[j]
                colors_b_list = sorted(product_b.get('colors', ['']))
                colors_b_str = ','.join(colors_b_list)
                
                # 상품명, 상품코드, 색상이 모두 일치하는지 확인
                if (product_b['name'] == product_name and 
                    product_b.get('code', '') == product_code and
                    colors_b_str == colors_a_str):
                    
                    # product_b의 사이즈를 product_a에 병합
                    sizes_a = set(product_a.get('sizes', ['FREE']))
                    sizes_b = set(product_b.get('sizes', ['FREE']))
                    product_a['sizes'] = sorted(list(sizes_a.union(sizes_b)))
                    
                    # product_b의 이미지가 있고 product_a에 없으면 가져오기
                    if not product_a.get('image') and product_b.get('image'):
                        product_a['image'] = product_b['image']
                        product_a['image_source'] = product_b.get('image_source', 'manual')
                    
                    # product_b의 발주/재고 이동/입출고 기록을 product_a로 이동
                    product_b_id = product_b['id']
                    product_a_id = product_a['id']
                    
                    for order in self.orders:
                        if order['product_id'] == product_b_id:
                            order['product_id'] = product_a_id
                    
                    for movement in self.movements:
                        if movement['product_id'] == product_b_id:
                            movement['product_id'] = product_a_id
                    
                    for record in self.inbound_records:
                        if record['product_id'] == product_b_id:
                            record['product_id'] = product_a_id
                    
                    for record in self.outbound_records:
                        if record['product_id'] == product_b_id:
                            record['product_id'] = product_a_id
                    
                    # product_b를 삭제 목록에 추가
                    products_to_remove.append(product_b)
                    merged_count += 1
        
        # 중복 상품 삭제
        for product in products_to_remove:
            if product in self.products:
                self.products.remove(product)
        
        # 병합이 있었으면 저장
        if merged_count > 0:
            # self.data_manager.save_data()
            # ← 자동저장 제거됨
            self._refresh_data_shortcuts()
        
        return merged_count
    
    def create_products_tab(self):
        # 상단 제어 패널
        control_frame = ttk.Frame(self.products_frame)
        control_frame.pack(fill=tk.X, padx=10, pady=10)
        
        # 왼쪽: 검색 영역
        search_frame = ttk.LabelFrame(control_frame, text="🔍 상품 검색", padding=5)
        search_frame.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))
        
        self.product_search_var = tk.StringVar()
        
        # Entry 위젯 먼저 생성
        search_entry = ttk.Entry(search_frame, textvariable=self.product_search_var, width=40)
        search_entry.pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)
        
        # Entry 위젯 저장
        self.product_search_entry = search_entry
        
        # 무한 루프 방지 플래그
        self._updating_search = False
        
        # trace를 통한 실시간 검색 - 무한 루프 방지
        def on_search_change(*args):
            if self._updating_search:
                return
            try:
                self._updating_search = True
                self.refresh_products_list()
            except Exception as e:
                print(f"검색 오류: {e}")
            finally:
                self._updating_search = False
        
        self.product_search_var.trace("w", on_search_change)
        
        # 오른쪽: 버튼 영역
        button_frame = ttk.Frame(control_frame)
        button_frame.pack(side=tk.RIGHT, padx=5)
        
        ttk.Button(button_frame, text="➕ 상품 추가", command=self.add_product).pack(side=tk.LEFT, padx=3)
        ttk.Button(button_frame, text="📥 엑셀로 일괄 등록", command=self.import_products_from_excel).pack(side=tk.LEFT, padx=3)
        ttk.Button(button_frame, text="📤 엑셀로 저장", command=self.export_products_to_excel).pack(side=tk.LEFT, padx=3)
        ttk.Button(button_frame, text="✏️ 수정", command=self.edit_product).pack(side=tk.LEFT, padx=3)
        ttk.Button(button_frame, text="🗑️ 삭제", command=self.delete_product).pack(side=tk.LEFT, padx=3)
        ttk.Button(button_frame, text="💾 저장", command=self.manual_save, style="Accent.TButton").pack(side=tk.LEFT, padx=3)
        ttk.Button(button_frame, text="🔄 동기화", command=self.sync_from_db).pack(side=tk.LEFT, padx=3)
        
        # 이미지를 위한 프레임 생성
        main_frame = ttk.Frame(self.products_frame)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # 왼쪽에 이미지 표시 영역
        image_frame = ttk.Frame(main_frame, width=150)
        image_frame.pack(side=tk.LEFT, fill=tk.Y, padx=(0, 10))
        image_frame.pack_propagate(False)
        
        ttk.Label(image_frame, text="상품 이미지", font=("Arial", 10, "bold")).pack(pady=5)
        self.product_image_label = ttk.Label(image_frame, text="이미지 없음", relief=tk.SUNKEN)
        self.product_image_label.pack(pady=5, padx=5, fill=tk.BOTH, expand=True)
        
        # 오른쪽에 테이블
        table_frame = ttk.Frame(main_frame)
        table_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        # 동적 컬럼 생성
        field_columns = tuple([field['name'] for field in self.field_names])
        columns = ("상품명", "상품코드") + field_columns + ("현재고", "미입고")
        self.products_tree = ttk.Treeview(table_frame, columns=columns, show="headings", height=20)
        
        # 컬럼 폭 조정
        col_widths = {"상품명": 200, "상품코드": 150, "현재고": 100, "미입고": 100}
        for col in columns:
            self.products_tree.heading(col, text=col, command=lambda c=col: self.sort_treeview(self.products_tree, c))
            width = col_widths.get(col, 100)
            self.products_tree.column(col, width=width, anchor='center')
        
        scrollbar = ttk.Scrollbar(table_frame, orient=tk.VERTICAL, command=self.products_tree.yview)
        self.products_tree.configure(yscrollcommand=scrollbar.set)
        
        self.products_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        self.products_tree.bind("<Double-1>", self.on_products_tree_double_click)
        self.products_tree.bind("<<TreeviewSelect>>", self.on_product_select)
        
        # 우클릭 메뉴 (발주 메뉴 제거)
        self.products_menu = tk.Menu(self.root, tearoff=0)
        self.products_menu.add_command(label="상세보기", command=self.show_product_detail)
        self.products_menu.add_command(label="수정", command=self.edit_product)
        self.products_menu.add_command(label="클라우드 이미지 검색", command=self.search_cloud_image)
        self.products_menu.add_command(label="삭제", command=self.delete_product)
        self.products_tree.bind("<Button-3>", self.show_products_menu)
        
        # DEL 키로 삭제
        self.products_tree.bind("<Delete>", lambda e: self.delete_product())
        
        self.refresh_products_list()
        
    def on_product_select(self, event=None):
        """상품 선택 시 이미지 표시"""
        selected = self.products_tree.selection()
        if not selected:
            self.product_image_label.config(image='', text="이미지 없음")
            return
        
        item = self.products_tree.item(selected[0])
        product_name = item['values'][0]
        
        product = None
        for p in self.products:
            if p['name'] == product_name:
                product = p
                break
        
        if not product or not product.get('image'):
            self.product_image_label.config(image='', text="이미지 없음")
            return
        
        try:
            img_data = base64.b64decode(product['image'])
            img = Image.open(BytesIO(img_data))
            img.thumbnail((140, 140))
            photo = ImageTk.PhotoImage(img)
            self.product_image_label.config(image=photo, text='')
            self.product_image_label.image = photo  # 참조 유지
        except Exception as e:
            self.product_image_label.config(image='', text="이미지 오류")
    
    def create_orders_tab(self):
        top_frame = ttk.Frame(self.orders_frame)
        top_frame.pack(fill=tk.X, padx=10, pady=10)
        
        # 날짜 선택 프레임 (기본: 오늘)
        date_frame = ttk.Frame(top_frame)
        date_frame.pack(side=tk.LEFT, padx=5)
        
        ttk.Label(date_frame, text="날짜:", font=("Arial", 10, "bold")).pack(side=tk.LEFT, padx=5)
        
        # 날짜 콤보박스
        self.order_date_var = tk.StringVar(value=datetime.now().strftime('%Y-%m-%d'))
        self.order_date_combo = ttk.Combobox(date_frame, textvariable=self.order_date_var, width=15, state="readonly")
        self.order_date_combo.pack(side=tk.LEFT, padx=5)
        self.order_date_combo.bind("<<ComboboxSelected>>", lambda e: self.refresh_orders_list())
        
        ttk.Button(date_frame, text="🔍 기간검색", command=self.search_order_period).pack(side=tk.LEFT, padx=5)
        ttk.Button(date_frame, text="🔄 동기화", command=self.sync_from_db).pack(side=tk.LEFT, padx=5)
        
        # 매장별 발주 현황 버튼
        ttk.Button(top_frame, text="📊 매장별 발주 현황", command=self.show_store_orders).pack(side=tk.RIGHT, padx=5)
        ttk.Button(top_frame, text="📄 발주장 출력 (Excel)", command=self.export_orders_excel).pack(side=tk.RIGHT, padx=5)
        ttk.Button(top_frame, text="📤 엑셀로 발주등록", command=self.import_order_excel).pack(side=tk.RIGHT, padx=5)
        ttk.Button(top_frame, text="📥 엑셀양식 저장", command=self.export_order_template).pack(side=tk.RIGHT, padx=5)
        ttk.Button(top_frame, text="➕ 발주 추가", command=self.add_order).pack(side=tk.RIGHT, padx=5)
        
        columns = ("발주일자", "상품명", "상품코드", "색상", "사이즈", "발주매장", "발주수량", "미입고수량", "메모")
        self.orders_tree = ttk.Treeview(self.orders_frame, columns=columns, show="headings", height=20)
        self.orders_tree.bind("<Double-1>", self.on_orders_tree_double_click)
        
        # 컬럼 폭 조정 (내용에 맞게)
        col_widths = {"발주일자": 100, "상품명": 150, "상품코드": 120, "색상": 80, "사이즈": 70,
                     "발주매장": 100, "발주수량": 80, "미입고수량": 90, "메모": 150}
        for col in columns:
            self.orders_tree.heading(col, text=col, command=lambda c=col: self.sort_treeview(self.orders_tree, c))
            self.orders_tree.column(col, width=col_widths.get(col, 90), anchor='center')
        
        scrollbar = ttk.Scrollbar(self.orders_frame, orient=tk.VERTICAL, command=self.orders_tree.yview)
        self.orders_tree.configure(yscrollcommand=scrollbar.set)
        
        self.orders_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(10,0), pady=10)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y, padx=(0,10), pady=10)
        
        # 더블클릭으로 셀 편집
        self.orders_tree.bind("<Double-1>", self.on_order_double_click)
        
        # 우클릭 메뉴 (수정 기능 추가)
        self.orders_menu = tk.Menu(self.root, tearoff=0)
        self.orders_menu.add_command(label="상품 변경", command=self.change_order_product)
        self.orders_menu.add_command(label="색상 변경", command=self.change_order_color)
        self.orders_menu.add_command(label="사이즈 변경", command=self.change_order_size)
        self.orders_menu.add_command(label="매장 변경", command=self.change_order_store)
        self.orders_menu.add_separator()
        self.orders_menu.add_command(label="삭제", command=self.delete_order)
        self.orders_tree.bind("<Button-3>", self.show_orders_menu)
        
        # DEL 키로 삭제
        self.orders_tree.bind("<Delete>", lambda e: self.delete_order())
        
        # 날짜 목록 초기화
        self.update_order_dates()
        self.refresh_orders_list()
    
    def create_inbound_tab(self):
        """입고 관리 탭 생성"""
        top_frame = ttk.Frame(self.inbound_frame)
        top_frame.pack(fill=tk.X, padx=10, pady=10)

        ttk.Label(top_frame, text="입고 관리", font=("Arial", 14, "bold")).pack(side=tk.LEFT, padx=5)

        ttk.Button(top_frame, text="📥 입고 등록", command=self.add_inbound).pack(side=tk.RIGHT, padx=5)
        ttk.Button(top_frame, text="📄 엑셀 파일 불러오기", command=self.import_inbound_excel).pack(side=tk.RIGHT, padx=5)
        ttk.Button(top_frame, text="📥 엑셀양식 저장", command=self.export_inbound_template).pack(side=tk.RIGHT, padx=5)

        # 상세 필터 프레임
        filter_frame = ttk.LabelFrame(self.inbound_frame, text="🔍 상세 검색", padding=5)
        filter_frame.pack(fill=tk.X, padx=10, pady=5)

        # 첫 번째 줄: 날짜 필터
        row1_frame = ttk.Frame(filter_frame)
        row1_frame.pack(fill=tk.X, pady=2)

        ttk.Label(row1_frame, text="날짜:", font=("Arial", 9)).pack(side=tk.LEFT, padx=5)
        self.inbound_date_filter_var = tk.StringVar(value="전체")
        inbound_date_combo = ttk.Combobox(row1_frame, textvariable=self.inbound_date_filter_var, width=12, state="readonly")
        inbound_date_combo.pack(side=tk.LEFT, padx=5)
        inbound_date_combo.bind("<<ComboboxSelected>>", lambda e: self.refresh_inbound_list())
        self.inbound_date_combo = inbound_date_combo

        ttk.Label(row1_frame, text="기간:", font=("Arial", 9)).pack(side=tk.LEFT, padx=(15,5))
        self.inbound_start_date = DateEntry(row1_frame, width=10, background='darkblue', foreground='white',
                                           borderwidth=2, date_pattern='yyyy-mm-dd')
        self.inbound_start_date.pack(side=tk.LEFT, padx=2)
        ttk.Label(row1_frame, text="~").pack(side=tk.LEFT)
        self.inbound_end_date = DateEntry(row1_frame, width=10, background='darkblue', foreground='white',
                                         borderwidth=2, date_pattern='yyyy-mm-dd')
        self.inbound_end_date.pack(side=tk.LEFT, padx=2)

        # 두 번째 줄: 상품 검색
        row2_frame = ttk.Frame(filter_frame)
        row2_frame.pack(fill=tk.X, pady=2)

        ttk.Label(row2_frame, text="상품:", font=("Arial", 9)).pack(side=tk.LEFT, padx=5)
        self.inbound_product_search_var = tk.StringVar()
        inbound_product_entry = ttk.Entry(row2_frame, textvariable=self.inbound_product_search_var, width=20)
        inbound_product_entry.pack(side=tk.LEFT, padx=5)
        ttk.Label(row2_frame, text="(상품명/코드)", font=("Arial", 8), foreground="gray").pack(side=tk.LEFT)

        # 검색/초기화 버튼
        btn_frame = ttk.Frame(row2_frame)
        btn_frame.pack(side=tk.RIGHT, padx=10)
        ttk.Button(btn_frame, text="🔍 검색", command=self.search_inbound_detail, width=8).pack(side=tk.LEFT, padx=2)
        ttk.Button(btn_frame, text="🔄 초기화", command=self.reset_inbound_filter, width=8).pack(side=tk.LEFT, padx=2)

        columns = ("입고일자", "상품명", "상품코드", "색상", "사이즈", "입고수량", "비고")
        self.inbound_tree = ttk.Treeview(self.inbound_frame, columns=columns, show="headings", height=20)

        col_widths = {"입고일자": 120, "상품명": 200, "상품코드": 150, "색상": 120, "사이즈": 100,
                     "입고수량": 120, "비고": 250}
        for col in columns:
            self.inbound_tree.heading(col, text=col, command=lambda c=col: self.sort_treeview(self.inbound_tree, c))
            self.inbound_tree.column(col, width=col_widths.get(col, 100), anchor='center')
        
        scrollbar = ttk.Scrollbar(self.inbound_frame, orient=tk.VERTICAL, command=self.inbound_tree.yview)
        self.inbound_tree.configure(yscrollcommand=scrollbar.set)
        
        self.inbound_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(10,0), pady=10)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y, padx=(0,10), pady=10)
        
        # 우클릭 메뉴 추가
        self.inbound_menu = tk.Menu(self.root, tearoff=0)
        self.inbound_menu.add_command(label="수량/비고 수정", command=self.edit_inbound)
        self.inbound_menu.add_command(label="상품 변경", command=self.change_inbound_product)
        self.inbound_menu.add_command(label="색상 변경", command=self.change_inbound_color)
        self.inbound_menu.add_command(label="사이즈 변경", command=self.change_inbound_size)
        self.inbound_menu.add_separator()
        self.inbound_menu.add_command(label="삭제", command=self.delete_inbound)
        self.inbound_tree.bind("<Button-3>", self.show_inbound_menu)
        
        # DEL 키로 삭제
        self.inbound_tree.bind("<Delete>", lambda e: self.delete_inbound())
        
        self.update_inbound_date_filter()
        self.refresh_inbound_list()
    
    def create_outbound_tab(self):
        """출고 관리 탭 생성"""
        top_frame = ttk.Frame(self.outbound_frame)
        top_frame.pack(fill=tk.X, padx=10, pady=10)

        ttk.Label(top_frame, text="출고 관리", font=("Arial", 14, "bold")).pack(side=tk.LEFT, padx=5)

        ttk.Button(top_frame, text="📤 출고 등록", command=self.add_outbound).pack(side=tk.RIGHT, padx=5)
        ttk.Button(top_frame, text="📄 엑셀 파일 불러오기", command=self.import_outbound_excel).pack(side=tk.RIGHT, padx=5)
        ttk.Button(top_frame, text="📥 엑셀양식 저장", command=self.export_outbound_template).pack(side=tk.RIGHT, padx=5)

        # 상세 필터 프레임
        filter_frame = ttk.LabelFrame(self.outbound_frame, text="🔍 상세 검색", padding=5)
        filter_frame.pack(fill=tk.X, padx=10, pady=5)

        # 첫 번째 줄: 날짜 필터
        row1_frame = ttk.Frame(filter_frame)
        row1_frame.pack(fill=tk.X, pady=2)

        ttk.Label(row1_frame, text="날짜:", font=("Arial", 9)).pack(side=tk.LEFT, padx=5)
        self.outbound_date_filter_var = tk.StringVar(value="전체")
        outbound_date_combo = ttk.Combobox(row1_frame, textvariable=self.outbound_date_filter_var, width=12, state="readonly")
        outbound_date_combo.pack(side=tk.LEFT, padx=5)
        outbound_date_combo.bind("<<ComboboxSelected>>", lambda e: self.refresh_outbound_list())
        self.outbound_date_combo = outbound_date_combo

        ttk.Label(row1_frame, text="기간:", font=("Arial", 9)).pack(side=tk.LEFT, padx=(15,5))
        self.outbound_start_date = DateEntry(row1_frame, width=10, background='darkblue', foreground='white',
                                            borderwidth=2, date_pattern='yyyy-mm-dd')
        self.outbound_start_date.pack(side=tk.LEFT, padx=2)
        ttk.Label(row1_frame, text="~").pack(side=tk.LEFT)
        self.outbound_end_date = DateEntry(row1_frame, width=10, background='darkblue', foreground='white',
                                          borderwidth=2, date_pattern='yyyy-mm-dd')
        self.outbound_end_date.pack(side=tk.LEFT, padx=2)

        # 두 번째 줄: 상품 검색
        row2_frame = ttk.Frame(filter_frame)
        row2_frame.pack(fill=tk.X, pady=2)

        ttk.Label(row2_frame, text="상품:", font=("Arial", 9)).pack(side=tk.LEFT, padx=5)
        self.outbound_product_search_var = tk.StringVar()
        outbound_product_entry = ttk.Entry(row2_frame, textvariable=self.outbound_product_search_var, width=20)
        outbound_product_entry.pack(side=tk.LEFT, padx=5)
        ttk.Label(row2_frame, text="(상품명/코드)", font=("Arial", 8), foreground="gray").pack(side=tk.LEFT)

        # 검색/초기화 버튼
        btn_frame = ttk.Frame(row2_frame)
        btn_frame.pack(side=tk.RIGHT, padx=10)
        ttk.Button(btn_frame, text="🔍 검색", command=self.search_outbound_detail, width=8).pack(side=tk.LEFT, padx=2)
        ttk.Button(btn_frame, text="🔄 초기화", command=self.reset_outbound_filter, width=8).pack(side=tk.LEFT, padx=2)

        columns = ("출고일자", "상품명", "상품코드", "색상", "사이즈", "출고수량", "비고")
        self.outbound_tree = ttk.Treeview(self.outbound_frame, columns=columns, show="headings", height=20)

        col_widths = {"출고일자": 120, "상품명": 200, "상품코드": 150, "색상": 120, "사이즈": 100,
                     "출고수량": 120, "비고": 250}
        for col in columns:
            self.outbound_tree.heading(col, text=col, command=lambda c=col: self.sort_treeview(self.outbound_tree, c))
            self.outbound_tree.column(col, width=col_widths.get(col, 100), anchor='center')
        
        scrollbar = ttk.Scrollbar(self.outbound_frame, orient=tk.VERTICAL, command=self.outbound_tree.yview)
        self.outbound_tree.configure(yscrollcommand=scrollbar.set)
        
        self.outbound_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(10,0), pady=10)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y, padx=(0,10), pady=10)
        
        # 우클릭 메뉴 추가
        self.outbound_menu = tk.Menu(self.root, tearoff=0)
        self.outbound_menu.add_command(label="수량/비고 수정", command=self.edit_outbound)
        self.outbound_menu.add_command(label="상품 변경", command=self.change_outbound_product)
        self.outbound_menu.add_command(label="색상 변경", command=self.change_outbound_color)
        self.outbound_menu.add_command(label="사이즈 변경", command=self.change_outbound_size)
        self.outbound_menu.add_separator()
        self.outbound_menu.add_command(label="삭제", command=self.delete_outbound)
        self.outbound_tree.bind("<Button-3>", self.show_outbound_menu)
        
        # DEL 키로 삭제
        self.outbound_tree.bind("<Delete>", lambda e: self.delete_outbound())
        
        self.update_outbound_date_filter()
        self.refresh_outbound_list()
        
    def create_stock_tab(self):
        # 상단 제어 패널
        control_frame = ttk.Frame(self.stock_frame)
        control_frame.pack(fill=tk.X, padx=10, pady=10)
        
        ttk.Label(control_frame, text="재고 현황", font=("Arial", 12, "bold")).pack(side=tk.LEFT, padx=10)
        
        # 검색 프레임
        search_frame = ttk.Frame(control_frame)
        search_frame.pack(side=tk.LEFT, padx=20)
        
        ttk.Label(search_frame, text="검색:", font=("Arial", 10)).pack(side=tk.LEFT, padx=5)
        self.stock_search_var = tk.StringVar()
        self.stock_search_var.trace("w", lambda *args: self.refresh_stock_list())
        stock_search_entry = ttk.Entry(search_frame, textvariable=self.stock_search_var, width=30)
        stock_search_entry.pack(side=tk.LEFT, padx=5)
        
        # 버튼 프레임
        button_frame = ttk.Frame(control_frame)
        button_frame.pack(side=tk.RIGHT)
        
        ttk.Button(button_frame, text="📄 엑셀 출력", command=self.export_stock_excel).pack(side=tk.LEFT, padx=3)
        ttk.Button(button_frame, text="🔄 동기화", command=self.sync_from_db).pack(side=tk.LEFT, padx=3)
        
        # 이미지를 위한 프레임 생성
        main_frame = ttk.Frame(self.stock_frame)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # 왼쪽에 이미지 표시 영역
        image_frame = ttk.Frame(main_frame, width=150)
        image_frame.pack(side=tk.LEFT, fill=tk.Y, padx=(0, 10))
        image_frame.pack_propagate(False)
        
        ttk.Label(image_frame, text="상품 이미지", font=("Arial", 10, "bold")).pack(pady=5)
        self.stock_image_label = ttk.Label(image_frame, text="이미지 없음", relief=tk.SUNKEN)
        self.stock_image_label.pack(pady=5, padx=5, fill=tk.BOTH, expand=True)
        
        # 오른쪽에 테이블
        table_frame = ttk.Frame(main_frame)
        table_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        columns = ("매장", "상품명", "상품코드", "색상", "사이즈", "현재고", "총입고", "총출고", "미입고")
        self.stock_tree = ttk.Treeview(table_frame, columns=columns, show="headings", height=25)
        
        col_widths = {"매장": 100, "상품명": 180, "상품코드": 120, "색상": 100, "사이즈": 80, 
                     "현재고": 80, "총입고": 80, "총출고": 80, "미입고": 80}
        
        # 정렬 기능을 위한 컬럼 헤딩 설정
        self.stock_sort_column = None
        self.stock_sort_reverse = False
        
        for col in columns:
            self.stock_tree.heading(col, text=col, 
                                   command=lambda c=col: self.sort_stock_tree(c))
            self.stock_tree.column(col, width=col_widths.get(col, 80), anchor='center')
        
        scrollbar = ttk.Scrollbar(table_frame, orient=tk.VERTICAL, command=self.stock_tree.yview)
        self.stock_tree.configure(yscrollcommand=scrollbar.set)
        
        self.stock_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # 재고 선택 시 이미지 표시
        self.stock_tree.bind("<<TreeviewSelect>>", self.on_stock_select)
        
        # 더블클릭으로 수량 수정
        self.stock_tree.bind("<Double-1>", self.on_stock_double_click)
        
        # 우클릭 메뉴 추가
        self.stock_menu = tk.Menu(self.root, tearoff=0)
        self.stock_menu.add_command(label="색상 변경", command=self.change_stock_color)
        self.stock_menu.add_command(label="사이즈 변경", command=self.change_stock_size)
        self.stock_tree.bind("<Button-3>", self.show_stock_menu)
        
        self.refresh_stock_list()
    
    def add_inbound(self):
        """입고 등록 다이얼로그 - 창고 입고"""
        dialog = tk.Toplevel(self.root)
        dialog.title("입고 등록 (창고)")
        dialog.geometry("600x850")
        dialog.transient(self.root)
        dialog.grab_set()
        utils.center_window(dialog)
        
        # 입고 날짜
        ttk.Label(dialog, text="입고 날짜:", font=("Arial", 10, "bold")).pack(pady=(20,5))
        date_entry = DateEntry(dialog, width=20, background='darkblue', foreground='white', 
                              borderwidth=2, date_pattern='yyyy-mm-dd')
        date_entry.pack(pady=5)
        
        # 창고 입고 안내
        info_label = ttk.Label(dialog, text="💡 모든 입고는 창고로 입고됩니다", 
                              font=("Arial", 9), foreground="blue")
        info_label.pack(pady=5)
        
        ttk.Label(dialog, text="상품 검색 (엔터로 검색):", font=("Arial", 10, "bold")).pack(pady=(10,5))
        
        # 검색창
        search_var = tk.StringVar()
        search_entry = ttk.Entry(dialog, textvariable=search_var, width=40)
        search_entry.pack(pady=5)
        ttk.Label(dialog, text="상품명 또는 코드 입력 후 엔터", 
                 font=("Arial", 8), foreground="gray").pack()
        
        # 검색 결과 리스트
        result_frame = ttk.LabelFrame(dialog, text="검색 결과", padding=10)
        result_frame.pack(pady=10, padx=20, fill=tk.BOTH)
        
        result_listbox = tk.Listbox(result_frame, height=5)
        result_listbox.pack(fill=tk.BOTH, expand=True)
        
        # 선택된 상품 정보
        selected_product = {'product': None}
        
        color_var = tk.StringVar()
        size_var = tk.StringVar()
        quantity_var = tk.StringVar()
        note_var = tk.StringVar()
        
        # 미입고 현황 프레임
        pending_frame = ttk.LabelFrame(dialog, text="📊 미입고 현황", padding=10)
        pending_tree = None
        
        def search_products(event=None):
            search_text = search_var.get().lower().strip()
            result_listbox.delete(0, tk.END)
            
            if not search_text:
                return
            
            for product in self.products:
                if search_text in product['name'].lower() or search_text in product.get('code', '').lower():
                    display_text = f"{product['name']} ({product.get('code', '없음')})"
                    result_listbox.insert(tk.END, display_text)
        
        def on_product_select(event):
            nonlocal pending_tree
            
            selection = result_listbox.curselection()
            if not selection:
                return
            
            selected_text = result_listbox.get(selection[0])
            
            # 상품 찾기
            for product in self.products:
                display_text = f"{product['name']} ({product.get('code', '없음')})"
                if display_text == selected_text:
                    selected_product['product'] = product
                    
                    # 기존 상품 정보 프레임 제거
                    for widget in dialog.winfo_children():
                        if isinstance(widget, ttk.Frame) and widget not in [result_frame, pending_frame]:
                            if hasattr(widget, 'custom_tag') and widget.custom_tag == 'product_info':
                                widget.destroy()
                    
                    info_frame = ttk.Frame(dialog)
                    info_frame.custom_tag = 'product_info'
                    info_frame.pack(pady=10, fill=tk.X, padx=20)
                    
                    ttk.Label(info_frame, text=f"선택: {product['name']}", 
                             font=("Arial", 10, "bold")).pack()
                    
                    # 색상
                    ttk.Label(info_frame, text="색상:", font=("Arial", 9)).pack(pady=(10,2))
                    colors = product.get('colors', [''])
                    color_combo = ttk.Combobox(info_frame, textvariable=color_var, width=30, state="readonly")
                    color_combo['values'] = colors
                    if colors:
                        color_var.set(colors[0])
                    color_combo.pack()
                    
                    # 사이즈
                    ttk.Label(info_frame, text="사이즈:", font=("Arial", 9)).pack(pady=(10,2))
                    sizes = product.get('sizes', ['FREE'])
                    size_combo = ttk.Combobox(info_frame, textvariable=size_var, width=30, state="readonly")
                    size_combo['values'] = sizes
                    if sizes:
                        size_var.set(sizes[0])
                    size_combo.pack()
                    
                    # 수량
                    ttk.Label(info_frame, text="입고 수량:", font=("Arial", 9)).pack(pady=(10,2))
                    quantity_entry = ttk.Entry(info_frame, textvariable=quantity_var, width=30)
                    quantity_entry.pack()
                    quantity_entry.focus()
                    
                    # 비고
                    ttk.Label(info_frame, text="비고:", font=("Arial", 9)).pack(pady=(10,2))
                    note_entry = ttk.Entry(info_frame, textvariable=note_var, width=30)
                    note_entry.pack()
                    
                    # 미입고 현황 표시
                    pending_frame.pack(pady=10, padx=20, fill=tk.BOTH, expand=True)
                    
                    def update_pending_display():
                        """미입고 현황 업데이트"""
                        for widget in pending_frame.winfo_children():
                            widget.destroy()
                        
                        columns = ("매장", "색상", "사이즈", "미입고")
                        pending_tree_new = ttk.Treeview(pending_frame, columns=columns, show="headings", height=6)
                        
                        for col in columns:
                            pending_tree_new.heading(col, text=col)
                            pending_tree_new.column(col, width=120, anchor='center')
                        
                        pending_tree_new.pack(fill=tk.BOTH, expand=True)
                        
                        # 매장별 미입고 데이터 계산
                        for store in self.stores:
                            for color in product.get('colors', ['']):
                                for size in product.get('sizes', ['FREE']):
                                    pending_qty = 0
                                    for order in self.orders:
                                        if (order['product_id'] == product['id'] and
                                            order.get('store_id') == store['id'] and
                                            order.get('color', '') == color and
                                            order.get('size', 'FREE') == size and
                                            order.get('status') != 'completed'):
                                            pending_qty += (order['quantity'] - order['shipped_quantity'])
                                    
                                    if pending_qty > 0:
                                        pending_tree_new.insert('', tk.END, values=(
                                            store['name'],
                                            color if color else '-',
                                            size,
                                            pending_qty
                                        ))
                        
                        return pending_tree_new
                    
                    pending_tree = update_pending_display()
                    
                    # 색상/사이즈 변경 시 업데이트
                    def on_variant_change(*args):
                        nonlocal pending_tree
                        pending_tree = update_pending_display()
                    
                    color_var.trace('w', on_variant_change)
                    size_var.trace('w', on_variant_change)
                    
                    # 엔터 키로 저장
                    quantity_entry.bind('<Return>', lambda e: save_inbound())
                    note_entry.bind('<Return>', lambda e: save_inbound())
                    
                    break
        
        search_entry.bind('<Return>', search_products)
        result_listbox.bind('<<ListboxSelect>>', on_product_select)
        
        def save_inbound():
            if not selected_product['product']:
                messagebox.showwarning("경고", "상품을 선택해주세요.")
                return
            
            try:
                quantity = int(quantity_var.get())
                if quantity <= 0:
                    raise ValueError
            except:
                messagebox.showwarning("경고", "올바른 수량을 입력해주세요.")
                return
            
            product = selected_product['product']
            
            # 입고 기록 생성 (창고 입고 - store_id는 None)
            inbound_record = {
                'id': len(self.inbound_records) + 1,
                'date': date_entry.get_date().strftime('%Y-%m-%d'),
                'product_id': product['id'],
                'product_name': product['name'],
                'product_code': product.get('code', ''),
                'color': color_var.get(),
                'size': size_var.get() if size_var.get() else 'FREE',
                'quantity': quantity,
                'store_id': None,  # 창고 입고
                'note': note_var.get()
            }
            
            # DB에 저장
            self.data_manager.add_inbound(inbound_record)
            
            # 재고 이동 기록 (창고 입고)
            movement = {
                'date': inbound_record['date'],
                'product_id': product['id'],
                'color': inbound_record['color'],
                'size': inbound_record['size'],
                'quantity': quantity,
                'from_location': '외부',
                'to_location': '창고',
                'type': 'in',
                'note': note_var.get()
            }
            # DB에 저장
            self.data_manager.add_movement(movement)
            
            # 미입고 수량 업데이트 (모든 매장의 발주에 대해)
            remaining_qty = quantity
            for order in self.orders:
                if (order['product_id'] == product['id'] and 
                    order.get('color', '') == inbound_record['color'] and 
                    order.get('size', 'FREE') == inbound_record['size'] and
                    order['status'] != 'completed' and remaining_qty > 0):
                    
                    pending = order['quantity'] - order['shipped_quantity']
                    if pending > 0:
                        shipped = min(pending, remaining_qty)
                        order['shipped_quantity'] += shipped
                        remaining_qty -= shipped
                        
                        # 상태 업데이트
                        if order['shipped_quantity'] >= order['quantity']:
                            order['status'] = 'completed'
                        elif order['shipped_quantity'] > 0:
                            order['status'] = 'partial'
            
            # self.data_manager.save_data()
            # ← 자동저장 제거됨
            self._refresh_data_shortcuts()
            self.refresh_inbound_list()
            self.refresh_orders_list()
            self.refresh_products_list()
            self.refresh_stock_list()
            
            # 입력 필드 초기화
            quantity_var.set("")
            note_var.set("")
            messagebox.showinfo("완료", "입고가 등록되었습니다.\n계속 추가할 수 있습니다.")
        
        btn_frame = ttk.Frame(dialog)
        btn_frame.pack(pady=20)
        ttk.Button(btn_frame, text="💾 저장", command=save_inbound, width=15).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="❌ 닫기", command=dialog.destroy, width=15).pack(side=tk.LEFT, padx=5)
    
    def add_outbound(self):
        """출고 등록 다이얼로그"""
        dialog = tk.Toplevel(self.root)
        dialog.title("출고 등록")
        dialog.geometry("600x900")
        dialog.transient(self.root)
        dialog.grab_set()
        utils.center_window(dialog)
        
        # 출고 날짜
        ttk.Label(dialog, text="출고 날짜:", font=("Arial", 10, "bold")).pack(pady=(20,5))
        date_entry = DateEntry(dialog, width=20, background='darkblue', foreground='white', 
                              borderwidth=2, date_pattern='yyyy-mm-dd')
        date_entry.pack(pady=5)
        
        # 매장 선택
        ttk.Label(dialog, text="출고 매장:", font=("Arial", 10, "bold")).pack(pady=(10,5))
        store_var = tk.StringVar()
        store_combo = ttk.Combobox(dialog, textvariable=store_var, width=40, state="readonly")
        store_combo['values'] = [s['name'] for s in self.stores]
        if self.stores:
            store_var.set(self.stores[0]['name'])
        store_combo.pack(pady=5)
        
        ttk.Label(dialog, text="상품 검색 (엔터로 검색):", font=("Arial", 10, "bold")).pack(pady=(10,5))
        
        # 검색창
        search_var = tk.StringVar()
        search_entry = ttk.Entry(dialog, textvariable=search_var, width=40)
        search_entry.pack(pady=5)
        ttk.Label(dialog, text="상품명 또는 코드 입력 후 엔터", 
                 font=("Arial", 8), foreground="gray").pack()
        
        # 검색 결과 리스트
        result_frame = ttk.LabelFrame(dialog, text="검색 결과", padding=10)
        result_frame.pack(pady=10, padx=20, fill=tk.BOTH)
        
        result_listbox = tk.Listbox(result_frame, height=5)
        result_listbox.pack(fill=tk.BOTH, expand=True)
        
        # 선택된 상품 정보
        selected_product = {'product': None}
        
        color_var = tk.StringVar()
        size_var = tk.StringVar()
        quantity_var = tk.StringVar()
        note_var = tk.StringVar()
        
        # 미입고 현황 프레임
        pending_frame = ttk.LabelFrame(dialog, text="📊 매장별 재고 현황", padding=10)
        pending_tree = None
        
        def search_products(event=None):
            search_text = search_var.get().lower().strip()
            result_listbox.delete(0, tk.END)
            
            if not search_text:
                return
            
            for product in self.products:
                if search_text in product['name'].lower() or search_text in product.get('code', '').lower():
                    display_text = f"{product['name']} ({product.get('code', '없음')})"
                    result_listbox.insert(tk.END, display_text)
        
        def on_product_select(event):
            nonlocal pending_tree
            
            selection = result_listbox.curselection()
            if not selection:
                return
            
            selected_text = result_listbox.get(selection[0])
            
            # 상품 찾기
            for product in self.products:
                display_text = f"{product['name']} ({product.get('code', '없음')})"
                if display_text == selected_text:
                    selected_product['product'] = product
                    
                    # 기존 상품 정보 프레임 제거
                    for widget in dialog.winfo_children():
                        if isinstance(widget, ttk.Frame) and widget not in [result_frame, pending_frame]:
                            if hasattr(widget, 'custom_tag') and widget.custom_tag == 'product_info':
                                widget.destroy()
                    
                    info_frame = ttk.Frame(dialog)
                    info_frame.custom_tag = 'product_info'
                    info_frame.pack(pady=10, fill=tk.X, padx=20)
                    
                    ttk.Label(info_frame, text=f"선택: {product['name']}", 
                             font=("Arial", 10, "bold")).pack()
                    
                    # 색상
                    ttk.Label(info_frame, text="색상:", font=("Arial", 9)).pack(pady=(10,2))
                    colors = product.get('colors', [''])
                    color_combo = ttk.Combobox(info_frame, textvariable=color_var, width=30, state="readonly")
                    color_combo['values'] = colors
                    if colors:
                        color_var.set(colors[0])
                    color_combo.pack()
                    
                    # 사이즈
                    ttk.Label(info_frame, text="사이즈:", font=("Arial", 9)).pack(pady=(10,2))
                    sizes = product.get('sizes', ['FREE'])
                    size_combo = ttk.Combobox(info_frame, textvariable=size_var, width=30, state="readonly")
                    size_combo['values'] = sizes
                    if sizes:
                        size_var.set(sizes[0])
                    size_combo.pack()
                    
                    # 수량
                    ttk.Label(info_frame, text="출고 수량:", font=("Arial", 9)).pack(pady=(10,2))
                    quantity_entry = ttk.Entry(info_frame, textvariable=quantity_var, width=30)
                    quantity_entry.pack()
                    quantity_entry.focus()
                    
                    # 비고
                    ttk.Label(info_frame, text="비고:", font=("Arial", 9)).pack(pady=(10,2))
                    note_entry = ttk.Entry(info_frame, textvariable=note_var, width=30)
                    note_entry.pack()
                    
                    # 매장별 재고 현황 표시
                    pending_frame.pack(pady=10, padx=20, fill=tk.BOTH, expand=True)
                    
                    def update_stock_display():
                        """매장별 재고 현황 업데이트"""
                        for widget in pending_frame.winfo_children():
                            widget.destroy()
                        
                        columns = ("매장", "색상", "사이즈", "현재고", "미입고")
                        stock_tree_new = ttk.Treeview(pending_frame, columns=columns, show="headings", height=6)
                        
                        for col in columns:
                            stock_tree_new.heading(col, text=col)
                            stock_tree_new.column(col, width=100, anchor='center')
                        
                        stock_tree_new.pack(fill=tk.BOTH, expand=True)
                        
                        # 매장별 재고 데이터 계산
                        for store in self.stores:
                            for color in product.get('colors', ['']):
                                for size in product.get('sizes', ['FREE']):
                                    # 현재고 계산
                                    stock = self.data_manager.calculate_stock_by_variant(product['id'], color, size)
                                    
                                    # 미입고 계산
                                    pending_qty = 0
                                    for order in self.orders:
                                        if (order['product_id'] == product['id'] and
                                            order.get('store_id') == store['id'] and
                                            order.get('color', '') == color and
                                            order.get('size', 'FREE') == size and
                                            order.get('status') != 'completed'):
                                            pending_qty += (order['quantity'] - order['shipped_quantity'])
                                    
                                    if stock > 0 or pending_qty > 0:
                                        stock_tree_new.insert('', tk.END, values=(
                                            store['name'],
                                            color if color else '-',
                                            size,
                                            stock,
                                            pending_qty
                                        ))
                        
                        return stock_tree_new
                    
                    pending_tree = update_stock_display()
                    
                    # 색상/사이즈 변경 시 업데이트
                    def on_variant_change(*args):
                        nonlocal pending_tree
                        pending_tree = update_stock_display()
                    
                    color_var.trace('w', on_variant_change)
                    size_var.trace('w', on_variant_change)
                    
                    # 엔터 키로 저장
                    quantity_entry.bind('<Return>', lambda e: save_outbound())
                    note_entry.bind('<Return>', lambda e: save_outbound())
                    
                    break
        
        search_entry.bind('<Return>', search_products)
        result_listbox.bind('<<ListboxSelect>>', on_product_select)
        
        def save_outbound():
            if not selected_product['product']:
                messagebox.showwarning("경고", "상품을 선택해주세요.")
                return
            
            if not store_var.get() and self.stores:
                messagebox.showwarning("경고", "출고 매장을 선택해주세요.")
                return
            
            try:
                quantity = int(quantity_var.get())
                if quantity <= 0:
                    raise ValueError
            except:
                messagebox.showwarning("경고", "올바른 수량을 입력해주세요.")
                return
            
            product = selected_product['product']
            
            # 선택된 매장 찾기
            store_id = None
            store_name = store_var.get()
            for s in self.stores:
                if s['name'] == store_name:
                    store_id = s['id']
                    break
            
            # 출고 기록 생성
            outbound_record = {
                'date': date_entry.get_date().strftime('%Y-%m-%d'),
                'product_id': product['id'],
                'product_name': product['name'],
                'product_code': product.get('code', ''),
                'color': color_var.get(),
                'size': size_var.get() if size_var.get() else 'FREE',
                'quantity': quantity,
                'store_id': store_id,
                'note': note_var.get()
            }
            
            # DB에 저장
            self.data_manager.add_outbound(outbound_record)
            
            # 매장명 가져오기
            store = self.data_manager.get_store_by_id(store_id)
            store_name = store['name'] if store else '매장'
            
            # 재고 이동 기록 (출고)
            movement = {
                'date': outbound_record['date'],
                'product_id': product['id'],
                'color': outbound_record['color'],
                'size': outbound_record['size'],
                'quantity': quantity,
                'from_location': '창고',
                'to_location': store_name,
                'type': 'out',
                'note': note_var.get()
            }
            # DB에 저장
            self.data_manager.add_movement(movement)
            
            # self.data_manager.save_data()
            # ← 자동저장 제거됨
            self._refresh_data_shortcuts()
            self.refresh_outbound_list()
            self.refresh_products_list()
            self.refresh_stock_list()
            self.update_outbound_date_filter()
            
            # 입력 필드 초기화
            quantity_var.set("")
            note_var.set("")
            messagebox.showinfo("완료", "출고가 등록되었습니다.\n계속 추가할 수 있습니다.")
        
        btn_frame = ttk.Frame(dialog)
        btn_frame.pack(pady=20)
        ttk.Button(btn_frame, text="💾 저장", command=save_outbound, width=15).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="❌ 닫기", command=dialog.destroy, width=15).pack(side=tk.LEFT, padx=5)
    
    def import_inbound_excel(self):
        """입고장 엑셀 파일 불러오기"""
        filename = filedialog.askopenfilename(
            title="입고장 엑셀 파일 선택",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        
        if not filename:
            return
        
        try:
            wb = openpyxl.load_workbook(filename)
            ws = wb.active
            
            imported_count = 0
            errors = []
            
            # 첫 행은 헤더로 간주하고 스킵
            # 양식: 모델명, 컬러, 사이즈, 입고수량
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    if not row[0]:  # 모델명이 없으면 스킵
                        continue
                    
                    # 엑셀 컬럼: 모델명, 컬러, 사이즈, 입고수량
                    product_name = str(row[0]) if row[0] else ""
                    color = str(row[1]) if row[1] else ""
                    size = str(row[2]) if row[2] else "FREE"
                    quantity = int(row[3]) if row[3] else 0
                    
                    # 오늘 날짜 사용
                    date_str = datetime.now().strftime('%Y-%m-%d')
                    
                    # 상품 찾기 (모델명으로만 검색)
                    product = None
                    for p in self.products:
                        if p['name'] == product_name:
                            product = p
                            break
                    
                    if not product:
                        errors.append(f"행 {row_idx}: 상품을 찾을 수 없음 ({product_name})")
                        continue
                    
                    if quantity <= 0:
                        errors.append(f"행 {row_idx}: 잘못된 수량")
                        continue
                    
                    # 입고 기록 생성
                    inbound_record = {
                        'date': date_str,
                        'product_id': product['id'],
                        'product_name': product['name'],
                        'product_code': product.get('code', ''),
                        'color': color,
                        'size': size,
                        'quantity': quantity,
                        'note': ''
                    }
                    
                    # DB에 저장
                    self.data_manager.add_inbound(inbound_record)
                    
                    # 재고 이동 기록
                    movement = {
                        'date': date_str,
                        'product_id': product['id'],
                        'color': color,
                        'size': size,
                        'quantity': quantity,
                        'from_location': '외부',
                        'to_location': '창고',
                        'type': 'in',
                        'note': ''
                    }
                    # DB에 저장
                    self.data_manager.add_movement(movement)
                    
                    # 미입고 수량 업데이트
                    remaining_qty = quantity
                    for order in self.orders:
                        if (order['product_id'] == product['id'] and 
                            order.get('color', '') == color and 
                            order.get('size', 'FREE') == size and
                            order['status'] != 'completed' and remaining_qty > 0):
                            
                            pending = order['quantity'] - order['shipped_quantity']
                            if pending > 0:
                                shipped = min(pending, remaining_qty)
                                order['shipped_quantity'] += shipped
                                remaining_qty -= shipped
                                
                                # 상태 업데이트
                                if order['shipped_quantity'] >= order['quantity']:
                                    order['status'] = 'completed'
                                elif order['shipped_quantity'] > 0:
                                    order['status'] = 'partial'
                    
                    imported_count += 1
                    
                except Exception as e:
                    errors.append(f"행 {row_idx}: {str(e)}")
            
            # self.data_manager.save_data()
            # ← 자동저장 제거됨
            self._refresh_data_shortcuts()
            self.refresh_inbound_list()
            self.refresh_products_list()
            self.refresh_stock_list()
            self.refresh_orders_list()
            self.update_inbound_date_filter()
            
            result_msg = f"총 {imported_count}개의 입고 기록을 가져왔습니다."
            if errors:
                result_msg += f"\n\n오류 {len(errors)}건:\n" + "\n".join(errors[:10])
                if len(errors) > 10:
                    result_msg += f"\n... 외 {len(errors)-10}건"
            
            messagebox.showinfo("완료", result_msg)
            
        except Exception as e:
            messagebox.showerror("오류", f"파일을 읽는 중 오류가 발생했습니다:\n{str(e)}")
    
    def import_outbound_excel(self):
        """출고장 엑셀 파일 불러오기"""
        filename = filedialog.askopenfilename(
            title="출고장 엑셀 파일 선택",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        
        if not filename:
            return
        
        try:
            wb = openpyxl.load_workbook(filename)
            ws = wb.active
            
            imported_count = 0
            errors = []
            
            # 첫 행은 헤더로 간주하고 스킵
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    if not row[0]:  # 날짜가 없으면 스킵
                        continue
                    
                    # 엑셀 컬럼: 날짜, 상품명, 상품코드, 색상, 사이즈, 수량, 비고
                    date_val = row[0]
                    product_name = str(row[1]) if row[1] else ""
                    product_code = str(row[2]) if row[2] else ""
                    color = str(row[3]) if row[3] else ""
                    size = str(row[4]) if row[4] else "FREE"
                    quantity = int(row[5]) if row[5] else 0
                    note = str(row[6]) if row[6] else ""
                    
                    # 날짜 처리
                    if isinstance(date_val, datetime):
                        date_str = date_val.strftime('%Y-%m-%d')
                    else:
                        date_str = str(date_val)
                    
                    # 상품 찾기
                    product = None
                    for p in self.products:
                        if p['name'] == product_name or p.get('code') == product_code:
                            product = p
                            break
                    
                    if not product:
                        errors.append(f"행 {row_idx}: 상품을 찾을 수 없음 ({product_name})")
                        continue
                    
                    if quantity <= 0:
                        errors.append(f"행 {row_idx}: 잘못된 수량")
                        continue
                    
                    # 출고 기록 생성
                    outbound_record = {
                        'date': date_str,
                        'product_id': product['id'],
                        'product_name': product['name'],
                        'product_code': product.get('code', ''),
                        'color': color,
                        'size': size,
                        'quantity': quantity,
                        'note': note
                    }
                    
                    # DB에 저장
                    self.data_manager.add_outbound(outbound_record)
                    
                    # 재고 이동 기록
                    movement = {
                        'date': date_str,
                        'product_id': product['id'],
                        'color': color,
                        'size': size,
                        'quantity': quantity,
                        'from_location': '창고',
                        'to_location': '출고',
                        'type': 'out',
                        'note': note
                    }
                    # DB에 저장
                    self.data_manager.add_movement(movement)
                    
                    imported_count += 1
                    
                except Exception as e:
                    errors.append(f"행 {row_idx}: {str(e)}")
            
            # self.data_manager.save_data()
            # ← 자동저장 제거됨
            self._refresh_data_shortcuts()
            self.refresh_outbound_list()
            self.refresh_products_list()
            self.refresh_stock_list()
            self.update_outbound_date_filter()
            
            result_msg = f"총 {imported_count}개의 출고 기록을 가져왔습니다."
            if errors:
                result_msg += f"\n\n오류 {len(errors)}건:\n" + "\n".join(errors[:10])
                if len(errors) > 10:
                    result_msg += f"\n... 외 {len(errors)-10}건"
            
            messagebox.showinfo("완료", result_msg)
            
        except Exception as e:
            messagebox.showerror("오류", f"파일을 읽는 중 오류가 발생했습니다:\n{str(e)}")
    
    def delete_inbound(self):
        """입고 기록 삭제 (다중 선택 지원)"""
        selected = self.inbound_tree.selection()
        if not selected:
            messagebox.showwarning("경고", "삭제할 입고 기록을 선택해주세요.")
            return
        
        delete_count = len(selected)
        if not messagebox.askyesno("확인", f"선택한 {delete_count}개의 입고 기록을 삭제하시겠습니까?"):
            return
        
        deleted_count = 0
        for item_id in selected:
            item = self.inbound_tree.item(item_id)
            values = item['values']
            date = values[0]
            product_name = values[1]
            color = values[3] if values[3] != '-' else ''
            size = values[4]
            quantity = values[5]
            
            # 입고 기록 찾아서 삭제
            for record in self.inbound_records[:]:
                record_color = record.get('color', '')
                if (record['date'] == date and 
                    record['product_name'] == product_name and 
                    record_color == color and 
                    record['size'] == size and 
                    record['quantity'] == quantity):
                    
                    self.inbound_records.remove(record)
                    
                    # 해당 재고 이동 기록도 삭제
                    for movement in self.movements[:]:
                        movement_color = movement.get('color', '')
                        if (movement['product_id'] == record['product_id'] and 
                            movement['date'] == date and 
                            movement['type'] == 'in' and
                            movement_color == color and 
                            movement['size'] == size and 
                            movement['quantity'] == quantity):
                            self.movements.remove(movement)
                            break
                    
                    deleted_count += 1
                    break
        
        if deleted_count > 0:
            # self.data_manager.save_data()
            # ← 자동저장 제거됨
            self._refresh_data_shortcuts()
            self.refresh_inbound_list()
            self.refresh_products_list()
            self.refresh_stock_list()
            messagebox.showinfo("완료", f"{deleted_count}개의 입고 기록이 삭제되었습니다.")
        else:
            messagebox.showwarning("경고", "삭제된 기록이 없습니다.")
    
    
    def delete_outbound(self):
        """출고 기록 삭제 (다중 선택 지원)"""
        selected = self.outbound_tree.selection()
        if not selected:
            messagebox.showwarning("경고", "삭제할 출고 기록을 선택해주세요.")
            return
        
        delete_count = len(selected)
        if not messagebox.askyesno("확인", f"선택한 {delete_count}개의 출고 기록을 삭제하시겠습니까?"):
            return
        
        deleted_count = 0
        for item_id in selected:
            item = self.outbound_tree.item(item_id)
            values = item['values']
            date = values[0]
            product_name = values[1]
            color = values[3] if values[3] != '-' else ''
            size = values[4]
            quantity = values[5]
            
            # 출고 기록 찾아서 삭제
            for record in self.outbound_records[:]:
                record_color = record.get('color', '')
                if (record['date'] == date and 
                    record['product_name'] == product_name and 
                    record_color == color and 
                    record['size'] == size and 
                    record['quantity'] == quantity):
                    
                    self.outbound_records.remove(record)
                    
                    # 해당 재고 이동 기록도 삭제
                    for movement in self.movements[:]:
                        movement_color = movement.get('color', '')
                        if (movement['product_id'] == record['product_id'] and 
                            movement['date'] == date and 
                            movement['type'] == 'out' and
                            movement_color == color and 
                            movement['size'] == size and 
                            movement['quantity'] == quantity):
                            self.movements.remove(movement)
                            break
                    
                    deleted_count += 1
                    break
        
        if deleted_count > 0:
            # self.data_manager.save_data()
            # ← 자동저장 제거됨
            self._refresh_data_shortcuts()
            self.refresh_outbound_list()
            self.refresh_products_list()
            self.refresh_stock_list()
            messagebox.showinfo("완료", f"{deleted_count}개의 출고 기록이 삭제되었습니다.")
        else:
            messagebox.showwarning("경고", "삭제된 기록이 없습니다.")
    
    
    def edit_inbound(self):
        """입고 기록 수정"""
        selected = self.inbound_tree.selection()
        if not selected:
            messagebox.showwarning("경고", "수정할 입고 기록을 선택해주세요.")
            return
        
        item = self.inbound_tree.item(selected[0])
        values = item['values']
        date = values[0]
        product_name = values[1]
        color = values[3] if values[3] != '-' else ''
        size = values[4]
        old_quantity = values[5]
        note = values[6] if len(values) > 6 else ""
        
        # 입고 기록 찾기
        record = None
        for r in self.inbound_records:
            record_color = r.get('color', '')
            if (r['date'] == date and 
                r['product_name'] == product_name and 
                record_color == color and 
                r['size'] == size and 
                r['quantity'] == old_quantity):
                record = r
                break
        
        if not record:
            messagebox.showerror("오류", "입고 기록을 찾을 수 없습니다.")
            return
        
        # 수정 다이얼로그
        dialog = tk.Toplevel(self.root)
        dialog.title("입고 수정")
        dialog.geometry("400x380")
        dialog.transient(self.root)
        dialog.grab_set()
        utils.center_window(dialog)
        
        ttk.Label(dialog, text=f"상품: {product_name}", font=("Arial", 11, "bold")).pack(pady=10)
        ttk.Label(dialog, text=f"색상: {color if color else '-'}", font=("Arial", 10)).pack(pady=5)
        ttk.Label(dialog, text=f"사이즈: {size}", font=("Arial", 10)).pack(pady=5)
        ttk.Label(dialog, text=f"입고일자: {date}", font=("Arial", 10)).pack(pady=5)
        
        ttk.Label(dialog, text="입고 수량 *", font=("Arial", 10, "bold")).pack(pady=(10,0))
        quantity_var = tk.StringVar(value=str(old_quantity))
        quantity_entry = ttk.Entry(dialog, textvariable=quantity_var, width=30)
        quantity_entry.pack(pady=5)
        
        ttk.Label(dialog, text="비고", font=("Arial", 10)).pack(pady=(10,0))
        note_var = tk.StringVar(value=note)
        note_entry = ttk.Entry(dialog, textvariable=note_var, width=30)
        note_entry.pack(pady=5)
        
        def save_changes():
            try:
                new_quantity = int(quantity_var.get())
                if new_quantity <= 0:
                    raise ValueError
            except:
                messagebox.showerror("오류", "올바른 수량을 입력해주세요.")
                return
            
            # 기록 업데이트
            record['quantity'] = new_quantity
            record['note'] = note_var.get()
            
            # Supabase에 저장
            self.data_manager.update_inbound_record_in_db(
                record['id'], 
                {'quantity': new_quantity, 'note': note_var.get()}
            )
            
            # 재고 이동 기록도 업데이트
            for movement in self.movements:
                movement_color = movement.get('color', '')
                if (movement['product_id'] == record['product_id'] and 
                    movement['date'] == date and 
                    movement['type'] == 'in' and
                    movement_color == color and 
                    movement['size'] == size and 
                    movement['quantity'] == old_quantity):
                    movement['quantity'] = new_quantity
                    movement['note'] = note_var.get()
                    
                    # Supabase에 저장
                    self.data_manager.update_movement_in_db(
                        movement['id'],
                        {'quantity': new_quantity, 'note': note_var.get()}
                    )
                    break
            
            self._refresh_data_shortcuts()
            self.refresh_inbound_list()
            self.refresh_products_list()
            self.refresh_stock_list()
            dialog.destroy()
            messagebox.showinfo("완료", "입고 기록이 수정되었습니다.")
        
        button_frame = ttk.Frame(dialog)
        button_frame.pack(pady=20)
        ttk.Button(button_frame, text="저장", command=save_changes).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="취소", command=dialog.destroy).pack(side=tk.LEFT, padx=5)
    
    def _get_selected_inbound_info(self):
        """선택된 입고 기록 정보 가져오기"""
        selected = self.inbound_tree.selection()
        if not selected:
            messagebox.showwarning("경고", "수정할 입고 기록을 선택해주세요.")
            return None
        
        item = self.inbound_tree.item(selected[0])
        values = item['values']
        date = str(values[0])
        product_name = str(values[1])
        product_code = str(values[2]) if values[2] != '-' else ''
        color = str(values[3]) if values[3] != '-' else ''
        size = str(values[4])
        quantity = values[5]
        
        # 상품 찾기 (상품코드가 있으면 상품코드로 우선 검색)
        product = None
        if product_code:
            for p in self.products:
                if p.get('code', '') == product_code:
                    product = p
                    break
        
        if not product:
            for p in self.products:
                if p['name'] == product_name and p.get('code', '') == product_code:
                    product = p
                    break
        
        if not product and not product_code:
            for p in self.products:
                if p['name'] == product_name:
                    product = p
                    break
        
        if not product:
            messagebox.showerror("오류", f"상품을 찾을 수 없습니다.\n상품명: {product_name}\n상품코드: {product_code}")
            return None
        
        # 입고 기록 찾기 (product_id로 검색)
        record = None
        for r in self.inbound_records:
            record_color = r.get('color', '')
            if (str(r['product_id']) == str(product['id']) and
                str(r['date']) == date and 
                record_color == color and 
                r['size'] == size and 
                r['quantity'] == quantity):
                record = r
                break
        
        if not record:
            messagebox.showerror("오류", "입고 기록을 찾을 수 없습니다.")
            return None
        
        return {
            'record': record,
            'product': product,
            'date': date,
            'color': color,
            'size': size,
            'quantity': quantity
        }
    
    def change_inbound_product(self):
        """입고 상품 변경"""
        info = self._get_selected_inbound_info()
        if not info:
            return
        
        record = info['record']
        old_quantity = info['quantity']
        old_color = info['color']
        old_size = info['size']
        old_date = info['date']
        
        # 상품 선택 다이얼로그
        dialog = tk.Toplevel(self.root)
        dialog.title("상품 변경")
        dialog.geometry("500x600")
        dialog.transient(self.root)
        dialog.grab_set()
        utils.center_window(dialog)
        
        ttk.Label(dialog, text="변경할 상품 선택", font=("Arial", 12, "bold")).pack(pady=10)
        
        # 검색
        search_frame = ttk.Frame(dialog)
        search_frame.pack(fill=tk.X, padx=10, pady=5)
        ttk.Label(search_frame, text="검색:").pack(side=tk.LEFT)
        search_var = tk.StringVar()
        search_entry = ttk.Entry(search_frame, textvariable=search_var, width=30)
        search_entry.pack(side=tk.LEFT, padx=5)
        
        # 상품 목록
        columns = ("상품명", "상품코드")
        tree = ttk.Treeview(dialog, columns=columns, show="headings", height=15)
        tree.heading("상품명", text="상품명")
        tree.heading("상품코드", text="상품코드")
        tree.column("상품명", width=250)
        tree.column("상품코드", width=150)
        tree.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        
        def refresh_list():
            tree.delete(*tree.get_children())
            search_text = search_var.get().lower()
            for p in self.products:
                name = p['name'].lower()
                code = p.get('code', '').lower()
                if search_text in name or search_text in code:
                    tree.insert('', tk.END, values=(p['name'], p.get('code', '')))
        
        refresh_list()
        search_var.trace('w', lambda *args: refresh_list())
        
        def select_product():
            selected = tree.selection()
            if not selected:
                messagebox.showwarning("경고", "상품을 선택해주세요.")
                return
            
            new_product_name = tree.item(selected[0])['values'][0]
            new_product = None
            for p in self.products:
                if p['name'] == new_product_name:
                    new_product = p
                    break
            
            if not new_product:
                return
            
            # 입고 기록 업데이트
            record['product_id'] = new_product['id']
            record['product_name'] = new_product['name']
            
            # 색상이 새 상품에 없으면 초기화
            new_colors = new_product.get('colors', [])
            if record.get('color', '') and record.get('color', '') not in new_colors:
                record['color'] = new_colors[0] if new_colors else ''
            
            # 사이즈가 새 상품에 없으면 초기화
            new_sizes = new_product.get('sizes', ['FREE'])
            if record.get('size', 'FREE') not in new_sizes:
                record['size'] = new_sizes[0] if new_sizes else 'FREE'
            
            # 재고 이동 기록도 업데이트
            for movement in self.movements:
                movement_color = movement.get('color', '')
                if (str(movement['product_id']) == str(info['product']['id']) and 
                    str(movement['date']) == old_date and 
                    movement['type'] == 'in' and
                    movement_color == old_color and 
                    movement['size'] == old_size and 
                    movement['quantity'] == old_quantity):
                    movement['product_id'] = new_product['id']
                    if record.get('color', '') != old_color:
                        movement['color'] = record.get('color', '')
                    if record.get('size', 'FREE') != old_size:
                        movement['size'] = record.get('size', 'FREE')
                    break
            
            self._refresh_data_shortcuts()
            self.refresh_inbound_list()
            self.refresh_products_list()
            self.refresh_stock_list()
            dialog.destroy()
            messagebox.showinfo("완료", "상품이 변경되었습니다.")
        
        btn_frame = ttk.Frame(dialog)
        btn_frame.pack(pady=10)
        ttk.Button(btn_frame, text="선택", command=select_product, width=10).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="취소", command=dialog.destroy, width=10).pack(side=tk.LEFT, padx=5)
    
    def change_inbound_color(self):
        """입고 색상 변경"""
        info = self._get_selected_inbound_info()
        if not info:
            return
        
        product = info['product']
        if not product:
            messagebox.showerror("오류", "상품을 찾을 수 없습니다.")
            return
        
        record = info['record']
        old_color = info['color']
        old_quantity = info['quantity']
        old_date = info['date']
        old_size = info['size']
        
        # 같은 상품명을 가진 모든 상품에서 색상 수집
        product_name = product['name']
        color_to_product = {}  # {색상: 상품} 매핑
        all_colors = []
        
        for p in self.products:
            if p['name'] == product_name:
                p_colors = p.get('colors', [])
                for c in p_colors:
                    if c and c not in all_colors:
                        all_colors.append(c)
                        color_to_product[c] = p
        
        if not all_colors:
            messagebox.showinfo("안내", "이 상품은 색상이 없습니다.")
            return
        
        # 색상 선택 다이얼로그
        dialog = tk.Toplevel(self.root)
        dialog.title("색상 변경")
        dialog.geometry("300x400")
        dialog.transient(self.root)
        dialog.grab_set()
        utils.center_window(dialog)
        
        ttk.Label(dialog, text=f"상품: {product_name}", font=("Arial", 11, "bold")).pack(pady=10)
        ttk.Label(dialog, text=f"현재 색상: {old_color if old_color else '-'}").pack(pady=5)
        ttk.Label(dialog, text="변경할 색상 선택:", font=("Arial", 10)).pack(pady=10)
        
        listbox = tk.Listbox(dialog, height=10, font=("Arial", 11))
        for color in all_colors:
            listbox.insert(tk.END, color)
        listbox.pack(fill=tk.BOTH, expand=True, padx=20, pady=5)
        
        # 현재 색상 선택
        if old_color in all_colors:
            listbox.selection_set(all_colors.index(old_color))
        
        def select_color():
            selected = listbox.curselection()
            if not selected:
                messagebox.showwarning("경고", "색상을 선택해주세요.")
                return
            
            new_color = all_colors[selected[0]]
            new_product = color_to_product.get(new_color, product)
            
            # 기록 업데이트
            record['color'] = new_color
            record['product_id'] = new_product['id']
            record['product_name'] = new_product['name']
            record['product_code'] = new_product.get('code', '')
            
            # 재고 이동 기록도 업데이트
            for movement in self.movements:
                movement_color = movement.get('color', '')
                if (str(movement['product_id']) == str(info['product']['id']) and 
                    str(movement['date']) == old_date and 
                    movement['type'] == 'in' and
                    movement_color == old_color and 
                    movement['size'] == old_size and 
                    movement['quantity'] == old_quantity):
                    movement['color'] = new_color
                    movement['product_id'] = new_product['id']
                    break
            
            self._refresh_data_shortcuts()
            self.refresh_inbound_list()
            self.refresh_stock_list()
            dialog.destroy()
            messagebox.showinfo("완료", f"색상이 '{new_color}'로 변경되었습니다.")
        
        btn_frame = ttk.Frame(dialog)
        btn_frame.pack(pady=10)
        ttk.Button(btn_frame, text="선택", command=select_color, width=10).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="취소", command=dialog.destroy, width=10).pack(side=tk.LEFT, padx=5)
    
    def change_inbound_size(self):
        """입고 사이즈 변경"""
        info = self._get_selected_inbound_info()
        if not info:
            return
        
        product = info['product']
        if not product:
            messagebox.showerror("오류", "상품을 찾을 수 없습니다.")
            return
        
        sizes = product.get('sizes', ['FREE'])
        record = info['record']
        old_color = info['color']
        old_quantity = info['quantity']
        old_date = info['date']
        old_size = info['size']
        
        # 사이즈 선택 다이얼로그
        dialog = tk.Toplevel(self.root)
        dialog.title("사이즈 변경")
        dialog.geometry("300x400")
        dialog.transient(self.root)
        dialog.grab_set()
        utils.center_window(dialog)
        
        ttk.Label(dialog, text=f"상품: {product['name']}", font=("Arial", 11, "bold")).pack(pady=10)
        ttk.Label(dialog, text=f"현재 사이즈: {info['size']}").pack(pady=5)
        ttk.Label(dialog, text="변경할 사이즈 선택:", font=("Arial", 10)).pack(pady=10)
        
        listbox = tk.Listbox(dialog, height=10, font=("Arial", 11))
        for size in sizes:
            listbox.insert(tk.END, size)
        listbox.pack(fill=tk.BOTH, expand=True, padx=20, pady=5)
        
        # 현재 사이즈 선택
        if info['size'] in sizes:
            listbox.selection_set(sizes.index(info['size']))
        
        def select_size():
            selected = listbox.curselection()
            if not selected:
                messagebox.showwarning("경고", "사이즈를 선택해주세요.")
                return
            
            new_size = sizes[selected[0]]
            record['size'] = new_size
            
            # 재고 이동 기록도 업데이트
            for movement in self.movements:
                movement_color = movement.get('color', '')
                if (str(movement['product_id']) == str(record['product_id']) and 
                    str(movement['date']) == old_date and 
                    movement['type'] == 'in' and
                    movement_color == old_color and 
                    movement['size'] == old_size and 
                    movement['quantity'] == old_quantity):
                    movement['size'] = new_size
                    break
            
            self._refresh_data_shortcuts()
            self.refresh_inbound_list()
            self.refresh_stock_list()
            dialog.destroy()
            messagebox.showinfo("완료", f"사이즈가 '{new_size}'로 변경되었습니다.")
        
        btn_frame = ttk.Frame(dialog)
        btn_frame.pack(pady=10)
        ttk.Button(btn_frame, text="선택", command=select_size, width=10).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="취소", command=dialog.destroy, width=10).pack(side=tk.LEFT, padx=5)
    
    def edit_outbound(self):
        """출고 기록 수정"""
        selected = self.outbound_tree.selection()
        if not selected:
            messagebox.showwarning("경고", "수정할 출고 기록을 선택해주세요.")
            return
        
        item = self.outbound_tree.item(selected[0])
        values = item['values']
        date = values[0]
        product_name = values[1]
        color = values[3] if values[3] != '-' else ''
        size = values[4]
        old_quantity = values[5]
        note = values[6] if len(values) > 6 else ""
        
        # 출고 기록 찾기
        record = None
        for r in self.outbound_records:
            record_color = r.get('color', '')
            if (r['date'] == date and 
                r['product_name'] == product_name and 
                record_color == color and 
                r['size'] == size and 
                r['quantity'] == old_quantity):
                record = r
                break
        
        if not record:
            messagebox.showerror("오류", "출고 기록을 찾을 수 없습니다.")
            return
        
        # 수정 다이얼로그
        dialog = tk.Toplevel(self.root)
        dialog.title("출고 수정")
        dialog.geometry("400x300")
        dialog.transient(self.root)
        dialog.grab_set()
        utils.center_window(dialog)
        
        ttk.Label(dialog, text=f"상품: {product_name}", font=("Arial", 11, "bold")).pack(pady=10)
        ttk.Label(dialog, text=f"색상: {color if color else '-'}", font=("Arial", 10)).pack(pady=5)
        ttk.Label(dialog, text=f"사이즈: {size}", font=("Arial", 10)).pack(pady=5)
        ttk.Label(dialog, text=f"출고일자: {date}", font=("Arial", 10)).pack(pady=5)
        
        ttk.Label(dialog, text="출고 수량 *", font=("Arial", 10, "bold")).pack(pady=(10,0))
        quantity_var = tk.StringVar(value=str(old_quantity))
        quantity_entry = ttk.Entry(dialog, textvariable=quantity_var, width=30)
        quantity_entry.pack(pady=5)
        
        ttk.Label(dialog, text="비고", font=("Arial", 10)).pack(pady=(10,0))
        note_var = tk.StringVar(value=note)
        note_entry = ttk.Entry(dialog, textvariable=note_var, width=30)
        note_entry.pack(pady=5)
        
        def save_changes():
            try:
                new_quantity = int(quantity_var.get())
                if new_quantity <= 0:
                    raise ValueError
            except:
                messagebox.showerror("오류", "올바른 수량을 입력해주세요.")
                return
            
            # 기록 업데이트
            record['quantity'] = new_quantity
            record['note'] = note_var.get()
            
            # 재고 이동 기록도 업데이트
            for movement in self.movements:
                movement_color = movement.get('color', '')
                if (movement['product_id'] == record['product_id'] and 
                    movement['date'] == date and 
                    movement['type'] == 'out' and
                    movement_color == color and 
                    movement['size'] == size and 
                    movement['quantity'] == old_quantity):
                    movement['quantity'] = new_quantity
                    movement['note'] = note_var.get()
                    break
            
            # self.data_manager.save_data()
            # ← 자동저장 제거됨
            self._refresh_data_shortcuts()
            self.refresh_outbound_list()
            self.refresh_products_list()
            self.refresh_stock_list()
            dialog.destroy()
            messagebox.showinfo("완료", "출고 기록이 수정되었습니다.")
        
        button_frame = ttk.Frame(dialog)
        button_frame.pack(pady=20)
        ttk.Button(button_frame, text="취소", command=dialog.destroy).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="저장", command=save_changes).pack(side=tk.LEFT, padx=5)
    
    def _get_selected_outbound_info(self):
        """선택된 출고 기록 정보 가져오기"""
        selected = self.outbound_tree.selection()
        if not selected:
            messagebox.showwarning("경고", "수정할 출고 기록을 선택해주세요.")
            return None
        
        item = self.outbound_tree.item(selected[0])
        values = item['values']
        date = str(values[0])
        product_name = str(values[1])
        product_code = str(values[2]) if values[2] != '-' else ''
        color = str(values[3]) if values[3] != '-' else ''
        size = str(values[4])
        quantity = values[5]
        
        # 상품 찾기 (상품코드가 있으면 상품코드로 우선 검색)
        product = None
        if product_code:
            for p in self.products:
                if p.get('code', '') == product_code:
                    product = p
                    break
        
        if not product:
            for p in self.products:
                if p['name'] == product_name and p.get('code', '') == product_code:
                    product = p
                    break
        
        if not product and not product_code:
            for p in self.products:
                if p['name'] == product_name:
                    product = p
                    break
        
        if not product:
            messagebox.showerror("오류", f"상품을 찾을 수 없습니다.\n상품명: {product_name}\n상품코드: {product_code}")
            return None
        
        # 출고 기록 찾기 (product_id로 검색)
        record = None
        for r in self.outbound_records:
            record_color = r.get('color', '')
            if (str(r['product_id']) == str(product['id']) and
                str(r['date']) == date and 
                record_color == color and 
                r['size'] == size and 
                r['quantity'] == quantity):
                record = r
                break
        
        if not record:
            messagebox.showerror("오류", "출고 기록을 찾을 수 없습니다.")
            return None
        
        return {
            'record': record,
            'product': product,
            'date': date,
            'color': color,
            'size': size,
            'quantity': quantity
        }
    
    def change_outbound_product(self):
        """출고 상품 변경"""
        info = self._get_selected_outbound_info()
        if not info:
            return
        
        record = info['record']
        old_quantity = info['quantity']
        old_color = info['color']
        old_size = info['size']
        old_date = info['date']
        
        # 상품 선택 다이얼로그
        dialog = tk.Toplevel(self.root)
        dialog.title("상품 변경")
        dialog.geometry("500x600")
        dialog.transient(self.root)
        dialog.grab_set()
        utils.center_window(dialog)
        
        ttk.Label(dialog, text="변경할 상품 선택", font=("Arial", 12, "bold")).pack(pady=10)
        
        # 검색
        search_frame = ttk.Frame(dialog)
        search_frame.pack(fill=tk.X, padx=10, pady=5)
        ttk.Label(search_frame, text="검색:").pack(side=tk.LEFT)
        search_var = tk.StringVar()
        search_entry = ttk.Entry(search_frame, textvariable=search_var, width=30)
        search_entry.pack(side=tk.LEFT, padx=5)
        
        # 상품 목록
        columns = ("상품명", "상품코드")
        tree = ttk.Treeview(dialog, columns=columns, show="headings", height=15)
        tree.heading("상품명", text="상품명")
        tree.heading("상품코드", text="상품코드")
        tree.column("상품명", width=250)
        tree.column("상품코드", width=150)
        tree.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        
        def refresh_list():
            tree.delete(*tree.get_children())
            search_text = search_var.get().lower()
            for p in self.products:
                name = p['name'].lower()
                code = p.get('code', '').lower()
                if search_text in name or search_text in code:
                    tree.insert('', tk.END, values=(p['name'], p.get('code', '')))
        
        refresh_list()
        search_var.trace('w', lambda *args: refresh_list())
        
        def select_product():
            selected = tree.selection()
            if not selected:
                messagebox.showwarning("경고", "상품을 선택해주세요.")
                return
            
            new_product_name = tree.item(selected[0])['values'][0]
            new_product = None
            for p in self.products:
                if p['name'] == new_product_name:
                    new_product = p
                    break
            
            if not new_product:
                return
            
            # 출고 기록 업데이트
            record['product_id'] = new_product['id']
            record['product_name'] = new_product['name']
            
            # 색상이 새 상품에 없으면 초기화
            new_colors = new_product.get('colors', [])
            if record.get('color', '') and record.get('color', '') not in new_colors:
                record['color'] = new_colors[0] if new_colors else ''
            
            # 사이즈가 새 상품에 없으면 초기화
            new_sizes = new_product.get('sizes', ['FREE'])
            if record.get('size', 'FREE') not in new_sizes:
                record['size'] = new_sizes[0] if new_sizes else 'FREE'
            
            # 재고 이동 기록도 업데이트
            for movement in self.movements:
                movement_color = movement.get('color', '')
                if (str(movement['product_id']) == str(info['product']['id']) and 
                    str(movement['date']) == old_date and 
                    movement['type'] == 'out' and
                    movement_color == old_color and 
                    movement['size'] == old_size and 
                    movement['quantity'] == old_quantity):
                    movement['product_id'] = new_product['id']
                    if record.get('color', '') != old_color:
                        movement['color'] = record.get('color', '')
                    if record.get('size', 'FREE') != old_size:
                        movement['size'] = record.get('size', 'FREE')
                    break
            
            self._refresh_data_shortcuts()
            self.refresh_outbound_list()
            self.refresh_products_list()
            self.refresh_stock_list()
            dialog.destroy()
            messagebox.showinfo("완료", "상품이 변경되었습니다.")
        
        btn_frame = ttk.Frame(dialog)
        btn_frame.pack(pady=10)
        ttk.Button(btn_frame, text="선택", command=select_product, width=10).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="취소", command=dialog.destroy, width=10).pack(side=tk.LEFT, padx=5)
    
    def change_outbound_color(self):
        """출고 색상 변경"""
        info = self._get_selected_outbound_info()
        if not info:
            return
        
        product = info['product']
        if not product:
            messagebox.showerror("오류", "상품을 찾을 수 없습니다.")
            return
        
        record = info['record']
        old_color = info['color']
        old_quantity = info['quantity']
        old_date = info['date']
        old_size = info['size']
        
        # 같은 상품명을 가진 모든 상품에서 색상 수집
        product_name = product['name']
        color_to_product = {}  # {색상: 상품} 매핑
        all_colors = []
        
        for p in self.products:
            if p['name'] == product_name:
                p_colors = p.get('colors', [])
                for c in p_colors:
                    if c and c not in all_colors:
                        all_colors.append(c)
                        color_to_product[c] = p
        
        if not all_colors:
            messagebox.showinfo("안내", "이 상품은 색상이 없습니다.")
            return
        
        # 색상 선택 다이얼로그
        dialog = tk.Toplevel(self.root)
        dialog.title("색상 변경")
        dialog.geometry("300x400")
        dialog.transient(self.root)
        dialog.grab_set()
        utils.center_window(dialog)
        
        ttk.Label(dialog, text=f"상품: {product_name}", font=("Arial", 11, "bold")).pack(pady=10)
        ttk.Label(dialog, text=f"현재 색상: {old_color if old_color else '-'}").pack(pady=5)
        ttk.Label(dialog, text="변경할 색상 선택:", font=("Arial", 10)).pack(pady=10)
        
        listbox = tk.Listbox(dialog, height=10, font=("Arial", 11))
        for color in all_colors:
            listbox.insert(tk.END, color)
        listbox.pack(fill=tk.BOTH, expand=True, padx=20, pady=5)
        
        # 현재 색상 선택
        if old_color in all_colors:
            listbox.selection_set(all_colors.index(old_color))
        
        def select_color():
            selected = listbox.curselection()
            if not selected:
                messagebox.showwarning("경고", "색상을 선택해주세요.")
                return
            
            new_color = all_colors[selected[0]]
            new_product = color_to_product.get(new_color, product)
            
            # 기록 업데이트
            record['color'] = new_color
            record['product_id'] = new_product['id']
            record['product_name'] = new_product['name']
            record['product_code'] = new_product.get('code', '')
            
            # 재고 이동 기록도 업데이트
            for movement in self.movements:
                movement_color = movement.get('color', '')
                if (str(movement['product_id']) == str(info['product']['id']) and 
                    str(movement['date']) == old_date and 
                    movement['type'] == 'out' and
                    movement_color == old_color and 
                    movement['size'] == old_size and 
                    movement['quantity'] == old_quantity):
                    movement['color'] = new_color
                    movement['product_id'] = new_product['id']
                    break
            
            self._refresh_data_shortcuts()
            self.refresh_outbound_list()
            self.refresh_stock_list()
            dialog.destroy()
            messagebox.showinfo("완료", f"색상이 '{new_color}'로 변경되었습니다.")
        
        btn_frame = ttk.Frame(dialog)
        btn_frame.pack(pady=10)
        ttk.Button(btn_frame, text="선택", command=select_color, width=10).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="취소", command=dialog.destroy, width=10).pack(side=tk.LEFT, padx=5)
    
    def change_outbound_size(self):
        """출고 사이즈 변경"""
        info = self._get_selected_outbound_info()
        if not info:
            return
        
        product = info['product']
        if not product:
            messagebox.showerror("오류", "상품을 찾을 수 없습니다.")
            return
        
        sizes = product.get('sizes', ['FREE'])
        record = info['record']
        old_color = info['color']
        old_quantity = info['quantity']
        old_date = info['date']
        old_size = info['size']
        
        # 사이즈 선택 다이얼로그
        dialog = tk.Toplevel(self.root)
        dialog.title("사이즈 변경")
        dialog.geometry("300x400")
        dialog.transient(self.root)
        dialog.grab_set()
        utils.center_window(dialog)
        
        ttk.Label(dialog, text=f"상품: {product['name']}", font=("Arial", 11, "bold")).pack(pady=10)
        ttk.Label(dialog, text=f"현재 사이즈: {info['size']}").pack(pady=5)
        ttk.Label(dialog, text="변경할 사이즈 선택:", font=("Arial", 10)).pack(pady=10)
        
        listbox = tk.Listbox(dialog, height=10, font=("Arial", 11))
        for size in sizes:
            listbox.insert(tk.END, size)
        listbox.pack(fill=tk.BOTH, expand=True, padx=20, pady=5)
        
        # 현재 사이즈 선택
        if info['size'] in sizes:
            listbox.selection_set(sizes.index(info['size']))
        
        def select_size():
            selected = listbox.curselection()
            if not selected:
                messagebox.showwarning("경고", "사이즈를 선택해주세요.")
                return
            
            new_size = sizes[selected[0]]
            record['size'] = new_size
            
            # 재고 이동 기록도 업데이트
            for movement in self.movements:
                movement_color = movement.get('color', '')
                if (str(movement['product_id']) == str(record['product_id']) and 
                    str(movement['date']) == old_date and 
                    movement['type'] == 'out' and
                    movement_color == old_color and 
                    movement['size'] == old_size and 
                    movement['quantity'] == old_quantity):
                    movement['size'] = new_size
                    break
            
            self._refresh_data_shortcuts()
            self.refresh_outbound_list()
            self.refresh_stock_list()
            dialog.destroy()
            messagebox.showinfo("완료", f"사이즈가 '{new_size}'로 변경되었습니다.")
        
        btn_frame = ttk.Frame(dialog)
        btn_frame.pack(pady=10)
        ttk.Button(btn_frame, text="선택", command=select_size, width=10).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="취소", command=dialog.destroy, width=10).pack(side=tk.LEFT, padx=5)
    
    def show_inbound_menu(self, event):
        """입고 우클릭 메뉴 표시"""
        try:
            self.inbound_tree.selection_set(self.inbound_tree.identify_row(event.y))
            self.inbound_menu.post(event.x_root, event.y_root)
        finally:
            self.inbound_menu.grab_release()
    
    def show_outbound_menu(self, event):
        """출고 우클릭 메뉴 표시"""
        try:
            self.outbound_tree.selection_set(self.outbound_tree.identify_row(event.y))
            self.outbound_menu.post(event.x_root, event.y_root)
        finally:
            self.outbound_menu.grab_release()
    
    def update_inbound_date_filter(self):
        """입고 날짜 필터 업데이트"""
        dates = sorted(set(r['date'] for r in self.inbound_records), reverse=True)
        self.inbound_date_combo['values'] = ['전체'] + dates
    
    def update_outbound_date_filter(self):
        """출고 날짜 필터 업데이트"""
        dates = sorted(set(r['date'] for r in self.outbound_records), reverse=True)
        self.outbound_date_combo['values'] = ['전체'] + dates
    
    def refresh_inbound_list(self):
        """입고 목록 새로고침"""
        for item in self.inbound_tree.get_children():
            self.inbound_tree.delete(item)
        
        selected_date = self.inbound_date_filter_var.get()
        sorted_records = sorted(self.inbound_records, key=lambda x: x['date'], reverse=True)
        
        if selected_date != '전체':
            sorted_records = [r for r in sorted_records if r['date'] == selected_date]
        
        # 같은 날짜, 상품, 색상, 사이즈를 합치기
        merged_records = {}
        for record in sorted_records:
            key = (record['date'], record['product_id'], record['color'], record['size'])
            
            if key in merged_records:
                merged_records[key]['quantity'] += record['quantity']
                # 비고는 첫 번째 것만 사용
            else:
                merged_records[key] = {
                    'date': record['date'],
                    'product_name': record['product_name'],
                    'product_code': record['product_code'],
                    'color': record['color'],
                    'size': record['size'],
                    'quantity': record['quantity'],
                    'note': record.get('note', '')
                }
        
        # 병합된 데이터로 표시
        for key, merged in merged_records.items():
            self.inbound_tree.insert('', tk.END, values=(
                merged['date'],
                merged['product_name'],
                merged['product_code'],
                merged['color'] if merged['color'] else '-',
                merged['size'],
                merged['quantity'],
                merged.get('note', '')
            ))
    
    def refresh_outbound_list(self):
        """출고 목록 새로고침"""
        for item in self.outbound_tree.get_children():
            self.outbound_tree.delete(item)
        
        selected_date = self.outbound_date_filter_var.get()
        sorted_records = sorted(self.outbound_records, key=lambda x: x['date'], reverse=True)
        
        if selected_date != '전체':
            sorted_records = [r for r in sorted_records if r['date'] == selected_date]
        
        # 같은 날짜, 상품, 색상, 사이즈를 합치기
        merged_records = {}
        for record in sorted_records:
            key = (record['date'], record['product_id'], record['color'], record['size'])
            
            if key in merged_records:
                merged_records[key]['quantity'] += record['quantity']
                # 비고는 첫 번째 것만 사용
            else:
                merged_records[key] = {
                    'date': record['date'],
                    'product_name': record['product_name'],
                    'product_code': record['product_code'],
                    'color': record['color'],
                    'size': record['size'],
                    'quantity': record['quantity'],
                    'note': record.get('note', '')
                }
        
        # 병합된 데이터로 표시
        for key, merged in merged_records.items():
            self.outbound_tree.insert('', tk.END, values=(
                merged['date'],
                merged['product_name'],
                merged['product_code'],
                merged['color'] if merged['color'] else '-',
                merged['size'],
                merged['quantity'],
                merged.get('note', '')
            ))

    def search_inbound_detail(self):
        """입고 상세 검색"""
        for item in self.inbound_tree.get_children():
            self.inbound_tree.delete(item)

        # 필터 조건 가져오기
        start_date = self.inbound_start_date.get_date().strftime('%Y-%m-%d')
        end_date = self.inbound_end_date.get_date().strftime('%Y-%m-%d')
        product_search = self.inbound_product_search_var.get().lower().strip()

        # 기간 내 데이터 필터링
        filtered_records = [r for r in self.inbound_records
                          if start_date <= r['date'] <= end_date]

        # 상품명/코드 필터링
        if product_search:
            filtered_records = [r for r in filtered_records
                              if product_search in r.get('product_name', '').lower()
                              or product_search in r.get('product_code', '').lower()]

        sorted_records = sorted(filtered_records, key=lambda x: x['date'], reverse=True)

        # 같은 날짜, 상품, 색상, 사이즈를 합치기
        merged_records = {}
        for record in sorted_records:
            key = (record['date'], record['product_id'], record['color'], record['size'])

            if key in merged_records:
                merged_records[key]['quantity'] += record['quantity']
            else:
                merged_records[key] = {
                    'date': record['date'],
                    'product_name': record['product_name'],
                    'product_code': record['product_code'],
                    'color': record['color'],
                    'size': record['size'],
                    'quantity': record['quantity'],
                    'note': record.get('note', '')
                }

        # 병합된 데이터로 표시
        for key, merged in merged_records.items():
            self.inbound_tree.insert('', tk.END, values=(
                merged['date'],
                merged['product_name'],
                merged['product_code'],
                merged['color'] if merged['color'] else '-',
                merged['size'],
                merged['quantity'],
                merged.get('note', '')
            ))

        messagebox.showinfo("검색 완료", f"{start_date} ~ {end_date} 기간 내 {len(merged_records)}건 검색됨")

    def reset_inbound_filter(self):
        """입고 필터 초기화"""
        self.inbound_date_filter_var.set("전체")
        self.inbound_product_search_var.set("")
        self.inbound_start_date.set_date(datetime.now())
        self.inbound_end_date.set_date(datetime.now())
        self.refresh_inbound_list()

    def search_outbound_detail(self):
        """출고 상세 검색"""
        for item in self.outbound_tree.get_children():
            self.outbound_tree.delete(item)

        # 필터 조건 가져오기
        start_date = self.outbound_start_date.get_date().strftime('%Y-%m-%d')
        end_date = self.outbound_end_date.get_date().strftime('%Y-%m-%d')
        product_search = self.outbound_product_search_var.get().lower().strip()

        # 기간 내 데이터 필터링
        filtered_records = [r for r in self.outbound_records
                          if start_date <= r['date'] <= end_date]

        # 상품명/코드 필터링
        if product_search:
            filtered_records = [r for r in filtered_records
                              if product_search in r.get('product_name', '').lower()
                              or product_search in r.get('product_code', '').lower()]

        sorted_records = sorted(filtered_records, key=lambda x: x['date'], reverse=True)

        # 같은 날짜, 상품, 색상, 사이즈를 합치기
        merged_records = {}
        for record in sorted_records:
            key = (record['date'], record['product_id'], record['color'], record['size'])

            if key in merged_records:
                merged_records[key]['quantity'] += record['quantity']
            else:
                merged_records[key] = {
                    'date': record['date'],
                    'product_name': record['product_name'],
                    'product_code': record['product_code'],
                    'color': record['color'],
                    'size': record['size'],
                    'quantity': record['quantity'],
                    'note': record.get('note', '')
                }

        # 병합된 데이터로 표시
        for key, merged in merged_records.items():
            self.outbound_tree.insert('', tk.END, values=(
                merged['date'],
                merged['product_name'],
                merged['product_code'],
                merged['color'] if merged['color'] else '-',
                merged['size'],
                merged['quantity'],
                merged.get('note', '')
            ))

        messagebox.showinfo("검색 완료", f"{start_date} ~ {end_date} 기간 내 {len(merged_records)}건 검색됨")

    def reset_outbound_filter(self):
        """출고 필터 초기화"""
        self.outbound_date_filter_var.set("전체")
        self.outbound_product_search_var.set("")
        self.outbound_start_date.set_date(datetime.now())
        self.outbound_end_date.set_date(datetime.now())
        self.refresh_outbound_list()

    def show_stock_menu(self, event):
        """재고현황 우클릭 메뉴 표시"""
        try:
            self.stock_tree.selection_set(self.stock_tree.identify_row(event.y))
            self.stock_menu.post(event.x_root, event.y_root)
        finally:
            self.stock_menu.grab_release()
    
    def on_stock_double_click(self, event):
        """재고현황 더블클릭 시 수량 편집"""
        region = self.stock_tree.identify("region", event.x, event.y)
        if region != "cell":
            return
        
        column = self.stock_tree.identify_column(event.x)
        col_index = int(column.replace('#', '')) - 1
        columns = self.stock_tree['columns']
        col_name = columns[col_index]
        
        # 편집 가능한 컬럼: 현재고, 총입고, 총출고, 미입고
        editable_columns = ["현재고", "총입고", "총출고", "미입고"]
        if col_name not in editable_columns:
            return
        
        item_id = self.stock_tree.identify_row(event.y)
        if not item_id:
            return
        
        item = self.stock_tree.item(item_id)
        values = list(item['values'])
        current_value = values[col_index]
        
        # 상품 정보 가져오기
        product_name = str(values[1])
        product_code = str(values[2]) if values[2] != '-' else ''
        color = str(values[3]) if values[3] != '-' else ''
        size = str(values[4])
        
        # 상품 찾기
        product = None
        if product_code:
            for p in self.products:
                if p.get('code', '') == product_code:
                    product = p
                    break
        if not product:
            for p in self.products:
                if p['name'] == product_name:
                    product = p
                    break
        
        if not product:
            return
        
        # 셀 위치 계산
        bbox = self.stock_tree.bbox(item_id, column)
        if not bbox:
            return
        
        x, y, width, height = bbox
        
        # Entry 생성
        entry = ttk.Entry(self.stock_tree, width=10)
        entry.place(x=x, y=y, width=width, height=height)
        entry.insert(0, str(current_value))
        entry.select_range(0, tk.END)
        entry.focus()
        
        def save_edit(event=None):
            try:
                new_value = int(entry.get())
                if new_value < 0:
                    raise ValueError
            except:
                entry.destroy()
                return
            
            old_value = int(current_value)
            diff = new_value - old_value
            
            if diff == 0:
                entry.destroy()
                return
            
            # 컬럼에 따라 다르게 처리
            if col_name == "현재고":
                # 현재고 변경 = 입고 또는 출고 기록 추가
                if diff > 0:
                    # 재고 증가 = 입고 추가
                    self._add_stock_adjustment(product, color, size, diff, 'in')
                else:
                    # 재고 감소 = 출고 추가
                    self._add_stock_adjustment(product, color, size, abs(diff), 'out')
            elif col_name == "총입고":
                # 총입고 변경
                if diff > 0:
                    self._add_stock_adjustment(product, color, size, diff, 'in')
                else:
                    messagebox.showwarning("경고", "총입고는 줄일 수 없습니다.\n개별 입고 기록을 수정해주세요.")
                    entry.destroy()
                    return
            elif col_name == "총출고":
                # 총출고 변경
                if diff > 0:
                    self._add_stock_adjustment(product, color, size, diff, 'out')
                else:
                    messagebox.showwarning("경고", "총출고는 줄일 수 없습니다.\n개별 출고 기록을 수정해주세요.")
                    entry.destroy()
                    return
            elif col_name == "미입고":
                # 미입고 변경 = 발주 수량 조정
                if diff != 0:
                    self._adjust_pending_orders(product, color, size, diff)
            
            entry.destroy()
            self.refresh_stock_list()
            self.refresh_inbound_list()
            self.refresh_outbound_list()
            self.refresh_orders_list()
        
        def cancel_edit(event=None):
            entry.destroy()
        
        entry.bind("<Return>", save_edit)
        entry.bind("<Escape>", cancel_edit)
        entry.bind("<FocusOut>", cancel_edit)
    
    def _add_stock_adjustment(self, product, color, size, quantity, movement_type):
        """재고 조정을 위한 입고/출고 기록 추가"""
        today = datetime.now().strftime('%Y-%m-%d')
        
        if movement_type == 'in':
            # 입고 기록 추가
            record = {
                'date': today,
                'product_id': product['id'],
                'product_name': product['name'],
                'product_code': product.get('code', ''),
                'color': color,
                'size': size,
                'quantity': quantity,
                'note': '재고 조정'
            }
            # DB에 저장
            self.data_manager.add_inbound(record)
            
            # 재고 이동 기록
            movement = {
                'date': today,
                'product_id': product['id'],
                'color': color,
                'size': size,
                'quantity': quantity,
                'from_location': '조정',
                'to_location': '창고',
                'type': 'in',
                'note': '재고 조정'
            }
            # DB에 저장
            self.data_manager.add_movement(movement)
        else:
            # 출고 기록 추가
            record = {
                'date': today,
                'product_id': product['id'],
                'product_name': product['name'],
                'product_code': product.get('code', ''),
                'color': color,
                'size': size,
                'quantity': quantity,
                'note': '재고 조정'
            }
            # DB에 저장
            self.data_manager.add_outbound(record)
            
            # 재고 이동 기록
            movement = {
                'date': today,
                'product_id': product['id'],
                'color': color,
                'size': size,
                'quantity': quantity,
                'from_location': '창고',
                'to_location': '조정',
                'type': 'out',
                'note': '재고 조정'
            }
            # DB에 저장
            self.data_manager.add_movement(movement)
        
        self._refresh_data_shortcuts()
    
    def _adjust_pending_orders(self, product, color, size, diff):
        """미입고 수량 조정 (발주 추가 또는 입고 처리)"""
        today = datetime.now().strftime('%Y-%m-%d')
        
        if diff > 0:
            # 미입고 증가 = 새 발주 추가
            order_id = max([o.get('id', 0) for o in self.orders], default=0) + 1
            order = {
                'id': order_id,
                'date': today,
                'product_id': product['id'],
                'color': color,
                'size': size,
                'quantity': diff,
                'shipped_quantity': 0,
                'status': 'pending',
                'store_id': '',
                'note': '재고 조정'
            }
            self.orders.append(order)
        else:
            # 미입고 감소 = 기존 발주에 입고 처리
            remaining = abs(diff)
            for order in self.orders:
                if (str(order['product_id']) == str(product['id']) and
                    order.get('color', '') == color and
                    order.get('size', 'FREE') == size and
                    order.get('status', 'pending') == 'pending'):
                    
                    pending = order['quantity'] - order.get('shipped_quantity', 0)
                    if pending > 0:
                        to_ship = min(pending, remaining)
                        order['shipped_quantity'] = order.get('shipped_quantity', 0) + to_ship
                        if order['shipped_quantity'] >= order['quantity']:
                            order['status'] = 'completed'
                        remaining -= to_ship
                        if remaining <= 0:
                            break
        
        self._refresh_data_shortcuts()
    
    def _get_selected_stock_info(self):
        """선택된 재고현황 정보 가져오기"""
        selected = self.stock_tree.selection()
        if not selected:
            messagebox.showwarning("경고", "수정할 항목을 선택해주세요.")
            return None
        
        item = self.stock_tree.item(selected[0])
        values = item['values']
        # columns = ("매장", "상품명", "상품코드", "색상", "사이즈", "현재고", "총입고", "총출고", "미입고")
        store_name = str(values[0]) if values[0] != '-' else ''
        product_name = str(values[1])
        product_code = str(values[2]) if values[2] != '-' else ''
        color = str(values[3]) if values[3] != '-' else ''
        size = str(values[4])
        
        # 상품 찾기 (상품코드가 있으면 상품코드로 우선 검색)
        product = None
        if product_code:
            for p in self.products:
                if p.get('code', '') == product_code:
                    product = p
                    break
        
        if not product:
            for p in self.products:
                if p['name'] == product_name and p.get('code', '') == product_code:
                    product = p
                    break
        
        if not product and not product_code:
            for p in self.products:
                if p['name'] == product_name:
                    product = p
                    break
        
        if not product:
            messagebox.showerror("오류", f"상품을 찾을 수 없습니다.\n상품명: {product_name}\n상품코드: {product_code}")
            return None
        
        # 매장 ID 찾기
        store_id = ''
        for s in self.stores:
            if s['name'] == store_name:
                store_id = s.get('id', '')
                break
        
        return {
            'product': product,
            'store_name': store_name,
            'store_id': store_id,
            'color': color,
            'size': size
        }
    
    def change_stock_color(self):
        """재고현황 색상 변경 (관련 입고/출고/발주 모두 변경)"""
        info = self._get_selected_stock_info()
        if not info:
            return
        
        product = info['product']
        old_color = info['color']
        old_size = info['size']
        
        # 같은 상품명을 가진 모든 상품에서 색상 수집
        product_name = product['name']
        color_to_product = {}  # {색상: 상품} 매핑
        all_colors = []
        
        for p in self.products:
            if p['name'] == product_name:
                p_colors = p.get('colors', [])
                for c in p_colors:
                    if c and c not in all_colors:
                        all_colors.append(c)
                        color_to_product[c] = p
        
        if not all_colors:
            messagebox.showinfo("안내", "이 상품은 색상이 없습니다.")
            return
        
        # 색상 선택 다이얼로그
        dialog = tk.Toplevel(self.root)
        dialog.title("색상 변경")
        dialog.geometry("350x450")
        dialog.transient(self.root)
        dialog.grab_set()
        utils.center_window(dialog)
        
        ttk.Label(dialog, text=f"상품: {product_name}", font=("Arial", 11, "bold")).pack(pady=10)
        ttk.Label(dialog, text=f"현재 색상: {old_color if old_color else '-'}").pack(pady=5)
        ttk.Label(dialog, text="⚠️ 관련된 모든 입고/출고/발주 기록의\n색상이 함께 변경됩니다!", 
                  foreground="red", font=("Arial", 9)).pack(pady=5)
        ttk.Label(dialog, text="변경할 색상 선택:", font=("Arial", 10)).pack(pady=10)
        
        listbox = tk.Listbox(dialog, height=10, font=("Arial", 11))
        for color in all_colors:
            listbox.insert(tk.END, color)
        listbox.pack(fill=tk.BOTH, expand=True, padx=20, pady=5)
        
        if old_color in all_colors:
            listbox.selection_set(all_colors.index(old_color))
        
        def select_color():
            selected = listbox.curselection()
            if not selected:
                messagebox.showwarning("경고", "색상을 선택해주세요.")
                return
            
            new_color = all_colors[selected[0]]
            new_product = color_to_product.get(new_color, product)
            
            if new_color == old_color:
                dialog.destroy()
                return
            
            changed_count = 0
            
            # 입고 기록 변경
            for r in self.inbound_records:
                if (str(r['product_id']) == str(product['id']) and 
                    r.get('color', '') == old_color and
                    r.get('size', 'FREE') == old_size):
                    r['color'] = new_color
                    r['product_id'] = new_product['id']
                    r['product_name'] = new_product['name']
                    r['product_code'] = new_product.get('code', '')
                    changed_count += 1
            
            # 출고 기록 변경
            for r in self.outbound_records:
                if (str(r['product_id']) == str(product['id']) and 
                    r.get('color', '') == old_color and
                    r.get('size', 'FREE') == old_size):
                    r['color'] = new_color
                    r['product_id'] = new_product['id']
                    r['product_name'] = new_product['name']
                    r['product_code'] = new_product.get('code', '')
                    changed_count += 1
            
            # 재고 이동 변경
            for m in self.movements:
                if (str(m['product_id']) == str(product['id']) and 
                    m.get('color', '') == old_color and
                    m.get('size', 'FREE') == old_size):
                    m['color'] = new_color
                    m['product_id'] = new_product['id']
                    changed_count += 1
            
            # 발주 변경
            for o in self.orders:
                if (str(o['product_id']) == str(product['id']) and 
                    o.get('color', '') == old_color and
                    o.get('size', 'FREE') == old_size):
                    o['color'] = new_color
                    o['product_id'] = new_product['id']
                    changed_count += 1
            
            self._refresh_data_shortcuts()
            self.refresh_stock_list()
            self.refresh_inbound_list()
            self.refresh_outbound_list()
            self.refresh_orders_list()
            dialog.destroy()
            messagebox.showinfo("완료", f"색상이 '{new_color}'로 변경되었습니다.\n({changed_count}건 변경)")
        
        btn_frame = ttk.Frame(dialog)
        btn_frame.pack(pady=10)
        ttk.Button(btn_frame, text="선택", command=select_color, width=10).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="취소", command=dialog.destroy, width=10).pack(side=tk.LEFT, padx=5)
    
    def change_stock_size(self):
        """재고현황 사이즈 변경 (관련 입고/출고/발주 모두 변경)"""
        info = self._get_selected_stock_info()
        if not info:
            return
        
        product = info['product']
        sizes = product.get('sizes', ['FREE'])
        
        old_color = info['color']
        old_size = info['size']
        
        # 사이즈 선택 다이얼로그
        dialog = tk.Toplevel(self.root)
        dialog.title("사이즈 변경")
        dialog.geometry("350x450")
        dialog.transient(self.root)
        dialog.grab_set()
        utils.center_window(dialog)
        
        ttk.Label(dialog, text=f"상품: {product['name']}", font=("Arial", 11, "bold")).pack(pady=10)
        ttk.Label(dialog, text=f"현재 사이즈: {old_size}").pack(pady=5)
        ttk.Label(dialog, text="⚠️ 관련된 모든 입고/출고/발주 기록의\n사이즈가 함께 변경됩니다!", 
                  foreground="red", font=("Arial", 9)).pack(pady=5)
        ttk.Label(dialog, text="변경할 사이즈 선택:", font=("Arial", 10)).pack(pady=10)
        
        listbox = tk.Listbox(dialog, height=10, font=("Arial", 11))
        for size in sizes:
            listbox.insert(tk.END, size)
        listbox.pack(fill=tk.BOTH, expand=True, padx=20, pady=5)
        
        if old_size in sizes:
            listbox.selection_set(sizes.index(old_size))
        
        def select_size():
            selected = listbox.curselection()
            if not selected:
                messagebox.showwarning("경고", "사이즈를 선택해주세요.")
                return
            
            new_size = sizes[selected[0]]
            if new_size == old_size:
                dialog.destroy()
                return
            
            changed_count = 0
            
            # 입고 기록 변경
            for r in self.inbound_records:
                if (str(r['product_id']) == str(product['id']) and 
                    r.get('color', '') == old_color and
                    r.get('size', 'FREE') == old_size):
                    r['size'] = new_size
                    changed_count += 1
            
            # 출고 기록 변경
            for r in self.outbound_records:
                if (str(r['product_id']) == str(product['id']) and 
                    r.get('color', '') == old_color and
                    r.get('size', 'FREE') == old_size):
                    r['size'] = new_size
                    changed_count += 1
            
            # 재고 이동 변경
            for m in self.movements:
                if (str(m['product_id']) == str(product['id']) and 
                    m.get('color', '') == old_color and
                    m.get('size', 'FREE') == old_size):
                    m['size'] = new_size
                    changed_count += 1
            
            # 발주 변경
            for o in self.orders:
                if (str(o['product_id']) == str(product['id']) and 
                    o.get('color', '') == old_color and
                    o.get('size', 'FREE') == old_size):
                    o['size'] = new_size
                    changed_count += 1
            
            self._refresh_data_shortcuts()
            self.refresh_stock_list()
            self.refresh_inbound_list()
            self.refresh_outbound_list()
            self.refresh_orders_list()
            dialog.destroy()
            messagebox.showinfo("완료", f"사이즈가 '{new_size}'로 변경되었습니다.\n({changed_count}건 변경)")
        
        btn_frame = ttk.Frame(dialog)
        btn_frame.pack(pady=10)
        ttk.Button(btn_frame, text="선택", command=select_size, width=10).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="취소", command=dialog.destroy, width=10).pack(side=tk.LEFT, padx=5)

    def on_stock_select(self, event=None):
        """재고 현황에서 상품 선택 시 이미지 표시"""
        selected = self.stock_tree.selection()
        if not selected:
            self.stock_image_label.config(image='', text="이미지 없음")
            return
        
        item = self.stock_tree.item(selected[0])
        # columns = ("매장", "상품명", "상품코드", "색상", "사이즈", ...)
        product_name = item['values'][1]
        
        product = None
        for p in self.products:
            if p['name'] == product_name:
                product = p
                break
        
        if not product or not product.get('image'):
            self.stock_image_label.config(image='', text="이미지 없음")
            return
        
        try:
            img_data = base64.b64decode(product['image'])
            img = Image.open(BytesIO(img_data))
            img.thumbnail((140, 140))
            photo = ImageTk.PhotoImage(img)
            self.stock_image_label.config(image=photo, text='')
            self.stock_image_label.image = photo  # 참조 유지
        except Exception as e:
            self.stock_image_label.config(image='', text="이미지 오류")
    
    def show_product_detail(self, event=None):
        selected = self.products_tree.selection()
        if not selected:
            messagebox.showwarning("경고", "상품을 선택해주세요.")
            return
        
        item = self.products_tree.item(selected[0])
        product_name = item['values'][0]
        product_code = item['values'][1]
        
        # 상품명과 상품코드로 정확한 상품 찾기
        product = None
        for p in self.products:
            if p['name'] == product_name and p.get('code', '') == product_code:
                product = p
                break
        
        if not product:
            return
        
        dialog = tk.Toplevel(self.root)
        dialog.title(f"상품 상세 - {product['name']}")
        dialog.geometry("700x600")
        dialog.transient(self.root)
        dialog.grab_set()
        utils.center_window(dialog)
        
        info_frame = ttk.Frame(dialog)
        info_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        if product.get('image'):
            try:
                img_data = base64.b64decode(product['image'])
                img = Image.open(BytesIO(img_data))
                img.thumbnail((200, 200))
                photo = ImageTk.PhotoImage(img)
                img_label = ttk.Label(info_frame, image=photo)
                img_label.image = photo
                img_label.pack(pady=10)
            except:
                pass
        
        ttk.Label(info_frame, text=f"상품명: {product['name']}", font=("Arial", 12, "bold")).pack(pady=5)
        ttk.Label(info_frame, text=f"상품코드: {product.get('code', '-')}").pack()
        ttk.Label(info_frame, text=f"매입처: {product.get('supplier', '-')}").pack()
        
        # 메모 표시
        if product.get('memo'):
            memo_frame = ttk.LabelFrame(info_frame, text="📝 메모", padding=10)
            memo_frame.pack(pady=10, fill=tk.X)
            ttk.Label(memo_frame, text=product.get('memo'), wraplength=600, justify=tk.LEFT).pack()
        
        ttk.Label(info_frame, text="색상/사이즈별 재고:", font=("Arial", 10, "bold")).pack(pady=(10,5))
        
        tree_frame = ttk.Frame(info_frame)
        tree_frame.pack(fill=tk.BOTH, expand=True, pady=5)
        
        columns = ("색상", "사이즈", "현재고", "미입고")
        detail_tree = ttk.Treeview(tree_frame, columns=columns, show="headings", height=10)
        
        for col in columns:
            detail_tree.heading(col, text=col)
            detail_tree.column(col, width=150, anchor='center')
        
        scrollbar = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=detail_tree.yview)
        detail_tree.configure(yscrollcommand=scrollbar.set)
        
        detail_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        colors = product.get('colors', [''])
        sizes = product.get('sizes', ['FREE'])
        
        for color in colors:
            for size in sizes:
                stock = self.data_manager.calculate_stock_by_variant(product['id'], color, size)
                pending = self.data_manager.calculate_pending_by_variant(product['id'], color, size)
                
                detail_tree.insert('', tk.END, values=(
                    color if color else '-',
                    size,
                    stock,
                    pending
                ))
        
        ttk.Button(info_frame, text="닫기", command=dialog.destroy).pack(pady=10)
    
    def add_product(self):
        dialog = tk.Toplevel(self.root)
        dialog.title("상품 추가")
        dialog.geometry("700x1000")
        dialog.transient(self.root)
        dialog.grab_set()
        utils.center_window(dialog)

        # 스크롤 가능한 프레임 생성
        content_frame = utils.make_scrollable_dialog(dialog, max_height=900)

        ttk.Label(content_frame, text="상품명:", font=("Arial", 10)).pack(pady=(20,5))
        name_var = tk.StringVar()
        ttk.Entry(content_frame, textvariable=name_var, width=50, justify='center').pack(pady=5)

        # 상품코드와 체크박스
        code_frame = ttk.Frame(content_frame)
        code_frame.pack(pady=(10,5))

        ttk.Label(code_frame, text="상품코드:", font=("Arial", 10)).pack(side=tk.LEFT, padx=(0,10))

        code_var = tk.StringVar()
        code_entry = ttk.Entry(code_frame, textvariable=code_var, width=40, justify='center')
        code_entry.pack(side=tk.LEFT, padx=5)

        # 자동 분리 체크박스
        auto_split_var = tk.BooleanVar(value=self.data_manager.get_auto_split_setting())
        auto_split_check = ttk.Checkbutton(
            code_frame,
            text="자동분리",
            variable=auto_split_var,
            command=lambda: self.data_manager.set_auto_split_setting(auto_split_var.get())
        )
        auto_split_check.pack(side=tk.LEFT, padx=5)

        ttk.Label(content_frame, text="매입처:", font=("Arial", 10)).pack(pady=(10,5))
        supplier_var = tk.StringVar()
        ttk.Entry(content_frame, textvariable=supplier_var, width=50, justify='center').pack(pady=5)

        ttk.Label(content_frame, text=f"{self.field_names[0]['name'] if len(self.field_names) > 0 else '색상'} (쉼표로 구분):", font=("Arial", 10)).pack(pady=(10,5))
        colors_var = tk.StringVar()
        ttk.Entry(content_frame, textvariable=colors_var, width=50, justify='center').pack(pady=5)
        ttk.Label(content_frame, text=f"예: 빨강, 파랑, 초록 (비워두면 없음)", font=("Arial", 8), foreground="gray").pack()

        ttk.Label(content_frame, text=f"{self.field_names[1]['name'] if len(self.field_names) > 1 else '사이즈'} (쉼표로 구분):", font=("Arial", 10)).pack(pady=(10,5))
        sizes_var = tk.StringVar(value="FREE")
        ttk.Entry(content_frame, textvariable=sizes_var, width=50, justify='center').pack(pady=5)
        ttk.Label(content_frame, text=f"예: S, M, L, XL (비워두면 FREE)", font=("Arial", 8), foreground="gray").pack()

        # 메모 필드 추가
        ttk.Label(content_frame, text="주문 관련 메모:", font=("Arial", 10)).pack(pady=(10,5))
        memo_text = tk.Text(content_frame, width=50, height=4, wrap=tk.WORD)
        memo_text.pack(pady=5, padx=20)
        ttk.Label(content_frame, text="예: 단골 고객 선호 상품, 계절 상품, 특별 주문 사항 등",
                 font=("Arial", 8), foreground="gray").pack()

        # 오더 수량 단위 필드 추가
        ttk.Label(content_frame, text="오더 수량 단위 (선택사항):", font=("Arial", 10)).pack(pady=(10,5))
        order_unit_var = tk.StringVar(value="")
        ttk.Entry(content_frame, textvariable=order_unit_var, width=50, justify='center').pack(pady=5)
        ttk.Label(content_frame, text="예: 40 (40장 단위로 발주해야 하는 경우, 비워두면 단위 없음)",
                 font=("Arial", 8), foreground="gray").pack()

        # 이미지 관련 변수
        image_data = tk.StringVar()
        image_source = tk.StringVar(value="none")
        search_status = tk.StringVar(value="이미지: 아직 선택되지 않음")

        ttk.Label(content_frame, text="━━━━━ 이미지 설정 ━━━━━", font=("Arial", 10, "bold")).pack(pady=(15, 5))
        ttk.Label(content_frame, textvariable=search_status, font=("Arial", 9), foreground="blue").pack(pady=5)

        # 이미지 URL 입력 필드
        ttk.Label(content_frame, text="이미지 URL (또는 아래 버튼 사용):", font=("Arial", 9)).pack(pady=(10, 5))
        url_var = tk.StringVar()
        url_entry = ttk.Entry(content_frame, textvariable=url_var, width=50, justify='center')
        url_entry.pack(pady=5)
        
        # 내부 함수들 (버튼 정의 전에)
        def download_from_url():
            """URL에서 이미지 다운로드"""
            url = url_var.get().strip()
            if not url:
                messagebox.showwarning("경고", "이미지 URL을 입력해주세요.")
                return
            
            if not url.startswith('http'):
                messagebox.showwarning("경고", "올바른 URL을 입력해주세요. (http:// 또는 https://로 시작)")
                return
            
            search_status.set("⏳ 이미지 다운로드 중...")
            dialog.update()
            
            try:
                img_base64 = utils.download_image_from_url(url)
                
                if img_base64:
                    image_data.set(img_base64)
                    image_source.set("url")
                    search_status.set(f"✅ 이미지 다운로드 완료")
                    messagebox.showinfo("완료", "이미지가 다운로드되었습니다!")
                else:
                    search_status.set("❌ 이미지 다운로드 실패")
                    messagebox.showerror("오류", "이미지를 다운로드할 수 없습니다.\n URL을 확인해주세요.")
            except Exception as e:
                search_status.set("❌ 오류 발생")
                messagebox.showerror("오류", f"오류가 발생했습니다:\n{str(e)}")
        
        def get_from_clipboard():
            """클립보드에서 URL 가져오기"""
            try:
                import tkinter as tk_clip
                temp_root = tk_clip.Tk()
                temp_root.withdraw()
                clipboard_url = temp_root.clipboard_get()
                temp_root.destroy()
                
                if not clipboard_url.startswith('http'):
                    messagebox.showwarning("경고", "클립보드에 유효한 URL이 없습니다.")
                    return
                
                url_var.set(clipboard_url)
                messagebox.showinfo("완료", f"URL이 클립보드에서 로드되었습니다:\n{clipboard_url[:60]}...")
            except Exception as e:
                messagebox.showerror("오류", f"클립보드에서 URL을 가져올 수 없습니다:\n{str(e)}")
        
        def select_image_manual():
            """수동으로 이미지 파일 선택"""
            filename = filedialog.askopenfilename(
                title="이미지 선택",
                filetypes=[("Image files", "*.png *.jpg *.jpeg *.gif *.webp *.bmp"), ("All files", "*.*")]
            )
            if filename:
                try:
                    with open(filename, 'rb') as f:
                        image_data.set(base64.b64encode(f.read()).decode())
                    image_source.set("manual")
                    search_status.set(f"✅ 이미지 선택됨: {os.path.basename(filename)}")
                    messagebox.showinfo("완료", "이미지가 선택되었습니다.")
                except Exception as e:
                    messagebox.showerror("오류", f"이미지를 읽을 수 없습니다: {str(e)}")
        


        # 버튼 프레임 (함수 정의 후)
        button_frame = ttk.Frame(content_frame)
        button_frame.pack(pady=10)
        ttk.Button(button_frame, text="🔗 URL 다운로드", command=download_from_url, width=15).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="📋 클립보드", command=get_from_clipboard, width=12).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="📷 파일 선택", command=select_image_manual, width=12).pack(side=tk.LEFT, padx=5)


        def save_product():
            if not name_var.get():
                messagebox.showwarning("경고", "상품명을 입력해주세요.")
                return

            # 상품코드 자동 분리 기능
            colors_input = colors_var.get().strip()
            sizes_input = sizes_var.get().strip()

            if auto_split_var.get() and code_var.get():
                # 자동 분리 체크박스가 선택되어 있고 상품코드가 있는 경우
                detected_colors, detected_sizes = utils.auto_split_product_code(code_var.get())

                # 색상 자동 입력 (기존 입력값이 없는 경우에만)
                if detected_colors and not colors_input:
                    colors_var.set(', '.join(detected_colors))
                    colors_input = colors_var.get()

                # 사이즈 자동 입력 (기존 입력값이 없는 경우에만)
                if detected_sizes and not sizes_input:
                    sizes_var.set(', '.join(detected_sizes))
                    sizes_input = sizes_var.get()

            colors = [c.strip() for c in colors_input.split(',') if c.strip()]
            sizes = [s.strip() for s in sizes_input.split(',') if s.strip()]

            if not colors:
                colors = ['']
            if not sizes:
                sizes = ['FREE']
            
            # 오더 수량 단위 처리
            order_unit_value = order_unit_var.get().strip()
            if order_unit_value:
                try:
                    order_unit = int(order_unit_value)
                except ValueError:
                    order_unit = None
            else:
                order_unit = None

            # 자동 분리가 체크되어 있고 색상이 여러 개인 경우
            base_code = code_var.get()
            if auto_split_var.get() and base_code and len(colors) > 1 and colors != ['']:
                # 각 색상별로 별도의 상품 생성
                added_count = 0
                for idx, color in enumerate(colors, start=1):
                    product = {
                        'id': self.data_manager.get_next_product_id(),
                        'name': name_var.get(),
                        'code': f"{base_code}-{idx}",  # 코드에 -1, -2 등 추가
                        'supplier': supplier_var.get(),
                        'colors': [color],  # 각 상품은 하나의 색상만
                        'sizes': sizes,
                        'memo': memo_text.get("1.0", tk.END).strip(),
                        'image': image_data.get() if image_data.get() else None,
                        'image_source': image_source.get(),
                        'order_unit': order_unit
                    }
                    # DB에 저장
                    self.data_manager.add_product(product)
                    added_count += 1
                
                # self.data_manager.save_data()
                # ← 자동저장 제거됨
                self._refresh_data_shortcuts()
                self.refresh_products_list()
                
                messagebox.showinfo("완료", f"{added_count}개의 상품이 색상별로 분리되어 추가되었습니다.")
                dialog.destroy()
            else:
                # 자동 분리가 체크되지 않았거나 색상이 1개 이하인 경우 기존 방식대로 저장
                product = {
                    'name': name_var.get(),
                    'code': base_code,
                    'supplier': supplier_var.get(),
                    'colors': colors,
                    'sizes': sizes,
                    'memo': memo_text.get("1.0", tk.END).strip(),
                    'image': image_data.get() if image_data.get() else None,
                    'image_source': image_source.get(),
                    'order_unit': order_unit
                }

                # DB에 저장
                self.data_manager.add_product(product)
                # self.data_manager.save_data()
                # ← 자동저장 제거됨
                self._refresh_data_shortcuts()
                self.refresh_products_list()
                
                messagebox.showinfo("완료", "상품이 추가되었습니다.")
                dialog.destroy()
        
        # 저장/취소 버튼
        bottom_frame = ttk.Frame(content_frame)
        bottom_frame.pack(pady=20)
        ttk.Button(bottom_frame, text="💾 저장", command=save_product, width=15).pack(side=tk.LEFT, padx=5)
        ttk.Button(bottom_frame, text="❌ 취소", command=dialog.destroy, width=15).pack(side=tk.LEFT, padx=5)
        

    def edit_product(self):
        selected = self.products_tree.selection()
        if not selected:
            messagebox.showwarning("경고", "수정할 상품을 선택해주세요.")
            return

        item = self.products_tree.item(selected[0])
        product_name = item['values'][0]
        product_code = item['values'][1]

        # 상품명과 상품코드로 정확한 상품 찾기
        product = None
        for p in self.products:
            if p['name'] == product_name and p.get('code', '') == product_code:
                product = p
                break

        if not product:
            return

        dialog = tk.Toplevel(self.root)
        dialog.title("상품 수정")
        dialog.geometry("700x950")
        dialog.transient(self.root)
        dialog.grab_set()
        utils.center_window(dialog)

        # 스크롤 가능한 프레임 생성
        content_frame = utils.make_scrollable_dialog(dialog, max_height=900)

        ttk.Label(content_frame, text="상품명:", font=("Arial", 10)).pack(pady=(20,5))
        name_var = tk.StringVar(value=product['name'])
        ttk.Entry(content_frame, textvariable=name_var, width=50, justify='center').pack(pady=5)

        # 상품코드와 체크박스
        code_frame = ttk.Frame(content_frame)
        code_frame.pack(pady=(10,5))

        ttk.Label(code_frame, text="상품코드:", font=("Arial", 10)).pack(side=tk.LEFT, padx=(0,10))

        code_var = tk.StringVar(value=product.get('code', ''))
        code_entry = ttk.Entry(code_frame, textvariable=code_var, width=40, justify='center')
        code_entry.pack(side=tk.LEFT, padx=5)

        # 자동 분리 체크박스
        auto_split_var = tk.BooleanVar(value=self.data_manager.get_auto_split_setting())
        auto_split_check = ttk.Checkbutton(
            code_frame,
            text="자동분리",
            variable=auto_split_var,
            command=lambda: self.data_manager.set_auto_split_setting(auto_split_var.get())
        )
        auto_split_check.pack(side=tk.LEFT, padx=5)

        ttk.Label(content_frame, text="매입처:", font=("Arial", 10)).pack(pady=(10,5))
        supplier_var = tk.StringVar(value=product.get('supplier', ''))
        ttk.Entry(content_frame, textvariable=supplier_var, width=50, justify='center').pack(pady=5)

        ttk.Label(content_frame, text=f"{self.field_names[0]['name'] if len(self.field_names) > 0 else '색상'} (쉼표로 구분):", font=("Arial", 10)).pack(pady=(10,5))
        colors_var = tk.StringVar(value=', '.join(product.get('colors', [''])))
        ttk.Entry(content_frame, textvariable=colors_var, width=50, justify='center').pack(pady=5)

        ttk.Label(content_frame, text=f"{self.field_names[1]['name'] if len(self.field_names) > 1 else '사이즈'} (쉼표로 구분):", font=("Arial", 10)).pack(pady=(10,5))
        sizes_var = tk.StringVar(value=', '.join(product.get('sizes', ['FREE'])))
        ttk.Entry(content_frame, textvariable=sizes_var, width=50, justify='center').pack(pady=5)

        # 메모 필드 추가
        ttk.Label(content_frame, text="주문 관련 메모:", font=("Arial", 10)).pack(pady=(10,5))
        memo_text = tk.Text(content_frame, width=50, height=4, wrap=tk.WORD)
        memo_text.insert("1.0", product.get('memo', ''))
        memo_text.pack(pady=5, padx=20)

        # 오더 수량 단위 필드 추가
        ttk.Label(content_frame, text="오더 수량 단위 (선택사항):", font=("Arial", 10)).pack(pady=(10,5))
        order_unit_var = tk.StringVar(value=str(product.get('order_unit', '')) if product.get('order_unit') else '')
        ttk.Entry(content_frame, textvariable=order_unit_var, width=50, justify='center').pack(pady=5)
        ttk.Label(content_frame, text="예: 40 (40장 단위로 발주해야 하는 경우)", font=("Arial", 8), foreground="gray").pack()

        image_data = tk.StringVar(value=product.get('image', ''))
        image_source = tk.StringVar(value=product.get('image_source', 'none'))
        search_status = tk.StringVar()

        # 현재 이미지 상태 표시
        if product.get('image'):
            search_status.set(f"✅ 이미지 있음 ({product.get('image_source', 'unknown')})")
        else:
            search_status.set("❌ 이미지 없음")

        ttk.Label(content_frame, text="━━━━━ 이미지 설정 ━━━━━", font=("Arial", 10, "bold")).pack(pady=(15, 5))
        ttk.Label(content_frame, textvariable=search_status, font=("Arial", 9), foreground="blue").pack(pady=5)

        # 이미지 URL 입력 필드
        ttk.Label(content_frame, text="이미지 URL:", font=("Arial", 9)).pack(pady=(10, 5))
        url_var = tk.StringVar()
        url_entry = ttk.Entry(content_frame, textvariable=url_var, width=50, justify='center')
        url_entry.pack(pady=5)
        
        def download_from_url():
            """URL에서 이미지 다운로드"""
            url = url_var.get().strip()
            if not url:
                messagebox.showwarning("경고", "이미지 URL을 입력해주세요.")
                return
            
            if not url.startswith('http'):
                messagebox.showwarning("경고", "올바른 URL을 입력해주세요. (http:// 또는 https://로 시작)")
                return
            
            search_status.set("⏳ 이미지 다운로드 중...")
            dialog.update()
            
            try:
                img_base64 = utils.download_image_from_url(url)
                
                if img_base64:
                    image_data.set(img_base64)
                    image_source.set("url")
                    search_status.set(f"✅ 이미지 다운로드 완료")
                    messagebox.showinfo("완료", "이미지가 다운로드되었습니다!")
                else:
                    search_status.set("❌ 이미지 다운로드 실패")
                    messagebox.showerror("오류", "이미지를 다운로드할 수 없습니다.\n URL을 확인해주세요.")
            except Exception as e:
                search_status.set("❌ 오류 발생")
                messagebox.showerror("오류", f"오류가 발생했습니다:\n{str(e)}")
        
        def select_image_manual():
            """수동으로 이미지 파일 선택"""
            filename = filedialog.askopenfilename(
                title="이미지 선택",
                filetypes=[("Image files", "*.png *.jpg *.jpeg *.gif *.webp *.bmp"), ("All files", "*.*")]
            )
            if filename:
                try:
                    with open(filename, 'rb') as f:
                        image_data.set(base64.b64encode(f.read()).decode())
                    image_source.set("manual")
                    search_status.set(f"✅ 이미지 선택됨: {os.path.basename(filename)}")
                    messagebox.showinfo("완료", "이미지가 선택되었습니다.")
                except Exception as e:
                    messagebox.showerror("오류", f"이미지를 읽을 수 없습니다: {str(e)}")
        
        button_frame = ttk.Frame(content_frame)
        button_frame.pack(pady=10)
        ttk.Button(button_frame, text="🔗 URL 다운로드", command=download_from_url, width=15).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="📷 파일 선택", command=select_image_manual, width=15).pack(side=tk.LEFT, padx=5)

        def save_changes():
            if not name_var.get():
                messagebox.showwarning("경고", "상품명을 입력해주세요.")
                return

            # 상품코드 자동 분리 기능
            colors_input = colors_var.get().strip()
            sizes_input = sizes_var.get().strip()

            if auto_split_var.get() and code_var.get():
                # 자동 분리 체크박스가 선택되어 있고 상품코드가 있는 경우
                detected_colors, detected_sizes = utils.auto_split_product_code(code_var.get())

                # 색상 자동 입력 (기존 입력값이 없는 경우에만)
                if detected_colors and not colors_input:
                    colors_var.set(', '.join(detected_colors))
                    colors_input = colors_var.get()

                # 사이즈 자동 입력 (기존 입력값이 없는 경우에만)
                if detected_sizes and not sizes_input:
                    sizes_var.set(', '.join(detected_sizes))
                    sizes_input = sizes_var.get()

            colors = [c.strip() for c in colors_input.split(',') if c.strip()]
            sizes = [s.strip() for s in sizes_input.split(',') if s.strip()]

            if not colors:
                colors = ['']
            if not sizes:
                sizes = ['FREE']
            
            # 오더 수량 단위 처리
            order_unit_value = order_unit_var.get().strip()
            if order_unit_value:
                try:
                    order_unit = int(order_unit_value)
                except ValueError:
                    order_unit = None
            else:
                order_unit = None

            base_code = code_var.get()
            
            # 자동 분리가 체크되어 있고 색상이 여러 개인 경우
            if auto_split_var.get() and base_code and len(colors) > 1 and colors != ['']:
                # 기존 상품 DB에서 삭제
                if hasattr(self.data_manager, 'delete_product_from_db'):
                    try:
                        self.data_manager.delete_product_from_db(product['id'])
                    except Exception as e:
                        print(f"상품 삭제 오류: {e}")
                        self.products.remove(product)
                
                # 각 색상별로 별도의 상품 생성
                for idx, color in enumerate(colors, start=1):
                    new_product = {
                        'name': name_var.get(),
                        'code': f"{base_code}-{idx}",  # 코드에 -1, -2 등 추가
                        'supplier': supplier_var.get(),
                        'colors': [color],  # 각 상품은 하나의 색상만
                        'sizes': sizes,
                        'memo': memo_text.get("1.0", tk.END).strip(),
                        'image': image_data.get() if image_data.get() else None,
                        'image_source': image_source.get(),
                        'order_unit': order_unit
                    }
                    # DB에 저장
                    self.data_manager.add_product(new_product)
                
                # self.data_manager.save_data()
                # ← 자동저장 제거됨
                self._refresh_data_shortcuts()
                self.refresh_products_list()
                self.refresh_orders_list()
                self.refresh_stock_list()
                
                messagebox.showinfo("완료", f"상품이 {len(colors)}개로 색상별로 분리되어 수정되었습니다.")
                dialog.destroy()
            else:
                # 자동 분리가 체크되지 않았거나 색상이 1개 이하인 경우 기존 방식대로 저장
                update_data = {
                    'name': name_var.get(),
                    'code': base_code,
                    'supplier': supplier_var.get(),
                    'colors': colors,
                    'sizes': sizes,
                    'memo': memo_text.get("1.0", tk.END).strip(),
                    'image_url': image_data.get() if image_data.get() else None,
                    'order_unit': order_unit
                }
                
                # DB 업데이트
                try:
                    result = self.data_manager.update_product_in_db(product['id'], update_data)
                    if result:
                        print(f"✅ DB 상품 수정: ID={product['id']}, 이름={update_data['name']}")
                    else:
                        print(f"⚠️ DB 수정 실패: ID={product['id']}")
                        messagebox.showerror("오류", "상품 수정에 실패했습니다.")
                        return
                except Exception as e:
                    print(f"❌ 상품 업데이트 오류: {e}")
                    import traceback
                    traceback.print_exc()
                    messagebox.showerror("오류", f"상품 수정 오류: {e}")
                    return
                
                # DB에서 최신 데이터로 갱신
                self._refresh_data_shortcuts()
                self.refresh_products_list()
                self.refresh_orders_list()
                self.refresh_stock_list()
                
                messagebox.showinfo("완료", "상품이 수정되었습니다.")
                dialog.destroy()
        
        btn_frame = ttk.Frame(content_frame)
        btn_frame.pack(pady=20)
        ttk.Button(btn_frame, text="저장", command=save_changes).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="취소", command=dialog.destroy).pack(side=tk.LEFT, padx=5)
    
    def delete_product(self):
        selected = self.products_tree.selection()
        if not selected:
            messagebox.showwarning("경고", "삭제할 상품을 선택해주세요.")
            return
        
        # 다중 선택 처리
        delete_count = len(selected)
        
        # 삭제 확인
        if delete_count > 1:
            if not messagebox.askyesno("확인", f"선택한 {delete_count}개의 항목을 삭제하시겠습니까?"):
                return
        
        # 삭제할 항목 정보 수집 (상품명, 상품코드, 색상, 사이즈)
        items_to_delete = []
        for item_id in selected:
            item = self.products_tree.item(item_id)
            items_to_delete.append({
                'name': item['values'][0],
                'code': item['values'][1],
                'color': item['values'][2] if item['values'][2] != '-' else '',
                'size': item['values'][3]
            })
        
        # 단일 선택일 때만 재고/발주 확인
        if delete_count == 1:
            item_info = items_to_delete[0]
            product_name = item_info['name']
            product_code = item_info['code']
            selected_color = item_info['color']
            selected_size = item_info['size']
            
            # 상품 찾기 (상품명 + 상품코드)
            product = None
            for p in self.products:
                if p['name'] == product_name and p.get('code', '') == product_code:
                    product = p
                    break
            
            if product:
                has_stock = self.data_manager.calculate_stock_by_variant(product['id'], selected_color, selected_size) != 0
                has_orders = self.data_manager.calculate_pending_by_variant(product['id'], selected_color, selected_size) != 0
                
                warning_msg = f"선택한 '{product_name} ({product_code}) - {selected_color if selected_color else '색상없음'} - {selected_size}'를 삭제하시겠습니까?"
                if has_stock or has_orders:
                    warning_msg += "\n\n경고: 이 색상/사이즈 조합에 재고 또는 발주 기록이 있습니다."
                
                if not messagebox.askyesno("확인", warning_msg):
                    return
        
        # 삭제할 상품들 수집
        products_to_delete = []
        product_ids_to_delete = []
        
        for item_info in items_to_delete:
            product_name = item_info['name']
            product_code = item_info['code']
            
            # 상품 찾기 (상품명 + 상품코드)
            product = None
            for p in self.products:
                if p['name'] == product_name and p.get('code', '') == product_code:
                    product = p
                    break
            
            if product and product not in products_to_delete:
                products_to_delete.append(product)
                product_ids_to_delete.append(product['id'])
        
        # 상품 삭제 (DB에서)
        deleted_count = 0
        for product in products_to_delete:
            try:
                result = self.data_manager.delete_product_from_db(product['id'])
                if result:
                    print(f"✅ DB 상품 삭제: ID={product['id']}, 이름={product['name']}")
                    deleted_count += 1
                else:
                    print(f"⚠️ DB 삭제 실패: ID={product['id']}")
            except Exception as e:
                print(f"❌ DB 삭제 오류: {e}")
        
        # DB에서 최신 데이터로 갱신
        self._refresh_data_shortcuts()
        self.refresh_products_list()
        self.refresh_orders_list()
        self.refresh_stock_list()
        self.refresh_inbound_list()
        self.refresh_outbound_list()
        
        if deleted_count > 0:
            messagebox.showinfo("완료", f"{deleted_count}개의 상품이 삭제되었습니다.")
        else:
            messagebox.showwarning("경고", "삭제된 항목이 없습니다.")
    
    def add_order(self):
        # 매장 목록 확인
        print(f"📍 발주 추가 - 매장 목록: {len(self.stores)}개")
        if not self.stores:
            messagebox.showwarning("매장 없음", 
                "등록된 매장이 없습니다.\n\n"
                "매장 관리 탭에서 먼저 매장을 추가해주세요.")
            return
        
        dialog = tk.Toplevel(self.root)
        dialog.title("발주 추가")
        dialog.geometry("600x900")
        dialog.transient(self.root)
        dialog.grab_set()
        utils.center_window(dialog)
        
        # 발주 날짜 선택
        ttk.Label(dialog, text="발주 날짜:", font=("Arial", 10, "bold")).pack(pady=(20,5))
        date_entry = DateEntry(dialog, width=20, background='darkblue', foreground='white', 
                              borderwidth=2, date_pattern='yyyy-mm-dd')
        date_entry.pack(pady=5)
        
        # 매장 선택 필드
        ttk.Label(dialog, text="발주 매장:", font=("Arial", 10, "bold")).pack(pady=(10,5))
        store_var = tk.StringVar()
        store_combo = ttk.Combobox(dialog, textvariable=store_var, width=40, state="readonly")
        store_combo['values'] = [s['name'] for s in self.stores]
        if self.stores:
            store_var.set(self.stores[0]['name'])
        store_combo.pack(pady=5)
        
        ttk.Label(dialog, text="상품 검색 (엔터로 검색):", font=("Arial", 10, "bold")).pack(pady=(10,5))
        
        # 검색창
        search_var = tk.StringVar()
        search_entry = ttk.Entry(dialog, textvariable=search_var, width=40)
        search_entry.pack(pady=5)
        ttk.Label(dialog, text="상품명 또는 코드 입력 후 엔터", 
                 font=("Arial", 8), foreground="gray").pack()
        
        # 검색 결과 리스트
        result_frame = ttk.LabelFrame(dialog, text="검색 결과", padding=10)
        result_frame.pack(pady=10, padx=20, fill=tk.BOTH)

        result_listbox = tk.Listbox(result_frame, height=5)
        result_listbox.pack(fill=tk.BOTH, expand=True)

        # 동적 콘텐츠를 위한 컨테이너 프레임 (info_frame, pending_frame 순서 보장)
        dynamic_content_frame = ttk.Frame(dialog)
        dynamic_content_frame.pack(pady=5, padx=20, fill=tk.BOTH, expand=True)

        # 선택된 상품 정보
        selected_product = {'product': None}

        color_var = tk.StringVar()
        size_var = tk.StringVar()
        quantity_var = tk.StringVar()

        color_combo = None
        size_combo = None

        # pending_frame과 pending_tree는 상품 선택 후 생성
        pending_frame_container = {'frame': None, 'tree': None}
        
        def search_products(event=None):
            search_text = search_var.get().lower().strip()
            result_listbox.delete(0, tk.END)
            
            if not search_text:
                return
            
            # 매칭된 상품을 저장 (ID를 키로 사용)
            found_products_map = {}
            for product in self.products:
                if search_text in product['name'].lower() or search_text in product.get('code', '').lower():
                    display_text = f"{product['name']} ({product.get('code', '없음')}) [ID:{product['id']}]"
                    result_listbox.insert(tk.END, display_text)
                    # display_text를 키로 하여 실제 상품 저장
                    found_products_map[display_text] = product
            
            # 찾은 상품 맵을 저장
            result_listbox.found_products_map = found_products_map
            print(f"검색 결과: {len(found_products_map)}개 상품 찾음")
        
        def on_product_select(event):
            nonlocal color_combo, size_combo
            
            selection = result_listbox.curselection()
            if not selection:
                return
            
            selected_text = result_listbox.get(selection[0])
            print(f"=== 상품 선택 디버깅 ===")
            print(f"선택된 텍스트: {selected_text}")
            
            # 저장된 맵에서 직접 가져오기
            found_products_map = getattr(result_listbox, 'found_products_map', {})
            found_product = found_products_map.get(selected_text)
            
            if not found_product:
                # 맵에서 찾지 못한 경우 기존 방식으로 재시도
                print("맵에서 찾지 못함, 직접 검색 시도...")
                for product in self.products:
                    # ID 포함된 텍스트에서 매칭
                    if f"[ID:{product['id']}]" in selected_text:
                        found_product = product
                        print(f"재검색으로 매칭된 상품: {product['name']}, ID: {product['id']}")
                        break
            else:
                print(f"맵에서 찾은 상품: {found_product['name']}, ID: {found_product['id']}")
            
            if not found_product:
                print("상품을 찾을 수 없음!")
                messagebox.showwarning("경고", "상품을 찾을 수 없습니다.")
                return
            
            # 선택된 상품 저장
            selected_product['product'] = found_product
            print(f"selected_product에 저장됨: {selected_product['product']['name']}, ID: {selected_product['product']['id']}")
            print("=====================")
            
            # 기존 동적 콘텐츠 삭제 (dynamic_content_frame 내부만)
            for widget in dynamic_content_frame.winfo_children():
                widget.destroy()

            # info_frame을 dynamic_content_frame 내부에 생성 (위쪽)
            info_frame = ttk.Frame(dynamic_content_frame)
            info_frame.pack(pady=10, fill=tk.X)
            
            ttk.Label(info_frame, text=f"선택: {found_product['name']} [ID: {found_product['id']}]", 
                     font=("Arial", 10, "bold")).pack()
            
            # 색상
            ttk.Label(info_frame, text="색상:", font=("Arial", 9)).pack(pady=(10,2))
            colors = found_product.get('colors', [''])
            color_combo = ttk.Combobox(info_frame, textvariable=color_var, width=30, state="readonly")
            color_combo['values'] = colors
            if colors:
                color_var.set(colors[0])
            color_combo.pack()
            
            # 사이즈
            ttk.Label(info_frame, text="사이즈:", font=("Arial", 9)).pack(pady=(10,2))
            sizes = found_product.get('sizes', ['FREE'])
            size_combo = ttk.Combobox(info_frame, textvariable=size_var, width=30, state="readonly")
            size_combo['values'] = sizes
            if sizes:
                size_var.set(sizes[0])
            size_combo.pack()
            
            # 수량
            ttk.Label(info_frame, text="발주 수량:", font=("Arial", 9)).pack(pady=(10,2))
            quantity_entry = ttk.Entry(info_frame, textvariable=quantity_var, width=30)
            quantity_entry.pack()
            quantity_entry.focus()
            
            # 미입고 현황 프레임 - dynamic_content_frame 내부에 생성 (아래쪽, info_frame 다음)
            pending_frame = ttk.LabelFrame(dynamic_content_frame, text="📊 미입고 현황", padding=10)
            pending_frame.pack(pady=10, fill=tk.BOTH, expand=True)
            pending_frame_container['frame'] = pending_frame
            
            def update_pending_display():
                """미입고 현황 업데이트 (입력한 수량 포함)"""
                if not pending_frame_container['frame']:
                    return None
                
                # 위젯이 파괴되었는지 확인
                try:
                    if not pending_frame_container['frame'].winfo_exists():
                        return None
                except:
                    return None
                
                # 기존 트리 삭제
                for widget in pending_frame_container['frame'].winfo_children():
                    widget.destroy()
                
                columns = ("매장", "색상", "사이즈", "현재 미입고", "입력 수량", "합계")
                pending_tree_new = ttk.Treeview(pending_frame_container['frame'], columns=columns, show="headings", height=6)
                
                col_widths = {"매장": 100, "색상": 80, "사이즈": 80, "현재 미입고": 90, "입력 수량": 90, "합계": 90}
                for col in columns:
                    pending_tree_new.heading(col, text=col)
                    pending_tree_new.column(col, width=col_widths.get(col, 90), anchor='center')
                
                pending_tree_new.pack(fill=tk.BOTH, expand=True)
                
                # 현재 입력한 수량
                try:
                    input_qty = int(quantity_var.get()) if quantity_var.get() else 0
                except:
                    input_qty = 0
                
                current_store_id = None
                if store_var.get():
                    for s in self.stores:
                        if s['name'] == store_var.get():
                            current_store_id = s['id']
                            break
                
                # 매장별 미입고 데이터 계산
                for store in self.stores:
                    for color in found_product.get('colors', ['']):
                        for size in found_product.get('sizes', ['FREE']):
                            # 현재 선택한 색상/사이즈와 일치하는지 확인
                            is_current = (store['id'] == current_store_id and 
                                         color == color_var.get() and 
                                         size == size_var.get())
                            
                            pending_qty = 0
                            for order in self.orders:
                                if (order['product_id'] == found_product['id'] and
                                    order.get('store_id') == store['id'] and
                                    order.get('color', '') == color and
                                    order.get('size', 'FREE') == size and
                                    order.get('status') != 'completed'):
                                    pending_qty += (order['quantity'] - order['shipped_quantity'])
                            
                            # 현재 입력 중인 수량 추가
                            add_qty = input_qty if is_current else 0
                            total_qty = pending_qty + add_qty
                            
                            # 미입고가 있거나 입력 중인 항목만 표시
                            if pending_qty > 0 or add_qty > 0:
                                # 현재 입력 중인 행은 강조
                                tag = 'current' if is_current else ''
                                item_id = pending_tree_new.insert('', tk.END, values=(
                                    store['name'],
                                    color if color else '-',
                                    size,
                                    pending_qty,
                                    add_qty,
                                    total_qty
                                ), tags=(tag,))
                                
                                if is_current:
                                    pending_tree_new.tag_configure('current', background='#e6f2ff')
                
                pending_frame_container['tree'] = pending_tree_new
                return pending_tree_new
            
            # 초기 표시
            update_pending_display()
            
            # 수량 변경 시 미입고 현황 업데이트
            def on_quantity_change(*args):
                try:
                    if dialog.winfo_exists():
                        update_pending_display()
                except:
                    pass
            
            quantity_var.trace('w', on_quantity_change)
            
            # 색상/사이즈 변경 시에도 업데이트
            def on_variant_change(*args):
                try:
                    if dialog.winfo_exists():
                        update_pending_display()
                except:
                    pass
            
            color_var.trace('w', on_variant_change)
            size_var.trace('w', on_variant_change)
            
            # 엔터 키로 저장
            quantity_entry.bind('<Return>', lambda e: save_order())
        
        search_entry.bind('<Return>', search_products)
        result_listbox.bind('<<ListboxSelect>>', on_product_select)
        
        def save_order():
            if not selected_product['product']:
                messagebox.showwarning("경고", "상품을 선택해주세요.")
                return

            if not store_var.get() and self.stores:
                messagebox.showwarning("경고", "발주 매장을 선택해주세요.")
                return

            try:
                quantity = int(quantity_var.get())
                if quantity <= 0:
                    raise ValueError
            except:
                messagebox.showwarning("경고", "올바른 수량을 입력해주세요.")
                return

            product = selected_product['product']

            # 디버깅: 선택된 상품 정보 출력
            print(f"=== 발주 저장 디버깅 ===")
            print(f"선택된 상품: {product['name']}")
            print(f"상품 ID: {product['id']}")
            print(f"선택된 색상: {color_var.get()}")
            print(f"선택된 사이즈: {size_var.get()}")

            # 선택된 매장 찾기 (정수로 변환)
            store_id = None
            store_name = store_var.get()
            for s in self.stores:
                if s['name'] == store_name:
                    try:
                        store_id = int(s['id'])
                    except (ValueError, TypeError):
                        store_id = s['id']
                    break
            
            print(f"선택된 매장: {store_name} → store_id={store_id}")

            # 오더 수량 단위 처리
            order_unit = product.get('order_unit')
            actual_order_qty = quantity
            pending_qty = 0

            if order_unit and order_unit > 0:
                # 기존 추가발주(pending_extra) 확인
                existing_pending = 0
                for order in self.orders:
                    if (order.get('product_id') == product['id'] and
                        order.get('color', '') == color_var.get() and
                        order.get('size', 'FREE') == size_var.get() and
                        order.get('store_id') == store_id and
                        order.get('status') == 'pending_extra'):
                        existing_pending += order.get('quantity', 0)

                # 총 수량 = 기존 추가발주 + 현재 발주 수량
                total_qty = existing_pending + quantity

                if total_qty >= order_unit:
                    # 오더 단위 이상이면 오더 단위로 발주
                    actual_order_qty = (total_qty // order_unit) * order_unit
                    pending_qty = total_qty % order_unit

                    # 기존 추가발주 삭제 (DB에서도 삭제)
                    orders_to_delete = [o for o in self.orders if (
                        o.get('product_id') == product['id'] and
                        o.get('color', '') == color_var.get() and
                        o.get('size', 'FREE') == size_var.get() and
                        o.get('store_id') == store_id and
                        o.get('status') == 'pending_extra')]
                    
                    for o in orders_to_delete:
                        if hasattr(self.data_manager, 'delete_order_from_db') and o.get('id'):
                            try:
                                self.data_manager.delete_order_from_db(o['id'])
                            except Exception as e:
                                print(f"추가발주 삭제 오류: {e}")
                    
                    self.orders = [o for o in self.orders if o not in orders_to_delete]

                    # 남은 수량은 추가발주로
                    if pending_qty > 0:
                        pending_order = {
                            'product_id': product['id'],
                            'color': color_var.get(),
                            'size': size_var.get() if size_var.get() else 'FREE',
                            'quantity': pending_qty,
                            'shipped_quantity': 0,
                            'status': 'pending_extra',
                            'store_id': store_id,
                            'date': date_entry.get_date().strftime('%Y-%m-%d'),
                            'note': f'추가발주 대기 (오더단위: {order_unit})'
                        }
                        # Supabase DB에 저장
                        if hasattr(self.data_manager, 'add_order_to_db'):
                            self.data_manager.add_order_to_db(pending_order)
                        else:
                            pending_order['id'] = len(self.orders) + 1
                            self.orders.append(pending_order)

                    info_msg = f"오더 단위 발주 처리:\n"
                    if existing_pending > 0:
                        info_msg += f"기존 추가발주: {existing_pending}장\n"
                    info_msg += f"실제 발주: {actual_order_qty}장 (오더 단위: {order_unit}장)\n"
                    if pending_qty > 0:
                        info_msg += f"추가발주 대기: {pending_qty}장"
                    messagebox.showinfo("오더 단위 처리", info_msg)
                else:
                    # 오더 단위 미만이면 추가 발주로만 기록
                    actual_order_qty = 0
                    pending_qty = total_qty

                    # 기존 추가발주 삭제 (DB에서도 삭제)
                    orders_to_delete2 = [o for o in self.orders if (
                        o.get('product_id') == product['id'] and
                        o.get('color', '') == color_var.get() and
                        o.get('size', 'FREE') == size_var.get() and
                        o.get('store_id') == store_id and
                        o.get('status') == 'pending_extra')]
                    
                    for o in orders_to_delete2:
                        if hasattr(self.data_manager, 'delete_order_from_db') and o.get('id'):
                            try:
                                self.data_manager.delete_order_from_db(o['id'])
                            except Exception as e:
                                print(f"추가발주 삭제 오류: {e}")
                    
                    self.orders = [o for o in self.orders if o not in orders_to_delete2]

                    # 새로운 추가발주 생성
                    pending_order = {
                        'product_id': product['id'],
                        'color': color_var.get(),
                        'size': size_var.get() if size_var.get() else 'FREE',
                        'quantity': pending_qty,
                        'shipped_quantity': 0,
                        'status': 'pending_extra',
                        'store_id': store_id,
                        'date': date_entry.get_date().strftime('%Y-%m-%d'),
                        'note': f'추가발주 대기 (오더단위: {order_unit})'
                    }
                    # Supabase DB에 저장
                    if hasattr(self.data_manager, 'add_order_to_db'):
                        self.data_manager.add_order_to_db(pending_order)
                    else:
                        pending_order['id'] = len(self.orders) + 1
                        self.orders.append(pending_order)

                    info_msg = f"오더 단위 미달:\n"
                    if existing_pending > 0:
                        info_msg += f"기존 추가발주: {existing_pending}장\n"
                    info_msg += f"누적 추가발주: {pending_qty}장\n"
                    info_msg += f"오더 단위까지 부족: {order_unit - pending_qty}장\n"
                    messagebox.showinfo("추가발주 누적", info_msg)

                    # 추가발주만 기록하고 종료
                    # self.data_manager.save_data()
                    # ← 자동저장 제거됨
                    self._refresh_data_shortcuts()
                    self.refresh_orders_list()
                    self.refresh_products_list()
                    quantity_var.set("")
                    return

            # 실제 발주 생성 (오더 수량이 있을 경우만)
            if actual_order_qty > 0:
                order = {
                    'date': date_entry.get_date().strftime('%Y-%m-%d'),
                    'product_id': product['id'],
                    'color': color_var.get(),
                    'size': size_var.get() if size_var.get() else 'FREE',
                    'quantity': actual_order_qty,
                    'shipped_quantity': 0,
                    'status': 'pending',
                    'store_id': store_id
                }

                print(f"생성된 발주: product_id={order['product_id']}, color={order['color']}, size={order['size']}, qty={order['quantity']}")
                print("=====================")

                # Supabase DB에 저장
                if hasattr(self.data_manager, 'add_order_to_db'):
                    try:
                        self.data_manager.add_order_to_db(order)
                        print("✅ Supabase DB에 발주 저장 완료")
                    except Exception as e:
                        print(f"❌ Supabase 저장 오류: {e}")
                        order['id'] = len(self.orders) + 1
                        self.orders.append(order)
                else:
                    order['id'] = len(self.orders) + 1
                    self.orders.append(order)

            # self.data_manager.save_data()
            # ← 자동저장 제거됨
            self._refresh_data_shortcuts()
            self.refresh_orders_list()
            self.refresh_products_list()

            # 입력 필드 초기화 (창은 닫지 않음)
            quantity_var.set("")
            if not order_unit:
                messagebox.showinfo("완료", f"발주가 추가되었습니다.\n상품: {product['name']}\n계속 추가할 수 있습니다.")
        
        btn_frame = ttk.Frame(dialog)
        btn_frame.pack(pady=20)
        ttk.Button(btn_frame, text="💾 저장", command=save_order, width=15).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="❌ 닫기", command=dialog.destroy, width=15).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="취소", command=dialog.destroy).pack(side=tk.LEFT, padx=5)
    
    def on_order_double_click(self, event):
        """발주 트리 더블클릭 시 셀 편집"""
        region = self.orders_tree.identify("region", event.x, event.y)
        if region != "cell":
            return
        
        column = self.orders_tree.identify_column(event.x)
        row_id = self.orders_tree.identify_row(event.y)
        
        if not row_id:
            return
        
        # 편집 가능한 컬럼: 발주일자(#1), 발주수량(#7)
        col_index = int(column.replace('#', '')) - 1
        col_name = self.orders_tree['columns'][col_index]
        
        if col_name not in ["발주일자", "발주수량"]:
            return
        
        # 현재 값 가져오기
        item_values = self.orders_tree.item(row_id)['values']
        current_value = item_values[col_index]
        
        # 셀 위치 계산
        x, y, width, height = self.orders_tree.bbox(row_id, column)
        
        # 편집 위젯 생성
        if col_name == "발주일자":
            # 날짜 선택
            edit_window = tk.Toplevel(self.root)
            edit_window.overrideredirect(True)
            edit_window.geometry(f"250x250+{self.orders_tree.winfo_rootx() + x}+{self.orders_tree.winfo_rooty() + y}")
            
            date_picker = DateEntry(edit_window, width=20, background='darkblue', 
                                   foreground='white', borderwidth=2, date_pattern='yyyy-mm-dd')
            try:
                date_parts = str(current_value).split('-')
                date_picker.set_date(datetime(int(date_parts[0]), int(date_parts[1]), int(date_parts[2])))
            except:
                pass
            date_picker.pack(pady=5)
            
            def save_date():
                new_date = date_picker.get_date().strftime('%Y-%m-%d')
                self.update_order_field(item_values, "date", new_date)
                edit_window.destroy()
            
            ttk.Button(edit_window, text="저장", command=save_date).pack(pady=5)
            ttk.Button(edit_window, text="취소", command=edit_window.destroy).pack(pady=5)
            
        else:  # 발주수량
            entry = ttk.Entry(self.orders_tree, width=width)
            entry.place(x=x, y=y, width=width, height=height)
            entry.insert(0, current_value)
            entry.select_range(0, tk.END)
            entry.focus()
            
            def save_value(event=None):
                new_value = entry.get()
                try:
                    new_quantity = int(new_value)
                    if new_quantity > 0:
                        self.update_order_field(item_values, "quantity", new_quantity)
                    entry.destroy()
                except ValueError:
                    messagebox.showwarning("경고", "올바른 숫자를 입력해주세요.")
                    entry.destroy()
            
            def cancel(event=None):
                entry.destroy()
            
            entry.bind("<Return>", save_value)
            entry.bind("<Escape>", cancel)
            entry.bind("<FocusOut>", save_value)
    
    def update_order_field(self, item_values, field, new_value):
        """발주 필드 업데이트"""
        # 디버그: item_values 전체 출력
        print(f"📋 item_values: {item_values}")
        
        order_date = str(item_values[0])
        product_name = str(item_values[1])
        # 상품명에서 "(추가발주: N장)" 부분 제거
        if " (추가발주:" in product_name:
            product_name = product_name.split(" (추가발주:")[0]
        product_code = str(item_values[2]) if item_values[2] != '-' else ''
        color = str(item_values[3]) if item_values[3] != '-' else ''
        size = str(item_values[4])
        store_name = str(item_values[5]) if len(item_values) > 5 and item_values[5] != '-' else ''
        current_total_quantity = item_values[6] if len(item_values) > 6 else 0
        
        print(f"📋 파싱: date={order_date}, name={product_name}, code={product_code}, color={color}, size={size}, store_name='{store_name}'")
        
        # 매장 ID 찾기 (정수로 변환)
        store_id = None
        for store in self.stores:
            if store['name'] == store_name:
                store_id = store.get('id')
                if store_id is not None:
                    try:
                        store_id = int(store_id)
                    except (ValueError, TypeError):
                        store_id = None
                break
        
        print(f"📋 매장 검색: store_name='{store_name}' → store_id={store_id}")
        print(f"📋 등록된 매장: {[(s['name'], s['id']) for s in self.stores]}")
        
        # 상품 찾기 (상품코드가 있으면 상품코드로 우선 검색)
        product = None
        if product_code:
            # 상품코드로 먼저 찾기
            for p in self.products:
                if p.get('code', '') == product_code:
                    product = p
                    break
        
        # 상품코드로 못 찾으면 상품명+상품코드 조합으로
        if not product:
            for p in self.products:
                if p['name'] == product_name and p.get('code', '') == product_code:
                    product = p
                    break
        
        # 그래도 못 찾으면 상품명만으로 (상품코드 없는 경우)
        if not product and not product_code:
            for p in self.products:
                if p['name'] == product_name:
                    product = p
                    break
        
        if not product:
            messagebox.showerror("오류", "상품을 찾을 수 없습니다.")
            return
        
        # 같은 날짜, 상품, 색상, 사이즈, 매장의 모든 발주 찾기
        matching_orders = []
        old_total = 0
        
        # 디버그 출력
        print(f"🔍 발주 검색: product_id={product['id']}, color={color}, size={size}, store_id={store_id}, date={order_date}")
        
        # 해당 날짜의 발주가 있는지 먼저 확인
        date_orders = [o for o in self.orders if str(o.get('date') or o.get('order_date') or '') == order_date]
        print(f"🔍 {order_date} 날짜의 발주: {len(date_orders)}개")
        
        for order in self.orders:
            # 비교를 위해 값 정규화
            order_product_id = order.get('product_id')
            order_color = order.get('color') or ''
            order_size = order.get('size') or 'FREE'
            order_store_id = order.get('store_id')
            order_order_date = str(order.get('date') or order.get('order_date') or '')
            
            # store_id를 정수로 변환
            if order_store_id is not None and order_store_id != '':
                try:
                    order_store_id = int(order_store_id)
                except (ValueError, TypeError):
                    order_store_id = None
            else:
                order_store_id = None
            
            # store_id가 None이면 store_id 조건 무시
            store_match = (store_id is None or order_store_id == store_id)
            
            if (order_product_id == product['id'] and 
                order_color == color and 
                order_size == size and
                store_match and
                order_order_date == order_date):
                matching_orders.append(order)
                old_total += order.get('quantity', 0)
        
        print(f"🔍 매칭된 발주: {len(matching_orders)}개, 총 수량: {old_total}")
        
        if not matching_orders:
            # 디버그: 해당 날짜 발주 상세
            for order in date_orders[:10]:
                print(f"  📅 {order_date} 발주: pid={order.get('product_id')}, color={order.get('color')}, size={order.get('size')}, store={order.get('store_id')}")
            messagebox.showwarning("경고", "수정할 발주를 찾을 수 없습니다.")
            return
        
        # 업데이트 처리
        if field == "date":
            # 날짜 변경은 모든 매칭되는 발주에 적용
            for order in matching_orders:
                order['date'] = new_value
        
        elif field == "quantity":
            # 수량 변경 - 비율대로 분배
            if len(matching_orders) == 1:
                # 발주가 1개면 직접 변경
                matching_orders[0]['quantity'] = new_value
                
                # 상태 업데이트
                order = matching_orders[0]
                if order['shipped_quantity'] >= new_value:
                    order['status'] = 'completed'
                elif order['shipped_quantity'] > 0:
                    order['status'] = 'partial'
                else:
                    order['status'] = 'pending'
            else:
                # 여러 발주가 합쳐진 경우 - 비율대로 분배
                if old_total > 0:
                    for order in matching_orders:
                        ratio = order['quantity'] / old_total
                        order['quantity'] = int(new_value * ratio)
                        
                        # 상태 업데이트
                        if order['shipped_quantity'] >= order['quantity']:
                            order['status'] = 'completed'
                        elif order['shipped_quantity'] > 0:
                            order['status'] = 'partial'
                        else:
                            order['status'] = 'pending'
                else:
                    # old_total이 0인 경우 균등 분배
                    each_quantity = new_value // len(matching_orders)
                    remainder = new_value % len(matching_orders)
                    
                    for i, order in enumerate(matching_orders):
                        order['quantity'] = each_quantity + (1 if i < remainder else 0)
                        
                        # 상태 업데이트
                        if order['shipped_quantity'] >= order['quantity']:
                            order['status'] = 'completed'
                        elif order['shipped_quantity'] > 0:
                            order['status'] = 'partial'
                        else:
                            order['status'] = 'pending'
        
        # Supabase에 저장
        for order in matching_orders:
            self.data_manager.update_order_in_db(order['id'], order)
        
        self._refresh_data_shortcuts()
        self.refresh_orders_list()
        self.refresh_products_list()
    
    def edit_order(self):
        """발주 수정 (삭제 예정 - 더블클릭 편집으로 대체)"""
        messagebox.showinfo("안내", "발주 수정은 더블클릭으로 직접 편집할 수 있습니다.\n\n"
                           "• 발주일자: 더블클릭하여 날짜 선택\n"
                           "• 발주수량: 더블클릭하여 직접 입력")
    
    def _get_selected_order_info(self):
        """선택된 발주 정보 가져오기"""
        selected = self.orders_tree.selection()
        if not selected:
            messagebox.showwarning("경고", "수정할 발주를 선택해주세요.")
            return None
        
        item = self.orders_tree.item(selected[0])
        values = item['values']
        order_date = str(values[0])
        product_name = str(values[1])
        # 상품명에서 "(추가발주: N장)" 부분 제거
        if " (추가발주:" in product_name:
            product_name = product_name.split(" (추가발주:")[0]
        product_code = str(values[2]) if values[2] != '-' else ''
        color = str(values[3]) if values[3] != '-' else ''
        size = str(values[4])
        store_name = str(values[5]) if values[5] != '-' else ''
        
        # 매장 ID 찾기
        store_id = ''
        for store in self.stores:
            if store['name'] == store_name:
                store_id = store.get('id', '')
                break
        
        # 상품 찾기 (상품코드가 있으면 상품코드로 우선 검색)
        product = None
        if product_code:
            # 상품코드로 먼저 찾기
            for p in self.products:
                if p.get('code', '') == product_code:
                    product = p
                    break
        
        # 상품코드로 못 찾으면 상품명+상품코드 조합으로
        if not product:
            for p in self.products:
                if p['name'] == product_name and p.get('code', '') == product_code:
                    product = p
                    break
        
        # 그래도 못 찾으면 상품명만으로 (상품코드 없는 경우)
        if not product and not product_code:
            for p in self.products:
                if p['name'] == product_name:
                    product = p
                    break
        
        if not product:
            messagebox.showerror("오류", f"상품을 찾을 수 없습니다.\n상품명: {product_name}\n상품코드: {product_code}")
            return None
        
        # 해당 발주들 찾기
        matching_orders = []
        for order in self.orders:
            if (str(order['product_id']) == str(product['id']) and 
                order.get('color', '') == color and 
                order.get('size', 'FREE') == size and
                str(order.get('store_id', '')) == str(store_id) and
                str(order['date']) == order_date):
                matching_orders.append(order)
        
        if not matching_orders:
            messagebox.showwarning("경고", "수정할 발주를 찾을 수 없습니다.")
            return None
        
        return {
            'orders': matching_orders,
            'product': product,
            'color': color,
            'size': size,
            'store_id': store_id,
            'store_name': store_name
        }
    
    def change_order_product(self):
        """발주 상품 변경"""
        info = self._get_selected_order_info()
        if not info:
            return
        
        # 상품 선택 다이얼로그
        dialog = tk.Toplevel(self.root)
        dialog.title("상품 변경")
        dialog.geometry("500x600")
        dialog.transient(self.root)
        dialog.grab_set()
        utils.center_window(dialog)
        
        ttk.Label(dialog, text="변경할 상품 선택", font=("Arial", 12, "bold")).pack(pady=10)
        
        # 검색
        search_frame = ttk.Frame(dialog)
        search_frame.pack(fill=tk.X, padx=10, pady=5)
        ttk.Label(search_frame, text="검색:").pack(side=tk.LEFT)
        search_var = tk.StringVar()
        search_entry = ttk.Entry(search_frame, textvariable=search_var, width=30)
        search_entry.pack(side=tk.LEFT, padx=5)
        
        # 상품 목록
        columns = ("상품명", "상품코드")
        tree = ttk.Treeview(dialog, columns=columns, show="headings", height=15)
        tree.heading("상품명", text="상품명")
        tree.heading("상품코드", text="상품코드")
        tree.column("상품명", width=250)
        tree.column("상품코드", width=150)
        tree.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        
        def refresh_list():
            tree.delete(*tree.get_children())
            search_text = search_var.get().lower()
            for p in self.products:
                name = p['name'].lower()
                code = p.get('code', '').lower()
                if search_text in name or search_text in code:
                    tree.insert('', tk.END, values=(p['name'], p.get('code', '')))
        
        refresh_list()
        search_var.trace('w', lambda *args: refresh_list())
        
        def select_product():
            selected = tree.selection()
            if not selected:
                messagebox.showwarning("경고", "상품을 선택해주세요.")
                return
            
            new_product_name = tree.item(selected[0])['values'][0]
            new_product = None
            for p in self.products:
                if p['name'] == new_product_name:
                    new_product = p
                    break
            
            if not new_product:
                return
            
            # 발주 업데이트
            for order in info['orders']:
                order['product_id'] = new_product['id']
                # 색상이 새 상품에 없으면 초기화
                new_colors = new_product.get('colors', [])
                if order.get('color', '') and order.get('color', '') not in new_colors:
                    order['color'] = new_colors[0] if new_colors else ''
                # 사이즈가 새 상품에 없으면 초기화
                new_sizes = new_product.get('sizes', ['FREE'])
                if order.get('size', 'FREE') not in new_sizes:
                    order['size'] = new_sizes[0] if new_sizes else 'FREE'
            
            self._refresh_data_shortcuts()
            self.refresh_orders_list()
            self.refresh_products_list()
            dialog.destroy()
            messagebox.showinfo("완료", "상품이 변경되었습니다.")
        
        btn_frame = ttk.Frame(dialog)
        btn_frame.pack(pady=10)
        ttk.Button(btn_frame, text="선택", command=select_product, width=10).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="취소", command=dialog.destroy, width=10).pack(side=tk.LEFT, padx=5)
    
    def change_order_color(self):
        """발주 색상 변경"""
        info = self._get_selected_order_info()
        if not info:
            return
        
        product = info['product']
        old_color = info['color']
        
        # 같은 상품명을 가진 모든 상품에서 색상 수집
        product_name = product['name']
        color_to_product = {}  # {색상: 상품} 매핑
        all_colors = []
        
        for p in self.products:
            if p['name'] == product_name:
                p_colors = p.get('colors', [])
                for c in p_colors:
                    if c and c not in all_colors:
                        all_colors.append(c)
                        color_to_product[c] = p
        
        if not all_colors:
            messagebox.showinfo("안내", "이 상품은 색상이 없습니다.")
            return
        
        # 색상 선택 다이얼로그
        dialog = tk.Toplevel(self.root)
        dialog.title("색상 변경")
        dialog.geometry("300x400")
        dialog.transient(self.root)
        dialog.grab_set()
        utils.center_window(dialog)
        
        ttk.Label(dialog, text=f"상품: {product_name}", font=("Arial", 11, "bold")).pack(pady=10)
        ttk.Label(dialog, text=f"현재 색상: {old_color if old_color else '-'}").pack(pady=5)
        ttk.Label(dialog, text="변경할 색상 선택:", font=("Arial", 10)).pack(pady=10)
        
        listbox = tk.Listbox(dialog, height=10, font=("Arial", 11))
        for color in all_colors:
            listbox.insert(tk.END, color)
        listbox.pack(fill=tk.BOTH, expand=True, padx=20, pady=5)
        
        # 현재 색상 선택
        if old_color in all_colors:
            listbox.selection_set(all_colors.index(old_color))
        
        def select_color():
            selected = listbox.curselection()
            if not selected:
                messagebox.showwarning("경고", "색상을 선택해주세요.")
                return
            
            new_color = all_colors[selected[0]]
            new_product = color_to_product.get(new_color, product)
            
            for order in info['orders']:
                order['color'] = new_color
                order['product_id'] = new_product['id']
            
            self._refresh_data_shortcuts()
            self.refresh_orders_list()
            dialog.destroy()
            messagebox.showinfo("완료", f"색상이 '{new_color}'로 변경되었습니다.")
        
        btn_frame = ttk.Frame(dialog)
        btn_frame.pack(pady=10)
        ttk.Button(btn_frame, text="선택", command=select_color, width=10).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="취소", command=dialog.destroy, width=10).pack(side=tk.LEFT, padx=5)
    
    def change_order_size(self):
        """발주 사이즈 변경"""
        info = self._get_selected_order_info()
        if not info:
            return
        
        product = info['product']
        sizes = product.get('sizes', ['FREE'])
        
        # 사이즈 선택 다이얼로그
        dialog = tk.Toplevel(self.root)
        dialog.title("사이즈 변경")
        dialog.geometry("300x400")
        dialog.transient(self.root)
        dialog.grab_set()
        utils.center_window(dialog)
        
        ttk.Label(dialog, text=f"상품: {product['name']}", font=("Arial", 11, "bold")).pack(pady=10)
        ttk.Label(dialog, text=f"현재 사이즈: {info['size']}").pack(pady=5)
        ttk.Label(dialog, text="변경할 사이즈 선택:", font=("Arial", 10)).pack(pady=10)
        
        listbox = tk.Listbox(dialog, height=10, font=("Arial", 11))
        for size in sizes:
            listbox.insert(tk.END, size)
        listbox.pack(fill=tk.BOTH, expand=True, padx=20, pady=5)
        
        # 현재 사이즈 선택
        if info['size'] in sizes:
            listbox.selection_set(sizes.index(info['size']))
        
        def select_size():
            selected = listbox.curselection()
            if not selected:
                messagebox.showwarning("경고", "사이즈를 선택해주세요.")
                return
            
            new_size = sizes[selected[0]]
            for order in info['orders']:
                order['size'] = new_size
            
            self._refresh_data_shortcuts()
            self.refresh_orders_list()
            dialog.destroy()
            messagebox.showinfo("완료", f"사이즈가 '{new_size}'로 변경되었습니다.")
        
        btn_frame = ttk.Frame(dialog)
        btn_frame.pack(pady=10)
        ttk.Button(btn_frame, text="선택", command=select_size, width=10).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="취소", command=dialog.destroy, width=10).pack(side=tk.LEFT, padx=5)
    
    def change_order_store(self):
        """발주 매장 변경"""
        info = self._get_selected_order_info()
        if not info:
            return
        
        if not self.stores:
            messagebox.showinfo("안내", "등록된 매장이 없습니다.")
            return
        
        # 매장 선택 다이얼로그
        dialog = tk.Toplevel(self.root)
        dialog.title("매장 변경")
        dialog.geometry("300x400")
        dialog.transient(self.root)
        dialog.grab_set()
        utils.center_window(dialog)
        
        ttk.Label(dialog, text=f"현재 매장: {info['store_name'] if info['store_name'] else '-'}", font=("Arial", 11)).pack(pady=10)
        ttk.Label(dialog, text="변경할 매장 선택:", font=("Arial", 10)).pack(pady=10)
        
        listbox = tk.Listbox(dialog, height=10, font=("Arial", 11))
        store_names = [s['name'] for s in self.stores]
        for name in store_names:
            listbox.insert(tk.END, name)
        listbox.pack(fill=tk.BOTH, expand=True, padx=20, pady=5)
        
        # 현재 매장 선택
        if info['store_name'] in store_names:
            listbox.selection_set(store_names.index(info['store_name']))
        
        def select_store():
            selected = listbox.curselection()
            if not selected:
                messagebox.showwarning("경고", "매장을 선택해주세요.")
                return
            
            new_store_name = store_names[selected[0]]
            new_store_id = ''
            for s in self.stores:
                if s['name'] == new_store_name:
                    new_store_id = s.get('id', '')
                    break
            
            for order in info['orders']:
                order['store_id'] = new_store_id
            
            self._refresh_data_shortcuts()
            self.refresh_orders_list()
            dialog.destroy()
            messagebox.showinfo("완료", f"매장이 '{new_store_name}'로 변경되었습니다.")
        
        btn_frame = ttk.Frame(dialog)
        btn_frame.pack(pady=10)
        ttk.Button(btn_frame, text="선택", command=select_store, width=10).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="취소", command=dialog.destroy, width=10).pack(side=tk.LEFT, padx=5)
    
    def delete_order(self):
        """발주 삭제 (다중 선택 지원)"""
        selected = self.orders_tree.selection()
        if not selected:
            messagebox.showwarning("경고", "삭제할 발주를 선택해주세요.")
            return
        
        delete_count = len(selected)
        if not messagebox.askyesno("확인", f"선택한 {delete_count}개의 발주를 삭제하시겠습니까?"):
            return
        
        # 삭제할 발주 정보 수집
        orders_to_delete = []
        for item_id in selected:
            item = self.orders_tree.item(item_id)
            values = item['values']
            orders_to_delete.append({
                'date': values[0],
                'product_name': values[1],
                'color': values[3] if values[3] != '-' else '',
                'size': values[4],
                'store_name': values[5] if values[5] != '-' else ''
            })
        
        # 발주 삭제 처리
        deleted_count = 0
        for order_info in orders_to_delete:
            # 매장 ID 찾기
            store_id = ''
            for store in self.stores:
                if store['name'] == order_info['store_name']:
                    store_id = str(store.get('id', ''))
                    break
            
            # 상품 찾기 (상품명 + 색상으로 정확하게 찾기)
            product = None
            order_color = order_info['color']
            
            for p in self.products:
                # 상품명이 일치하고
                if p['name'] == order_info['product_name']:
                    # 색상도 일치하는지 확인
                    product_colors = p.get('colors', [''])
                    # 색상이 비어있거나 색상이 일치하면 선택
                    if not order_color or order_color in product_colors:
                        product = p
                        break
            
            if not product:
                continue
            
            # 같은 날짜, 상품, 색상, 사이즈, 매장의 모든 발주 삭제
            for o in self.orders[:]:
                if (o['product_id'] == product['id'] and 
                    o['date'] == order_info['date'] and 
                    o.get('color', '') == order_info['color'] and 
                    o.get('size', 'FREE') == order_info['size'] and
                    str(o.get('store_id', '')) == store_id):
                    
                    # Supabase DB에서 삭제
                    order_id = o.get('id')
                    if order_id and hasattr(self.data_manager, 'delete_order_from_db'):
                        try:
                            self.data_manager.delete_order_from_db(order_id)
                        except Exception as e:
                            print(f"DB 삭제 오류: {e}")
                            # 메모리에서도 수동 삭제
                            if o in self.orders:
                                self.orders.remove(o)
                    else:
                        # JSON 모드: 메모리에서만 삭제
                        self.orders.remove(o)
                    
                    deleted_count += 1
        
        if deleted_count > 0:
            self._refresh_data_shortcuts()
            self.refresh_orders_list()
            self.refresh_products_list()
            messagebox.showinfo("완료", f"{deleted_count}개의 발주가 삭제되었습니다.")
        else:
            messagebox.showwarning("경고", "삭제된 발주가 없습니다.")
    
    def shipment_order(self):
        """발주 출고 처리 (구 버전 호환용 - 사용하지 않음)"""
        pass
    
    def receive_order(self):
        """발주 입고 처리 (구 버전 호환용 - 사용하지 않음)"""
        pass
    
    def export_orders_excel(self):
        # 선택된 날짜 가져오기
        selected_date = self.order_date_var.get() if hasattr(self, 'order_date_var') else datetime.now().strftime('%Y-%m-%d')
        
        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialfile=f"발주장_{selected_date}_{datetime.now().strftime('%H%M%S')}.xlsx"
        )
        
        if filename:
            try:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "발주장"
                
                # 발주매장 컬럼 추가
                headers = ["발주일자", "상품명", "상품코드", "색상", "사이즈", "발주매장", "발주수량", "미입고수량", "메모"]
                
                header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")
                border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_num)
                    cell.value = header
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.border = border
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # 선택된 날짜의 발주만 필터링
                filtered_orders = [o for o in self.orders if o.get('date', '') == selected_date]
                
                # 매장 포함하여 같은 날짜, 상품, 색상, 사이즈, 매장을 합치기
                merged_orders = {}
                for order in filtered_orders:
                    product = self.data_manager.get_product_by_id(order['product_id'])
                    if not product:
                        continue
                    
                    # 매장 ID를 키에 포함
                    key = (order.get('date', ''), order['product_id'], order.get('color', ''), order.get('size', 'FREE'), order.get('store_id', ''))
                    
                    if key in merged_orders:
                        merged_orders[key]['quantity'] += order['quantity']
                        merged_orders[key]['shipped_quantity'] += order['shipped_quantity']
                        # 메모 합치기 (중복 제외)
                        if order.get('note') and order.get('note') not in merged_orders[key]['notes']:
                            merged_orders[key]['notes'].append(order.get('note'))
                    else:
                        merged_orders[key] = {
                            'date': order.get('date', ''),
                            'product': product,
                            'color': order.get('color', ''),
                            'size': order.get('size', 'FREE'),
                            'quantity': order['quantity'],
                            'shipped_quantity': order['shipped_quantity'],
                            'store_id': order.get('store_id', ''),
                            'notes': [order.get('note')] if order.get('note') else []
                        }
                
                # 상품명순으로 정렬
                sorted_merged = sorted(merged_orders.items(), key=lambda x: x[1]['product']['name'])
                
                for row_num, (key, merged) in enumerate(sorted_merged, 2):
                    pending_qty = merged['quantity'] - merged['shipped_quantity']
                    store = self.data_manager.get_store_by_id(merged.get('store_id'))
                    store_name = store['name'] if store else '-'
                    note_text = ', '.join(merged['notes']) if merged['notes'] else ''
                    
                    row_data = [
                        merged['date'],
                        merged['product']['name'],
                        merged['product'].get('code', ''),
                        merged['color'] if merged['color'] else '-',
                        merged['size'],
                        store_name,
                        merged['quantity'],
                        pending_qty,
                        note_text
                    ]
                    
                    for col_num, value in enumerate(row_data, 1):
                        cell = ws.cell(row=row_num, column=col_num)
                        cell.value = value
                        cell.border = border  # 메모란도 border 적용
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # 컬럼 너비 자동 조정
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 30)
                    ws.column_dimensions[column_letter].width = adjusted_width
                
                wb.save(filename)
                
                # 저장 완료 후 폴더 열기 옵션
                if messagebox.askyesno("저장 완료", 
                                       f"발주장이 저장되었습니다.\n{filename}\n\n저장 폴더를 여시겠습니까?"):
                    import subprocess
                    folder_path = os.path.dirname(filename)
                    if platform.system() == 'Windows':
                        os.startfile(folder_path)
                    elif platform.system() == 'Darwin':  # macOS
                        subprocess.Popen(['open', folder_path])
                    else:  # Linux
                        subprocess.Popen(['xdg-open', folder_path])
                        
            except Exception as e:
                messagebox.showerror("오류", f"파일 저장 중 오류가 발생했습니다:\n{str(e)}")
    
    def refresh_products_list(self):
        for item in self.products_tree.get_children():
            self.products_tree.delete(item)
        
        search_term = self.product_search_var.get().lower()
        
        # 상품을 이름 순으로 정렬
        sorted_products = sorted(self.products, key=lambda x: x['name'])
        
        for product in sorted_products:
            if search_term and search_term not in product['name'].lower() and search_term not in product.get('code', '').lower():
                continue
            
            # 동적 필드 값 가져오기
            colors = product.get('colors', [''])
            sizes = product.get('sizes', ['FREE'])
            
            # 색상별로 사이즈를 통합하여 표시
            for color in colors:
                # 해당 색상의 모든 사이즈에 대한 재고와 미입고를 합산
                total_stock = 0
                total_pending = 0
                
                for size in sizes:
                    stock = self.data_manager.calculate_stock_by_variant(product['id'], color, size)
                    pending = self.data_manager.calculate_pending_by_variant(product['id'], color, size)
                    total_stock += stock
                    total_pending += pending
                
                # 동적 컬럼 생성
                values = [product['name'], product.get('code', '')]
                
                # 필드 값 추가
                if len(self.field_names) >= 1:
                    values.append(color if color else '-')
                
                # 사이즈는 "전체"로 표시 (통합되었음을 나타냄)
                if len(self.field_names) >= 2:
                    # 사이즈가 여러 개면 "통합", 한 개면 해당 사이즈 표시
                    if len(sizes) > 1:
                        values.append('통합')
                    else:
                        values.append(sizes[0] if sizes else 'FREE')
                
                # 추가 필드가 있다면 처리
                for i in range(2, len(self.field_names)):
                    extra_field_value = product.get(f'field{i+1}', '-')
                    values.append(extra_field_value)
                
                values.extend([total_stock, total_pending])
                
                self.products_tree.insert('', tk.END, values=tuple(values))
    
    def refresh_orders_list(self):
        for item in self.orders_tree.get_children():
            self.orders_tree.delete(item)
        
        # 선택된 날짜의 발주만 표시
        selected_date = self.order_date_var.get()
        
        # 해당 날짜의 발주 필터링
        filtered_orders = [o for o in self.orders if o.get('date', '') == selected_date]
        
        # 같은 날짜, 상품, 색상, 사이즈, 매장을 합치기
        merged_orders = {}
        for order in filtered_orders:
            product = self.data_manager.get_product_by_id(order['product_id'])
            if not product:
                print(f"⚠️ 발주에서 상품을 찾을 수 없음: product_id={order['product_id']}, 발주 날짜={order.get('date')}")
                continue
            
            key = (order.get('date', ''), order['product_id'], order.get('color', ''), order.get('size', 'FREE'), order.get('store_id', ''))
            
            if key in merged_orders:
                merged_orders[key]['quantity'] += order['quantity']
                merged_orders[key]['shipped_quantity'] += order['shipped_quantity']
                # 메모가 있으면 추가 (중복 제외)
                if order.get('note') and order.get('note') not in merged_orders[key]['notes']:
                    merged_orders[key]['notes'].append(order.get('note'))
            else:
                merged_orders[key] = {
                    'date': order.get('date', ''),
                    'product': product,
                    'product_id': order['product_id'],
                    'color': order.get('color', ''),
                    'size': order.get('size', 'FREE'),
                    'quantity': order['quantity'],
                    'shipped_quantity': order['shipped_quantity'],
                    'store_id': order.get('store_id', ''),
                    'notes': [order.get('note')] if order.get('note') else []
                }
        
        # 매장별로 정렬
        sorted_merged = sorted(merged_orders.items(), key=lambda x: (x[1]['store_id'] or '', x[1]['product']['name']))
        
        # 병합된 데이터로 표시
        for key, merged in sorted_merged:
            pending_qty = merged['quantity'] - merged['shipped_quantity']
            store = self.data_manager.get_store_by_id(merged['store_id'])

            # Check for pending extra (추가발주 대기)
            pending_extra = 0
            for order in self.orders:
                if (order.get('product_id') == merged.get('product_id') and
                    order.get('color', '') == merged.get('color', '') and
                    order.get('size', 'FREE') == merged.get('size', 'FREE') and
                    order.get('store_id') == merged.get('store_id') and
                    order.get('status') == 'pending_extra'):
                    pending_extra += order.get('quantity', 0)

            product_name = merged['product']['name']
            if pending_extra > 0:
                product_name += f" (추가발주: {pending_extra}장)"
            
            # 메모 합치기
            note_text = ', '.join(merged['notes']) if merged['notes'] else ''

            self.orders_tree.insert('', tk.END, values=(
                merged['date'],
                product_name,
                merged['product'].get('code', ''),
                merged['color'] if merged['color'] else '-',
                merged['size'],
                store['name'] if store else '-',
                merged['quantity'],
                pending_qty,
                note_text
            ))
    
    def sync_from_db(self):
        """Supabase에서 데이터 다시 로드 (동기화)"""
        try:
            print("🔄 Supabase에서 데이터 동기화 중...")
            # 캐시 무효화 후 새로 로드
            if hasattr(self.data_manager, 'invalidate_all_cache'):
                self.data_manager.invalidate_all_cache()
            self.data_manager.load_data()
            self._refresh_data_shortcuts()
            self.update_order_dates()
            self.refresh_orders_list()
            self.refresh_products_list()
            self.refresh_stock_list()
            if hasattr(self, 'refresh_inbound_list'):
                self.refresh_inbound_list()
            if hasattr(self, 'refresh_outbound_list'):
                self.refresh_outbound_list()
            print("✅ 동기화 완료")
            
            # 상태바에 표시
            if hasattr(self, 'status_label'):
                self.status_label.config(text="✅ 동기화 완료")
                self.root.after(3000, lambda: self.status_label.config(text=""))
        except Exception as e:
            print(f"❌ 동기화 오류: {e}")
            messagebox.showerror("동기화 오류", f"데이터 동기화 중 오류가 발생했습니다:\n{e}")
    
    def update_order_dates(self):
        """발주 날짜 목록 업데이트"""
        dates = set()
        for order in self.orders:
            if order.get('date'):
                dates.add(order.get('date'))
        
        # 오늘 날짜 추가
        today = datetime.now().strftime('%Y-%m-%d')
        dates.add(today)
        
        sorted_dates = sorted(list(dates), reverse=True)
        self.order_date_combo['values'] = sorted_dates
        
        # 현재 선택된 날짜가 목록에 없으면 오늘로 설정
        if self.order_date_var.get() not in sorted_dates:
            self.order_date_var.set(today)
        
        self.refresh_orders_list()
    
    def search_order_period(self):
        """기간 검색 다이얼로그"""
        dialog = tk.Toplevel(self.root)
        dialog.title("기간 검색")
        dialog.geometry("450x300")
        dialog.transient(self.root)
        dialog.grab_set()
        utils.center_window(dialog)

        ttk.Label(dialog, text="발주 기간 검색", font=("Arial", 12, "bold")).pack(pady=20)

        # 날짜 입력 프레임
        date_frame = ttk.Frame(dialog)
        date_frame.pack(pady=10)

        ttk.Label(date_frame, text="시작 날짜:", font=("Arial", 10)).grid(row=0, column=0, padx=10, pady=5)
        start_date_entry = DateEntry(date_frame, width=15, background='darkblue', foreground='white',
                                    borderwidth=2, date_pattern='yyyy-mm-dd')
        start_date_entry.grid(row=0, column=1, padx=10, pady=5)

        ttk.Label(date_frame, text="종료 날짜:", font=("Arial", 10)).grid(row=1, column=0, padx=10, pady=5)
        end_date_entry = DateEntry(date_frame, width=15, background='darkblue', foreground='white',
                                  borderwidth=2, date_pattern='yyyy-mm-dd')
        end_date_entry.grid(row=1, column=1, padx=10, pady=5)

        ttk.Label(dialog, text="달력에서 날짜를 선택하세요",
                 font=("Arial", 8), foreground="gray").pack(pady=5)

        def do_search():
            start = start_date_entry.get_date().strftime('%Y-%m-%d')
            end = end_date_entry.get_date().strftime('%Y-%m-%d')

            if start > end:
                messagebox.showwarning("경고", "시작 날짜가 종료 날짜보다 늦습니다.")
                return

            # 기간 내의 모든 발주 표시
            for item in self.orders_tree.get_children():
                self.orders_tree.delete(item)

            filtered_orders = [o for o in self.orders
                             if start <= o.get('date', '') <= end]

            # 같은 날짜, 상품, 색상, 사이즈, 매장을 합치기
            merged_orders = {}
            for order in filtered_orders:
                product = self.data_manager.get_product_by_id(order['product_id'])
                if not product:
                    continue

                key = (order.get('date', ''), order['product_id'], order.get('color', ''),
                      order.get('size', 'FREE'), order.get('store_id', ''))

                if key in merged_orders:
                    merged_orders[key]['quantity'] += order['quantity']
                    merged_orders[key]['shipped_quantity'] += order['shipped_quantity']
                else:
                    merged_orders[key] = {
                        'date': order.get('date', ''),
                        'product': product,
                        'product_id': order['product_id'],
                        'color': order.get('color', ''),
                        'size': order.get('size', 'FREE'),
                        'quantity': order['quantity'],
                        'shipped_quantity': order['shipped_quantity'],
                        'store_id': order.get('store_id', '')
                    }

            # 날짜, 매장별로 정렬
            sorted_merged = sorted(merged_orders.items(),
                                 key=lambda x: (x[1]['date'], x[1]['store_id']), reverse=True)

            # 병합된 데이터로 표시
            for key, merged in sorted_merged:
                pending_qty = merged['quantity'] - merged['shipped_quantity']
                store = self.data_manager.get_store_by_id(merged['store_id'])

                # Check for pending extra (추가발주 대기)
                pending_extra = 0
                for order in self.orders:
                    if (order.get('product_id') == merged.get('product_id') and
                        order.get('color', '') == merged.get('color', '') and
                        order.get('size', 'FREE') == merged.get('size', 'FREE') and
                        order.get('store_id') == merged.get('store_id') and
                        order.get('status') == 'pending_extra'):
                        pending_extra += order.get('quantity', 0)

                product_name = merged['product']['name']
                if pending_extra > 0:
                    product_name += f" (추가발주: {pending_extra}장)"

                self.orders_tree.insert('', tk.END, values=(
                    merged['date'],
                    product_name,
                    merged['product'].get('code', ''),
                    merged['color'] if merged['color'] else '-',
                    merged['size'],
                    store['name'] if store else '-',
                    merged['quantity'],
                    pending_qty
                ))

            messagebox.showinfo("완료", f"{start} ~ {end} 기간의 발주 {len(sorted_merged)}건을 조회했습니다.")

        def do_reset():
            """검색 초기화 및 오늘 날짜로 복원"""
            self.order_date_var.set(datetime.now().strftime('%Y-%m-%d'))
            self.refresh_orders_list()
            dialog.destroy()

        btn_frame = ttk.Frame(dialog)
        btn_frame.pack(pady=20)
        ttk.Button(btn_frame, text="🔍 검색", command=do_search, width=12).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="🔄 초기화", command=do_reset, width=12).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="❌ 닫기", command=dialog.destroy, width=12).pack(side=tk.LEFT, padx=5)
    
    def refresh_stock_list(self):
        for item in self.stock_tree.get_children():
            self.stock_tree.delete(item)
        
        # 검색어 가져오기
        search_term = self.stock_search_var.get().lower() if hasattr(self, 'stock_search_var') else ""
        
        # 1. 창고 재고 먼저 표시
        for product in self.products:
            # 검색 필터링 (상품명, 상품코드)
            if search_term and search_term not in product['name'].lower() and search_term not in product.get('code', '').lower():
                continue
            
            colors = product.get('colors', [''])
            sizes = product.get('sizes', ['FREE'])
            
            for color in colors:
                for size in sizes:
                    # 창고 재고 계산 (store_id = None)
                    warehouse_in = sum(m['quantity'] for m in self.movements 
                                      if m['product_id'] == product['id'] and m['type'] == 'in'
                                      and m.get('color', '') == color and m.get('size', 'FREE') == size
                                      and m.get('store_id') is None)
                    
                    warehouse_out = sum(m['quantity'] for m in self.movements 
                                       if m['product_id'] == product['id'] and m['type'] == 'out'
                                       and m.get('color', '') == color and m.get('size', 'FREE') == size
                                       and m.get('store_id') is None)
                    
                    warehouse_stock = warehouse_in - warehouse_out
                    
                    # 창고 대상 미입고 계산 (매장 지정이 없는 발주만)
                    warehouse_ordered = sum(o['quantity'] for o in self.orders
                                           if o['product_id'] == product['id']
                                           and o.get('color', '') == color
                                           and o.get('size', 'FREE') == size
                                           and not o.get('store_id'))  # 매장 지정이 없는 것만
                    
                    warehouse_shipped = sum(o['shipped_quantity'] for o in self.orders
                                           if o['product_id'] == product['id']
                                           and o.get('color', '') == color
                                           and o.get('size', 'FREE') == size
                                           and not o.get('store_id'))  # 매장 지정이 없는 것만
                    
                    warehouse_pending = warehouse_ordered - warehouse_shipped
                    
                    # 창고 재고가 있거나 미입고가 있는 경우만 표시
                    if warehouse_stock > 0 or warehouse_in > 0 or warehouse_out > 0 or warehouse_pending > 0:
                        self.stock_tree.insert('', tk.END, values=(
                            '🏢 창고',
                            product['name'],
                            product.get('code', ''),
                            color if color else '-',
                            size,
                            warehouse_stock,
                            warehouse_in,
                            warehouse_out,
                            warehouse_pending
                        ))
        
        # 2. 매장별 재고 표시
        for store in self.stores:
            # 매장명으로도 검색 가능
            store_match = not search_term or search_term in store['name'].lower()
            
            for product in self.products:
                # 검색 필터링 (상품명, 상품코드, 매장명)
                if search_term:
                    product_match = (search_term in product['name'].lower() or 
                                   search_term in product.get('code', '').lower())
                    if not (product_match or store_match):
                        continue
                
                colors = product.get('colors', [''])
                sizes = product.get('sizes', ['FREE'])
                
                for color in colors:
                    for size in sizes:
                        # 해당 매장의 재고 계산 (창고에서 출고된 것만)
                        store_in = sum(m['quantity'] for m in self.movements 
                                      if m['product_id'] == product['id'] and m['type'] == 'out'
                                      and m.get('color', '') == color and m.get('size', 'FREE') == size
                                      and m.get('store_id') == store['id'])
                        
                        store_stock = store_in
                        
                        # 해당 매장의 미입고 계산 (해당 매장으로 발주된 것만)
                        store_ordered = sum(o['quantity'] for o in self.orders
                                          if o['product_id'] == product['id']
                                          and o.get('store_id') == store['id']
                                          and o.get('color', '') == color
                                          and o.get('size', 'FREE') == size)
                        
                        store_shipped = sum(o['shipped_quantity'] for o in self.orders
                                          if o['product_id'] == product['id']
                                          and o.get('store_id') == store['id']
                                          and o.get('color', '') == color
                                          and o.get('size', 'FREE') == size)
                        
                        store_pending = store_ordered - store_shipped
                        
                        # 재고나 미입고가 있는 경우만 표시
                        if store_stock > 0 or store_pending > 0 or store_in > 0:
                            self.stock_tree.insert('', tk.END, values=(
                                store['name'],
                                product['name'],
                                product.get('code', ''),
                                color if color else '-',
                                size,
                                store_stock,
                                store_in,
                                0,  # 매장의 총출고 (현재는 추적 안함)
                                store_pending
                            ))
    
    def sort_stock_tree(self, col):
        """재고 현황 테이블 정렬"""
        # 현재 컬럼으로 정렬 중이면 reverse, 아니면 새로 정렬
        if self.stock_sort_column == col:
            self.stock_sort_reverse = not self.stock_sort_reverse
        else:
            self.stock_sort_column = col
            self.stock_sort_reverse = False
        
        # 현재 데이터 가져오기
        data = []
        for item in self.stock_tree.get_children():
            values = self.stock_tree.item(item)['values']
            data.append(values)
        
        # 컬럼 인덱스 찾기
        columns = ("매장", "상품명", "상품코드", "색상", "사이즈", "현재고", "총입고", "총출고", "미입고")
        col_index = columns.index(col)
        
        # 정렬 (숫자 컬럼은 숫자로, 문자 컬럼은 문자로)
        if col in ("현재고", "총입고", "총출고", "미입고"):
            # 숫자 정렬
            data.sort(key=lambda x: int(x[col_index]) if str(x[col_index]).isdigit() else 0, 
                     reverse=self.stock_sort_reverse)
        else:
            # 문자 정렬
            data.sort(key=lambda x: str(x[col_index]), reverse=self.stock_sort_reverse)
        
        # 트리 업데이트
        for item in self.stock_tree.get_children():
            self.stock_tree.delete(item)
        
        for values in data:
            self.stock_tree.insert('', tk.END, values=values)
        
        # 헤딩에 정렬 표시
        for column in columns:
            if column == col:
                heading = f"{column} {'▼' if self.stock_sort_reverse else '▲'}"
            else:
                heading = column
            self.stock_tree.heading(column, text=heading)

    def sort_treeview(self, tree, col):
        """TreeView 컬럼 클릭 시 정렬 (범용 메서드)"""
        # Get current sort order
        current_order = getattr(tree, '_sort_order', {})
        reverse = current_order.get(col, False)

        # Get all items
        items = [(tree.set(item, col), item) for item in tree.get_children('')]

        # Try numeric sort first, fall back to string sort
        try:
            items.sort(key=lambda x: float(x[0].replace(',', '').replace('원', '').replace('장', '').replace('개', '')), reverse=reverse)
        except (ValueError, AttributeError):
            items.sort(key=lambda x: x[0], reverse=reverse)

        # Rearrange items
        for index, (_, item) in enumerate(items):
            tree.move(item, '', index)

        # Toggle sort order for next click
        tree._sort_order = current_order
        tree._sort_order[col] = not reverse

        # Update column headings with sort indicator
        for column in tree['columns']:
            base_col = column.replace(' ▲', '').replace(' ▼', '')
            if column == col:
                indicator = '▼' if reverse else '▲'
                tree.heading(column, text=f"{base_col} {indicator}")
            else:
                tree.heading(column, text=base_col)

    def export_stock_excel(self):
        """재고 현황 엑셀 출력"""
        if not self.products:
            messagebox.showwarning("경고", "출력할 재고 데이터가 없습니다.")
            return
        
        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialfile=f"재고현황_{datetime.now().strftime('%Y%m%d')}.xlsx"
        )
        
        if filename:
            try:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "재고현황"
                
                # 헤더 스타일
                header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF", size=11)
                header_alignment = Alignment(horizontal="center", vertical="center")
                
                border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                headers = ["매장", "상품명", "상품코드", "색상", "사이즈", "현재고", "총입고", "총출고", "미입고"]
                
                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_num)
                    cell.value = header
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.border = border
                    cell.alignment = header_alignment
                
                # 데이터 추가 (창고 + 매장별)
                row_num = 2
                
                # 1. 창고 재고 먼저 추가
                for product in self.products:
                    colors = product.get('colors', [''])
                    sizes = product.get('sizes', ['FREE'])
                    
                    for color in colors:
                        for size in sizes:
                            # 창고 재고 계산
                            warehouse_in = sum(m['quantity'] for m in self.movements 
                                              if m['product_id'] == product['id'] and m['type'] == 'in'
                                              and m.get('color', '') == color and m.get('size', 'FREE') == size
                                              and m.get('store_id') is None)
                            
                            warehouse_out = sum(m['quantity'] for m in self.movements 
                                               if m['product_id'] == product['id'] and m['type'] == 'out'
                                               and m.get('color', '') == color and m.get('size', 'FREE') == size
                                               and m.get('store_id') is None)
                            
                            warehouse_stock = warehouse_in - warehouse_out
                            
                            total_ordered = sum(o['quantity'] for o in self.orders
                                               if o['product_id'] == product['id']
                                               and o.get('color', '') == color
                                               and o.get('size', 'FREE') == size)
                            
                            total_shipped = sum(o['shipped_quantity'] for o in self.orders
                                               if o['product_id'] == product['id']
                                               and o.get('color', '') == color
                                               and o.get('size', 'FREE') == size)
                            
                            warehouse_pending = total_ordered - total_shipped
                            
                            if warehouse_stock > 0 or warehouse_in > 0 or warehouse_out > 0 or warehouse_pending > 0:
                                row_data = [
                                    '🏢 창고',
                                    product['name'],
                                    product.get('code', ''),
                                    color if color else '-',
                                    size,
                                    warehouse_stock,
                                    warehouse_in,
                                    warehouse_out,
                                    warehouse_pending
                                ]
                                
                                for col_num, value in enumerate(row_data, 1):
                                    cell = ws.cell(row=row_num, column=col_num)
                                    cell.value = value
                                    cell.border = border
                                    cell.alignment = Alignment(horizontal="center", vertical="center")
                                
                                row_num += 1
                
                # 2. 매장별 재고 추가
                for store in self.stores:
                    for product in self.products:
                        colors = product.get('colors', [''])
                        sizes = product.get('sizes', ['FREE'])
                        
                        for color in colors:
                            for size in sizes:
                                # 매장 재고 계산
                                store_in = sum(m['quantity'] for m in self.movements 
                                              if m['product_id'] == product['id'] and m['type'] == 'out'
                                              and m.get('color', '') == color and m.get('size', 'FREE') == size
                                              and m.get('store_id') == store['id'])
                                
                                store_stock = store_in
                                
                                store_pending = sum((o['quantity'] - o['shipped_quantity']) for o in self.orders
                                                   if o['product_id'] == product['id']
                                                   and o.get('store_id') == store['id']
                                                   and o.get('color', '') == color
                                                   and o.get('size', 'FREE') == size
                                                   and o.get('status') != 'completed')
                                
                                # 재고나 미입고가 있는 경우만 출력
                                if store_stock > 0 or store_pending > 0:
                                    row_data = [
                                        store['name'],
                                        product['name'],
                                        product.get('code', ''),
                                        color if color else '-',
                                        size,
                                        store_stock,
                                        store_in,
                                        0,  # 매장 출고
                                        store_pending
                                    ]
                                    
                                    for col_num, value in enumerate(row_data, 1):
                                        cell = ws.cell(row=row_num, column=col_num)
                                        cell.value = value
                                        cell.border = border
                                        cell.alignment = Alignment(horizontal="center", vertical="center")
                                    
                                    row_num += 1
                
                # 컬럼 너비 자동 조정
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 30)
                    ws.column_dimensions[column_letter].width = adjusted_width
                
                wb.save(filename)
                messagebox.showinfo("완료", f"재고 현황이 저장되었습니다:\n{filename}")
            except Exception as e:
                messagebox.showerror("오류", f"파일 저장 중 오류가 발생했습니다:\n{str(e)}")
    
    def show_products_menu(self, event):
        try:
            self.products_tree.selection_set(self.products_tree.identify_row(event.y))
            self.products_menu.post(event.x_root, event.y_root)
        finally:
            self.products_menu.grab_release()
    
    def show_orders_menu(self, event):
        try:
            self.orders_tree.selection_set(self.orders_tree.identify_row(event.y))
            self.orders_menu.post(event.x_root, event.y_root)
        finally:
            self.orders_menu.grab_release()
    
    def import_products_from_excel(self):
        """엑셀 파일로 상품 일괄 등록"""
        file_path = filedialog.askopenfilename(
            title="엑셀 파일 선택",
            filetypes=[("Excel files", "*.xlsx *.xls")]
        )
        
        if not file_path:
            return
        
        try:
            import pandas as pd
            df = pd.read_excel(file_path)
            
            # 필수 컬럼 확인
            required_columns = ['상품명']
            for col in required_columns:
                if col not in df.columns:
                    messagebox.showerror("오류", f"필수 컬럼 '{col}'이(가) 없습니다.")
                    return
            
            # 상품명별로 데이터 그룹화
            product_dict = {}
            
            for idx, row in df.iterrows():
                try:
                    name = str(row.get('상품명', '')).strip()
                    if not name or name == 'nan':
                        continue

                    # 상품코드 처리
                    code = str(row.get('상품코드', '')).strip() if str(row.get('상품코드', '')) != 'nan' else ''

                    # 색상 처리
                    colors_str = str(row.get('색상', ''))
                    color = colors_str.strip() if colors_str and colors_str != 'nan' else ''

                    # 사이즈 처리
                    sizes_str = str(row.get('사이즈', 'FREE'))
                    size = sizes_str.strip() if sizes_str and sizes_str != 'nan' else 'FREE'

                    # 자동 분리 기능이 활성화되어 있고, 상품코드가 있는 경우
                    if self.data_manager.get_auto_split_setting() and code:
                        detected_colors, detected_sizes = utils.auto_split_product_code(code)

                        # 색상이 비어있으면 자동 감지된 색상 사용
                        if not color and detected_colors:
                            color = detected_colors[0] if len(detected_colors) > 0 else ''

                        # 사이즈가 FREE이거나 비어있으면 자동 감지된 사이즈 사용
                        if (not size or size == 'FREE') and detected_sizes:
                            size = detected_sizes[0] if len(detected_sizes) > 0 else 'FREE'
                    
                    # 이미지 처리
                    image_base64 = None
                    image_url = str(row.get('이미지url', ''))
                    if image_url and image_url != 'nan' and image_url.lower() != 'none':
                        # URL인 경우 다운로드 시도
                        if image_url.startswith('http'):
                            try:
                                image_base64 = utils.download_image_from_url(image_url)
                                if not image_base64:
                                    print(f"이미지 다운로드 실패: {image_url}")
                            except Exception as e:
                                print(f"이미지 다운로드 오류 ({image_url}): {str(e)}")
                        # 로컬 파일 경로인 경우
                        elif os.path.exists(image_url):
                            try:
                                img = Image.open(image_url)
                                img.thumbnail((300, 300))
                                buffered = BytesIO()
                                img.save(buffered, format="PNG")
                                image_base64 = base64.b64encode(buffered.getvalue()).decode()
                            except Exception as e:
                                print(f"이미지 로드 오류 ({image_url}): {str(e)}")
                    
                    # 상품명으로 그룹화
                    if name not in product_dict:
                        product_dict[name] = {
                            'name': name,
                            'code': str(row.get('상품코드', '')).strip() if str(row.get('상품코드', '')) != 'nan' else '',
                            'supplier': str(row.get('매입처', '')).strip() if str(row.get('매입처', '')) != 'nan' else '',
                            'colors': set(),
                            'sizes': set(),
                            'memo': '',
                            'image': image_base64,
                            'image_source': 'url' if image_base64 and image_url.startswith('http') else 'manual'
                        }
                    
                    # 색상과 사이즈 추가 (set으로 중복 제거)
                    if color:
                        product_dict[name]['colors'].add(color)
                    if size:
                        product_dict[name]['sizes'].add(size)
                    
                    # 이미지가 없었는데 새로 생긴 경우 업데이트
                    if not product_dict[name]['image'] and image_base64:
                        product_dict[name]['image'] = image_base64
                        product_dict[name]['image_source'] = 'url' if image_url.startswith('http') else 'manual'
                    
                except Exception as e:
                    print(f"행 {idx + 2} 처리 중 오류: {str(e)}")
                    continue
            
            # 상품 등록
            imported_count = 0
            updated_count = 0
            
            for name, data in product_dict.items():
                # set을 list로 변환
                colors_list = sorted(list(data['colors'])) if data['colors'] else ['']
                sizes_list = sorted(list(data['sizes'])) if data['sizes'] else ['FREE']
                
                # 자동 분리 기능이 활성화되어 있고 상품코드가 있으며 색상이 여러 개인 경우
                if (self.data_manager.get_auto_split_setting() and 
                    data['code'] and 
                    len(colors_list) > 1 and 
                    colors_list != ['']):
                    
                    # 각 색상별로 별도의 상품 생성
                    for idx, color in enumerate(colors_list, start=1):
                        # 같은 상품명과 상품코드-색상 조합이 있는지 확인
                        split_code = f"{data['code']}-{idx}"
                        existing_product = None
                        for product in self.products:
                            if (product['name'] == data['name'] and 
                                product.get('code', '') == split_code):
                                existing_product = product
                                break
                        
                        if existing_product:
                            # 기존 상품 업데이트 (사이즈 병합)
                            existing_sizes = set(existing_product.get('sizes', ['FREE']))
                            new_sizes = set(sizes_list)
                            existing_product['sizes'] = sorted(list(existing_sizes.union(new_sizes)))
                            
                            # 이미지가 없었으면 업데이트
                            if not existing_product.get('image') and data['image']:
                                existing_product['image'] = data['image']
                                existing_product['image_source'] = data['image_source']
                            
                            updated_count += 1
                        else:
                            # 새 상품 생성 (색상별로 분리)
                            product = {
                                'id': self.data_manager.get_next_product_id(),
                                'name': data['name'],
                                'code': split_code,
                                'supplier': data['supplier'],
                                'colors': [color],  # 각 상품은 하나의 색상만
                                'sizes': sizes_list,
                                'memo': data['memo'],
                                'image': data['image'],
                                'image_source': data['image_source']
                            }
                            
                            self.products.append(product)
                            imported_count += 1
                else:
                    # 자동 분리가 비활성화되어 있거나 색상이 1개 이하인 경우 기존 방식
                    colors_str = ','.join(colors_list)
                    
                    # 중복 체크: 상품명, 상품코드, 색상이 모두 일치하는 상품 찾기
                    existing_product = None
                    for product in self.products:
                        existing_colors_list = sorted(product.get('colors', ['']))
                        existing_colors_str = ','.join(existing_colors_list)
                        
                        if (product['name'] == data['name'] and 
                            product.get('code', '') == data['code'] and
                            existing_colors_str == colors_str):
                            existing_product = product
                            break
                    
                    if existing_product:
                        # 기존 상품 업데이트 (사이즈 병합)
                        existing_sizes = set(existing_product.get('sizes', ['FREE']))
                        new_sizes = set(sizes_list)
                        existing_product['sizes'] = sorted(list(existing_sizes.union(new_sizes)))
                        
                        # 이미지가 없었으면 업데이트
                        if not existing_product.get('image') and data['image']:
                            existing_product['image'] = data['image']
                            existing_product['image_source'] = data['image_source']
                        
                        updated_count += 1
                    else:
                        # 새 상품 생성
                        product = {
                            'id': self.data_manager.get_next_product_id(),
                            'name': data['name'],
                            'code': data['code'],
                            'supplier': data['supplier'],
                            'colors': colors_list,
                            'sizes': sizes_list,
                            'memo': data['memo'],
                            'image': data['image'],
                            'image_source': data['image_source']
                        }
                        
                        self.products.append(product)
                        imported_count += 1
            
            # self.data_manager.save_data()
            # ← 자동저장 제거됨
            self._refresh_data_shortcuts()
            self.refresh_products_list()
            self.refresh_stock_list()
            
            message = f"✅ 신규 등록: {imported_count}개\n✅ 업데이트: {updated_count}개"
            if imported_count > 0 or updated_count > 0:
                message += "\n\n💡 같은 상품명의 다른 색상/사이즈는 자동으로 하나의 상품으로 통합되었습니다."
            
            messagebox.showinfo("완료", message)
            
        except Exception as e:
            messagebox.showerror("오류", f"엑셀 파일 처리 중 오류가 발생했습니다:\n{str(e)}")
    
    def export_products_to_excel(self):
        """상품 목록을 엑셀로 저장"""
        if not self.products:
            messagebox.showwarning("경고", "저장할 상품이 없습니다.")
            return
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        default_filename = f"상품목록_{timestamp}.xlsx"
        
        filename = filedialog.asksaveasfilename(
            title="상품 목록 저장",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=default_filename
        )
        
        if not filename:
            return
        
        try:
            # 엑셀 데이터 준비
            data = []
            for product in self.products:
                colors = product.get('colors', [''])
                sizes = product.get('sizes', ['FREE'])
                
                # 색상/사이즈 조합별로 행 생성
                for color in colors:
                    for size in sizes:
                        row = {
                            '상품명': product['name'],
                            '상품코드': product.get('code', ''),
                            '매입처': product.get('supplier', ''),
                            '색상': color if color else '',
                            '사이즈': size,
                        }
                        
                        # 재고 정보 추가
                        stock = self.data_manager.calculate_stock_by_variant(product['id'], color, size)
                        pending = self.data_manager.calculate_pending_by_variant(product['id'], color, size)
                        
                        row['현재고'] = stock
                        row['미입고'] = pending
                        
                        data.append(row)
            
            # DataFrame 생성 및 저장
            import pandas as pd
            df = pd.DataFrame(data)
            
            # 엑셀 저장 (스타일 포함)
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='상품목록')
                
                # 워크시트 가져오기
                workbook = writer.book
                worksheet = writer.sheets['상품목록']
                
                # 헤더 스타일링
                header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF", size=11)
                header_alignment = Alignment(horizontal="center", vertical="center")
                
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_alignment
                
                # 컬럼 너비 자동 조정
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 30)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
            
            messagebox.showinfo("완료", 
                f"상품 목록이 저장되었습니다!\n\n"
                f"파일: {filename}\n"
                f"총 {len(data)}개 항목 (색상/사이즈별)")
            
        except Exception as e:
            messagebox.showerror("오류", f"엑셀 저장 중 오류가 발생했습니다:\n{str(e)}")
    
    def search_cloud_image(self):
        """클라우드 이미지 폴더에서 상품 이미지 검색 (로컬 + 클라우드 스토리지)"""
        selected = self.products_tree.selection()
        if not selected:
            messagebox.showwarning("경고", "상품을 선택해주세요.")
            return
        
        item = self.products_tree.item(selected[0])
        product_name = item['values'][0]
        
        product = None
        for p in self.products:
            if p['name'] == product_name:
                product = p
                break
        
        if not product:
            return
        
        # 이미지 검색 경로 목록
        from pathlib import Path
        search_paths = []
        
        # 1. 로컬 cloud_images 폴더
        if os.path.exists(self.cloud_image_folder):
            search_paths.append(("로컬", self.cloud_image_folder))
        
        # 2. 클라우드 스토리지의 images 폴더
        if self.cloud_path and os.path.exists(self.cloud_path):
            cloud_images_path = os.path.join(self.cloud_path, "images")
            if os.path.exists(cloud_images_path):
                search_paths.append((f"{self.cloud_type} 클라우드", cloud_images_path))
        
        # 이미지 파일 검색
        image_files = []
        for location_name, path in search_paths:
            for ext in ['*.jpg', '*.jpeg', '*.png', '*.gif', '*.bmp', '*.webp']:
                for img_file in Path(path).glob(ext):
                    image_files.append((location_name, img_file))
        
        if not image_files:
            msg = "이미지를 찾을 수 없습니다.\n\n검색한 위치:\n"
            for location_name, path in search_paths:
                msg += f"• {location_name}: {path}\n"
            if not search_paths:
                msg += "• 로컬: cloud_images 폴더 (없음)\n"
                if self.cloud_path:
                    msg += f"• {self.cloud_type} 클라우드: {os.path.join(self.cloud_path, 'images')} (없음)\n"
            msg += "\n이미지를 해당 폴더에 저장해주세요."
            messagebox.showinfo("정보", msg)
            return
        
        # 이미지 선택 다이얼로그
        dialog = tk.Toplevel(self.root)
        dialog.title(f"이미지 검색 - {product['name']}")
        dialog.geometry("900x700")
        dialog.transient(self.root)
        dialog.grab_set()
        utils.center_window(dialog)
        
        ttk.Label(dialog, text=f"상품: {product['name']}", font=("Arial", 12, "bold")).pack(pady=10)
        ttk.Label(dialog, text=f"{len(image_files)}개의 이미지 발견", font=("Arial", 10)).pack()
        
        # 스크롤 가능한 프레임
        canvas = tk.Canvas(dialog)
        scrollbar = ttk.Scrollbar(dialog, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)
        
        scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # 이미지 그리드 생성
        row, col = 0, 0
        max_cols = 4
        
        for location_name, img_path in image_files:
            try:
                img = Image.open(img_path)
                img.thumbnail((180, 180))
                photo = ImageTk.PhotoImage(img)
                
                frame = ttk.Frame(scrollable_frame, relief=tk.RIDGE, borderwidth=2)
                frame.grid(row=row, column=col, padx=8, pady=8)
                
                img_label = ttk.Label(frame, image=photo)
                img_label.image = photo
                img_label.pack()
                
                # 파일명과 위치 표시
                ttk.Label(frame, text=f"[{location_name}]", font=("Arial", 8), foreground="blue").pack()
                ttk.Label(frame, text=img_path.name, wraplength=160, font=("Arial", 8)).pack()
                
                def select_image(path=img_path, loc=location_name):
                    try:
                        img = Image.open(path)
                        img.thumbnail((300, 300))
                        buffered = BytesIO()
                        img.save(buffered, format="PNG")
                        img_str = base64.b64encode(buffered.getvalue()).decode()
                        
                        product['image'] = img_str
                        # self.data_manager.save_data()
                        # ← 자동저장 제거됨
                        self._refresh_data_shortcuts()
                        self.refresh_products_list()
                        dialog.destroy()
                        messagebox.showinfo("완료", f"이미지가 적용되었습니다.\n위치: {loc}")
                    except Exception as e:
                        messagebox.showerror("오류", f"이미지 적용 실패:\n{str(e)}")
                
                ttk.Button(frame, text="선택", command=select_image).pack(pady=5)
                
                col += 1
                if col >= max_cols:
                    col = 0
                    row += 1
                    
            except Exception as e:
                print(f"이미지 로드 실패 ({img_path}): {str(e)}")
                continue
        
        canvas.pack(side="left", fill="both", expand=True, padx=10, pady=10)
        scrollbar.pack(side="right", fill="y", pady=10)
        
        ttk.Button(dialog, text="닫기", command=dialog.destroy).pack(pady=10)
    
    def reset_order_date_filter(self):
        """발주 날짜 필터 초기화"""
        self.order_start_date_var.set("")
        self.order_end_date_var.set("")
        self.refresh_orders_list()
    
    def show_store_orders(self):
        """매장별 발주 현황 보기"""
        dialog = tk.Toplevel(self.root)
        dialog.title("매장별 발주 현황")
        dialog.geometry("800x600")
        dialog.transient(self.root)
        dialog.grab_set()
        utils.center_window(dialog)
        
        ttk.Label(dialog, text="매장별 발주 현황", font=("Arial", 14, "bold")).pack(pady=10)
        
        columns = ("매장명", "상품명", "색상", "사이즈", "총발주수량", "총출고수량", "미입고수량")
        tree = ttk.Treeview(dialog, columns=columns, show="headings", height=20)
        
        for col in columns:
            tree.heading(col, text=col)
            tree.column(col, width=100)
        
        scrollbar = ttk.Scrollbar(dialog, orient=tk.VERTICAL, command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        
        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(10,0), pady=10)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y, padx=(0,10), pady=10)
        
        # 매장별 발주 집계
        store_orders = {}
        for order in self.orders:
            store_id = order.get('store_id', '')
            if not store_id:
                continue
            
            key = (store_id, order['product_id'], order.get('color', ''), order.get('size', 'FREE'))
            
            if key not in store_orders:
                store_orders[key] = {
                    'store_id': store_id,
                    'product_id': order['product_id'],
                    'color': order.get('color', ''),
                    'size': order.get('size', 'FREE'),
                    'total_quantity': 0,
                    'total_shipped': 0
                }
            
            store_orders[key]['total_quantity'] += order['quantity']
            store_orders[key]['total_shipped'] += order['shipped_quantity']
        
        # 트리에 데이터 추가
        for key, data in store_orders.items():
            store = self.data_manager.get_store_by_id(data['store_id'])
            product = self.data_manager.get_product_by_id(data['product_id'])
            
            if store and product:
                pending = data['total_quantity'] - data['total_shipped']
                tree.insert('', tk.END, values=(
                    store['name'],
                    product['name'],
                    data['color'] if data['color'] else '-',
                    data['size'],
                    data['total_quantity'],
                    data['total_shipped'],
                    pending
                ))
        
        ttk.Button(dialog, text="닫기", command=dialog.destroy).pack(pady=10)
    
    def create_stores_tab(self):
        """매장 관리 탭 생성"""
        top_frame = ttk.Frame(self.stores_frame)
        top_frame.pack(fill=tk.X, padx=10, pady=10)
        
        ttk.Label(top_frame, text="매장 관리", font=("Arial", 12, "bold")).pack(side=tk.LEFT, padx=5)
        ttk.Button(top_frame, text="➕ 매장 추가", command=self.add_store).pack(side=tk.RIGHT, padx=5)
        
        columns = ("매장명", "매장코드", "주소", "연락처")
        self.stores_tree = ttk.Treeview(self.stores_frame, columns=columns, show="headings", height=20)
        
        for col in columns:
            self.stores_tree.heading(col, text=col)
            self.stores_tree.column(col, width=150)
        
        scrollbar = ttk.Scrollbar(self.stores_frame, orient=tk.VERTICAL, command=self.stores_tree.yview)
        self.stores_tree.configure(yscrollcommand=scrollbar.set)
        
        self.stores_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(10,0), pady=10)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y, padx=(0,10), pady=10)
        
        self.stores_menu = tk.Menu(self.root, tearoff=0)
        self.stores_menu.add_command(label="수정", command=self.edit_store)
        self.stores_menu.add_command(label="삭제", command=self.delete_store)
        self.stores_tree.bind("<Button-3>", self.show_stores_menu)
        
        self.refresh_stores_list()
    
    def add_store(self):
        """매장 추가 다이얼로그"""
        dialog = tk.Toplevel(self.root)
        dialog.title("매장 추가")
        dialog.geometry("500x350")
        dialog.transient(self.root)
        dialog.grab_set()
        utils.center_window(dialog)
        
        ttk.Label(dialog, text="매장명 *", font=("Arial", 10, "bold")).pack(pady=(10,0))
        name_var = tk.StringVar()
        ttk.Entry(dialog, textvariable=name_var, width=50).pack(pady=5)
        
        ttk.Label(dialog, text="매장코드", font=("Arial", 10)).pack(pady=(10,0))
        code_var = tk.StringVar()
        ttk.Entry(dialog, textvariable=code_var, width=50).pack(pady=5)
        
        ttk.Label(dialog, text="주소", font=("Arial", 10)).pack(pady=(10,0))
        address_var = tk.StringVar()
        ttk.Entry(dialog, textvariable=address_var, width=50).pack(pady=5)
        
        ttk.Label(dialog, text="연락처", font=("Arial", 10)).pack(pady=(10,0))
        phone_var = tk.StringVar()
        ttk.Entry(dialog, textvariable=phone_var, width=50).pack(pady=5)
        
        def save_store():
            name = name_var.get().strip()
            if not name:
                messagebox.showerror("오류", "매장명을 입력해주세요.")
                return
            
            store = {
                'id': str(len(self.stores) + 1),
                'name': name,
                'code': code_var.get().strip(),
                'address': address_var.get().strip(),
                'phone': phone_var.get().strip()
            }
            
            self.stores.append(store)
            # self.data_manager.save_data()
            # ← 자동저장 제거됨
            self._refresh_data_shortcuts()
            self.refresh_stores_list()
            dialog.destroy()
            messagebox.showinfo("완료", "매장이 추가되었습니다.")
        
        button_frame = ttk.Frame(dialog)
        button_frame.pack(pady=20)
        ttk.Button(button_frame, text="취소", command=dialog.destroy).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="저장", command=save_store).pack(side=tk.LEFT, padx=5)
    
    def edit_store(self):
        """매장 수정"""
        selected = self.stores_tree.selection()
        if not selected:
            messagebox.showwarning("경고", "수정할 매장을 선택해주세요.")
            return
        
        item = self.stores_tree.item(selected[0])
        store_name = item['values'][0]
        
        store = None
        for s in self.stores:
            if s['name'] == store_name:
                store = s
                break
        
        if not store:
            return
        
        dialog = tk.Toplevel(self.root)
        dialog.title("매장 수정")
        dialog.geometry("500x350")
        dialog.transient(self.root)
        dialog.grab_set()
        utils.center_window(dialog)
        
        ttk.Label(dialog, text="매장명 *", font=("Arial", 10, "bold")).pack(pady=(10,0))
        name_var = tk.StringVar(value=store.get('name', ''))
        ttk.Entry(dialog, textvariable=name_var, width=50).pack(pady=5)
        
        ttk.Label(dialog, text="매장코드", font=("Arial", 10)).pack(pady=(10,0))
        code_var = tk.StringVar(value=store.get('code', ''))
        ttk.Entry(dialog, textvariable=code_var, width=50).pack(pady=5)
        
        ttk.Label(dialog, text="주소", font=("Arial", 10)).pack(pady=(10,0))
        address_var = tk.StringVar(value=store.get('address', ''))
        ttk.Entry(dialog, textvariable=address_var, width=50).pack(pady=5)
        
        ttk.Label(dialog, text="연락처", font=("Arial", 10)).pack(pady=(10,0))
        phone_var = tk.StringVar(value=store.get('phone', ''))
        ttk.Entry(dialog, textvariable=phone_var, width=50).pack(pady=5)
        
        def save_store():
            name = name_var.get().strip()
            if not name:
                messagebox.showerror("오류", "매장명을 입력해주세요.")
                return
            
            store['name'] = name
            store['code'] = code_var.get().strip()
            store['address'] = address_var.get().strip()
            store['phone'] = phone_var.get().strip()
            
            # self.data_manager.save_data()
            # ← 자동저장 제거됨
            self._refresh_data_shortcuts()
            self.refresh_stores_list()
            dialog.destroy()
            messagebox.showinfo("완료", "매장이 수정되었습니다.")
        
        button_frame = ttk.Frame(dialog)
        button_frame.pack(pady=20)
        ttk.Button(button_frame, text="취소", command=dialog.destroy).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="저장", command=save_store).pack(side=tk.LEFT, padx=5)
    
    def delete_store(self):
        """매장 삭제"""
        selected = self.stores_tree.selection()
        if not selected:
            messagebox.showwarning("경고", "삭제할 매장을 선택해주세요.")
            return
        
        if messagebox.askyesno("확인", "정말 이 매장을 삭제하시겠습니까?"):
            item = self.stores_tree.item(selected[0])
            store_name = item['values'][0]
            
            self.stores = [s for s in self.stores if s['name'] != store_name]
            # self.data_manager.save_data()
            # ← 자동저장 제거됨
            self._refresh_data_shortcuts()
            self.refresh_stores_list()
            messagebox.showinfo("완료", "매장이 삭제되었습니다.")
    
    def refresh_stores_list(self):
        """매장 목록 새로고침"""
        for item in self.stores_tree.get_children():
            self.stores_tree.delete(item)
        
        for store in self.stores:
            self.stores_tree.insert('', tk.END, values=(
                store['name'],
                store.get('code', ''),
                store.get('address', ''),
                store.get('phone', '')
            ))
    
    def change_field_names(self):
        """필드명 관리 다이얼로그 (추가/삭제/변경)"""
        dialog = tk.Toplevel(self.root)
        dialog.title("필드명 관리")
        dialog.geometry("700x800")
        dialog.transient(self.root)
        dialog.grab_set()
        utils.center_window(dialog)
        
        # 안내
        info_frame = ttk.LabelFrame(dialog, text="ℹ️ 필드명 관리 안내", padding=15)
        info_frame.pack(fill=tk.X, padx=20, pady=20)
        
        info_text = """업종에 맞게 필드명을 변경하거나 추가하세요!

• 의류: 색상, 사이즈
• 식품: 맛, 용량, 포장
• 전자제품: 색상, 용량, 모델
• 도서: 분류, 판형, 언어
• 가구: 색상, 규격, 재질

필드를 추가하려면 '➕ 필드 추가' 버튼을 클릭하세요!"""
        
        ttk.Label(info_frame, text=info_text, justify=tk.LEFT, font=("Arial", 9)).pack()
        
        # 현재 필드 목록
        fields_frame = ttk.LabelFrame(dialog, text="현재 필드 목록", padding=15)
        fields_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)
        
        # 스크롤 가능한 필드 리스트
        canvas = tk.Canvas(fields_frame, height=300)
        scrollbar = ttk.Scrollbar(fields_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)
        
        scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas_frame = canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # Canvas 너비를 스크롤 프레임에 맞춤
        def configure_canvas_width(event):
            canvas.itemconfig(canvas_frame, width=event.width)
        canvas.bind('<Configure>', configure_canvas_width)
        
        # 필드 위젯 저장용
        field_entries = []
        
        def add_field_row(field_id, field_name, is_new=False):
            """필드 행 추가"""
            row_frame = ttk.Frame(scrollable_frame)
            row_frame.pack(fill=tk.X, pady=5, padx=10)
            
            row_num = len(field_entries) + 1
            ttk.Label(row_frame, text=f"필드 {row_num}:", width=10).pack(side=tk.LEFT, padx=5)
            
            field_var = tk.StringVar(value=field_name)
            field_entry = ttk.Entry(row_frame, textvariable=field_var, width=35)
            field_entry.pack(side=tk.LEFT, padx=5)
            
            entry_info = {'id': field_id, 'var': field_var, 'frame': row_frame}
            field_entries.append(entry_info)
            
            # 삭제 버튼 (최소 1개는 남겨야 함)
            def delete_field():
                if len(field_entries) > 1:
                    field_entries.remove(entry_info)
                    row_frame.destroy()
                    # 번호 다시 매기기
                    renumber_fields()
                else:
                    messagebox.showwarning("경고", "최소 1개의 필드는 필요합니다.")
            
            ttk.Button(row_frame, text="❌", command=delete_field, width=3).pack(side=tk.LEFT, padx=5)
            
            if is_new:
                field_entry.focus_set()
        
        def renumber_fields():
            """필드 번호 다시 매기기"""
            for idx, entry_info in enumerate(field_entries, 1):
                # 라벨 텍스트 업데이트
                for widget in entry_info['frame'].winfo_children():
                    if isinstance(widget, ttk.Label):
                        widget.config(text=f"필드 {idx}:")
                        break
        
        # 기존 필드 표시
        for field in self.field_names:
            add_field_row(field['id'], field['name'])
        
        def add_new_field():
            """새 필드 추가"""
            new_id = f"field{int(datetime.now().timestamp() * 1000)}"
            add_field_row(new_id, f"속성{len(field_entries) + 1}", is_new=True)
        
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # 필드 추가 버튼
        add_button_frame = ttk.Frame(dialog)
        add_button_frame.pack(fill=tk.X, padx=20, pady=10)
        ttk.Button(add_button_frame, text="➕ 필드 추가", command=add_new_field, width=20).pack()
        
        def save_changes():
            # 필드명 수집 (비어있거나 중복된 것 제외)
            new_fields = []
            seen_names = set()
            
            for entry_info in field_entries:
                field_name = entry_info['var'].get().strip()
                
                if not field_name:
                    continue  # 비어있으면 건너뛰기
                
                # 중복 체크
                if field_name in seen_names:
                    messagebox.showwarning("경고", f"'{field_name}'이(가) 중복됩니다. 중복된 필드명은 무시됩니다.")
                    continue
                
                seen_names.add(field_name)
                new_fields.append({
                    'id': entry_info['id'],
                    'name': field_name
                })
            
            if len(new_fields) == 0:
                messagebox.showwarning("경고", "최소 1개의 필드명을 입력해주세요.")
                return
            
            self.field_names = new_fields
            # self.data_manager.save_data()
            # ← 자동저장 제거됨
            self._refresh_data_shortcuts()
            
            # UI 새로고침 (상품 탭 다시 생성)
            self.notebook.forget(0)  # 상품 관리 탭 제거
            self.products_frame = ttk.Frame(self.notebook)
            self.notebook.insert(0, self.products_frame, text="📦 상품 관리")
            self.create_products_tab()
            self.notebook.select(0)  # 상품 관리 탭 선택
            
            field_summary = "\n".join([f"• {f['name']}" for f in new_fields])
            messagebox.showinfo("완료", 
                f"필드명이 저장되었습니다!\n\n"
                f"{field_summary}\n\n"
                f"상품 관리 화면이 업데이트되었습니다.")
            dialog.destroy()
        
        button_frame = ttk.Frame(dialog)
        button_frame.pack(pady=20)
        ttk.Button(button_frame, text="취소", command=dialog.destroy, width=15).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="저장", command=save_changes, width=15).pack(side=tk.LEFT, padx=5)
    
    def show_stores_menu(self, event):
        try:
            self.stores_tree.selection_set(self.stores_tree.identify_row(event.y))
            self.stores_menu.post(event.x_root, event.y_root)
        finally:
            self.stores_menu.grab_release()

    def upload_settlement_sheet(self):
        """출고장 업로드 (날짜 지정)"""
        # 날짜 선택 다이얼로그
        dialog = tk.Toplevel(self.root)
        dialog.title("출고장 업로드")
        dialog.geometry("500x300")
        dialog.transient(self.root)
        dialog.grab_set()
        utils.center_window(dialog)

        ttk.Label(dialog, text="출고장 엑셀 파일 업로드", font=("Arial", 14, "bold")).pack(pady=20)

        # 날짜 선택
        date_frame = ttk.Frame(dialog)
        date_frame.pack(pady=10)
        ttk.Label(date_frame, text="출고 날짜:").pack(side=tk.LEFT, padx=5)
        date_var = tk.StringVar(value=datetime.now().strftime('%Y-%m-%d'))
        date_entry = DateEntry(date_frame, textvariable=date_var, date_pattern='yyyy-mm-dd')
        date_entry.pack(side=tk.LEFT, padx=5)

        # 파일 선택
        file_path_var = tk.StringVar()
        file_frame = ttk.Frame(dialog)
        file_frame.pack(pady=10, fill=tk.X, padx=20)
        ttk.Label(file_frame, text="파일:").pack(side=tk.LEFT, padx=5)
        ttk.Entry(file_frame, textvariable=file_path_var, width=30).pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)

        def browse_file():
            filename = filedialog.askopenfilename(
                title="출고장 엑셀 파일 선택",
                filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
            )
            if filename:
                file_path_var.set(filename)

        ttk.Button(file_frame, text="찾아보기", command=browse_file).pack(side=tk.LEFT, padx=5)

        # 업로드 버튼
        def process_upload():
            file_path = file_path_var.get()
            upload_date = date_var.get()

            if not file_path:
                messagebox.showwarning("경고", "파일을 선택해주세요.")
                return

            if not os.path.exists(file_path):
                messagebox.showerror("오류", "파일을 찾을 수 없습니다.")
                return

            try:
                # 엑셀 파일 읽기
                wb = openpyxl.load_workbook(file_path)
                ws = wb.active

                # 출고 데이터 파싱 (예: 첫 행은 헤더, 이후 데이터)
                # 형식: 매장명, 상품명, 상품코드, 색상, 사이즈, 출고수량, 금액
                headers = [cell.value for cell in ws[1]]

                processed_count = 0
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row[0]:  # 매장명이 없으면 스킵
                        continue

                    # 출고 기록 생성
                    outbound_record = {
                        'date': upload_date,
                        'store_name': str(row[0]) if row[0] else '',
                        'product_name': str(row[1]) if len(row) > 1 and row[1] else '',
                        'product_code': str(row[2]) if len(row) > 2 and row[2] else '',
                        'color': str(row[3]) if len(row) > 3 and row[3] else '',
                        'size': str(row[4]) if len(row) > 4 and row[4] else '',
                        'quantity': int(row[5]) if len(row) > 5 and row[5] else 0,
                        'amount': float(row[6]) if len(row) > 6 and row[6] else 0,
                        'note': f'출고장 업로드 ({upload_date})'
                    }

                    # 출고 기록 추가
                    self.outbound_records.append(outbound_record)
                    processed_count += 1

                    # 매장별 잔액 업데이트 (금액 차감)
                    store_name = outbound_record['store_name']
                    amount = outbound_record['amount']
                    if store_name not in self.settlement_balances:
                        self.settlement_balances[store_name] = 0
                    self.settlement_balances[store_name] -= amount

                # 데이터 저장
                self.data_manager.outbound_records = self.outbound_records
                self.data_manager.settlement_balances = self.settlement_balances
                # self.data_manager.save_data()
                # ← 자동저장 제거됨
                self._refresh_data_shortcuts()

                # UI 새로고침
                self.refresh_outbound_list()

                messagebox.showinfo("완료", f"{processed_count}개의 출고 기록이 추가되었습니다.")
                dialog.destroy()

            except Exception as e:
                messagebox.showerror("오류", f"출고장 처리 중 오류가 발생했습니다:\n{str(e)}")

        button_frame = ttk.Frame(dialog)
        button_frame.pack(pady=20)
        ttk.Button(button_frame, text="취소", command=dialog.destroy, width=15).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="업로드", command=process_upload, width=15).pack(side=tk.LEFT, padx=5)

    def manage_settlement_balances(self):
        """잔액 확인/수정"""
        dialog = tk.Toplevel(self.root)
        dialog.title("정산 잔액 관리")
        dialog.geometry("800x600")
        dialog.transient(self.root)
        dialog.grab_set()
        utils.center_window(dialog)

        ttk.Label(dialog, text="매장별 정산 잔액", font=("Arial", 14, "bold")).pack(pady=20)

        # 잔액 테이블
        columns = ("매장명", "잔액", "상태")
        tree = ttk.Treeview(dialog, columns=columns, show="headings", height=15)

        tree.heading("매장명", text="매장명")
        tree.heading("잔액", text="잔액 (원)")
        tree.heading("상태", text="상태")

        tree.column("매장명", width=300, anchor='center')
        tree.column("잔액", width=200, anchor='e')
        tree.column("상태", width=150, anchor='center')

        scrollbar = ttk.Scrollbar(dialog, orient=tk.VERTICAL, command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)

        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(20, 0), pady=10)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y, padx=(0, 20), pady=10)

        def refresh_balance_list():
            tree.delete(*tree.get_children())
            for store_name, balance in sorted(self.settlement_balances.items()):
                status = "정상" if balance >= 0 else "마이너스"
                tree.insert('', 'end', values=(store_name, f"{balance:,.0f}", status))

        refresh_balance_list()

        # 버튼 프레임
        button_frame = ttk.Frame(dialog)
        button_frame.pack(pady=10, fill=tk.X, padx=20)

        def add_balance():
            """잔액 추가/수정"""
            selected = tree.selection()
            if not selected:
                # 새 매장 추가
                add_dialog = tk.Toplevel(dialog)
                add_dialog.title("잔액 추가")
                add_dialog.geometry("400x200")
                add_dialog.transient(dialog)
                add_dialog.grab_set()
                utils.center_window(add_dialog)

                ttk.Label(add_dialog, text="매장명:").pack(pady=5)
                store_var = tk.StringVar()
                ttk.Entry(add_dialog, textvariable=store_var, width=30).pack(pady=5)

                ttk.Label(add_dialog, text="잔액:").pack(pady=5)
                balance_var = tk.StringVar(value="0")
                ttk.Entry(add_dialog, textvariable=balance_var, width=30).pack(pady=5)

                def save_balance():
                    store_name = store_var.get().strip()
                    try:
                        balance = float(balance_var.get())
                        if store_name:
                            self.settlement_balances[store_name] = balance
                            self.data_manager.settlement_balances = self.settlement_balances
                            # self.data_manager.save_data()
                            # ← 자동저장 제거됨
                            self._refresh_data_shortcuts()
                            refresh_balance_list()
                            add_dialog.destroy()
                        else:
                            messagebox.showwarning("경고", "매장명을 입력해주세요.")
                    except ValueError:
                        messagebox.showerror("오류", "잔액은 숫자로 입력해주세요.")

                ttk.Button(add_dialog, text="저장", command=save_balance).pack(pady=10)
            else:
                # 기존 잔액 수정
                item = tree.item(selected[0])
                store_name = item['values'][0]
                current_balance = self.settlement_balances.get(store_name, 0)

                edit_dialog = tk.Toplevel(dialog)
                edit_dialog.title("잔액 수정")
                edit_dialog.geometry("400x200")
                edit_dialog.transient(dialog)
                edit_dialog.grab_set()
                utils.center_window(edit_dialog)

                ttk.Label(edit_dialog, text=f"매장: {store_name}").pack(pady=10)
                ttk.Label(edit_dialog, text="새 잔액:").pack(pady=5)
                balance_var = tk.StringVar(value=str(current_balance))
                ttk.Entry(edit_dialog, textvariable=balance_var, width=30).pack(pady=5)

                def save_balance():
                    try:
                        balance = float(balance_var.get())
                        self.settlement_balances[store_name] = balance
                        self.data_manager.settlement_balances = self.settlement_balances
                        # self.data_manager.save_data()
                        # ← 자동저장 제거됨
                        self._refresh_data_shortcuts()
                        refresh_balance_list()
                        edit_dialog.destroy()
                    except ValueError:
                        messagebox.showerror("오류", "잔액은 숫자로 입력해주세요.")

                ttk.Button(edit_dialog, text="저장", command=save_balance).pack(pady=10)

        def delete_balance():
            """잔액 삭제"""
            selected = tree.selection()
            if not selected:
                messagebox.showwarning("경고", "삭제할 매장을 선택해주세요.")
                return

            item = tree.item(selected[0])
            store_name = item['values'][0]

            if messagebox.askyesno("확인", f"'{store_name}'의 잔액 기록을 삭제하시겠습니까?"):
                if store_name in self.settlement_balances:
                    del self.settlement_balances[store_name]
                    self.data_manager.settlement_balances = self.settlement_balances
                    # self.data_manager.save_data()
                    # ← 자동저장 제거됨
                    self._refresh_data_shortcuts()
                    refresh_balance_list()

        ttk.Button(button_frame, text="추가/수정", command=add_balance).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="삭제", command=delete_balance).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="새로고침", command=refresh_balance_list).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="닫기", command=dialog.destroy).pack(side=tk.RIGHT, padx=5)

    def export_order_template(self):
        """발주 엑셀 양식 저장"""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        default_filename = f"발주양식_{timestamp}.xlsx"

        filename = filedialog.asksaveasfilename(
            title="발주 양식 저장",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=default_filename
        )

        if not filename:
            return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "발주 양식"

            # 헤더 설정
            headers = ["발주일자", "상품명", "상품코드", "색상", "사이즈", "발주매장", "발주수량"]
            ws.append(headers)

            # 헤더 스타일
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # 예시 데이터 추가
            example_row = [
                datetime.now().strftime('%Y-%m-%d'),
                "상품명 예시",
                "코드123",
                "빨강",
                "M",
                "본점" if self.stores else "매장명",
                "10"
            ]
            ws.append(example_row)

            # 컬럼 너비 조정
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 25
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 12
            ws.column_dimensions['E'].width = 10
            ws.column_dimensions['F'].width = 15
            ws.column_dimensions['G'].width = 12

            wb.save(filename)
            messagebox.showinfo("완료", f"발주 양식이 저장되었습니다:\n{filename}")

        except Exception as e:
            messagebox.showerror("오류", f"양식 저장 중 오류가 발생했습니다:\n{str(e)}")

    def export_inbound_template(self):
        """입고 엑셀 양식 저장"""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        default_filename = f"입고양식_{timestamp}.xlsx"

        filename = filedialog.asksaveasfilename(
            title="입고 양식 저장",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=default_filename
        )

        if not filename:
            return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "입고 양식"

            # 헤더 설정
            headers = ["날짜", "상품명", "상품코드", "색상", "사이즈", "수량", "비고"]
            ws.append(headers)

            # 헤더 스타일
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # 예시 데이터 추가
            example_row = [
                datetime.now().strftime('%Y-%m-%d'),
                "상품명 예시",
                "코드123",
                "빨강",
                "M",
                "20",
                "비고 예시"
            ]
            ws.append(example_row)

            # 컬럼 너비 조정
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 25
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 12
            ws.column_dimensions['E'].width = 10
            ws.column_dimensions['F'].width = 10
            ws.column_dimensions['G'].width = 30

            wb.save(filename)
            messagebox.showinfo("완료", f"입고 양식이 저장되었습니다:\n{filename}")

        except Exception as e:
            messagebox.showerror("오류", f"양식 저장 중 오류가 발생했습니다:\n{str(e)}")

    def export_outbound_template(self):
        """출고 엑셀 양식 저장"""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        default_filename = f"출고양식_{timestamp}.xlsx"

        filename = filedialog.asksaveasfilename(
            title="출고 양식 저장",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=default_filename
        )

        if not filename:
            return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "출고 양식"

            # 헤더 설정
            headers = ["날짜", "상품명", "상품코드", "색상", "사이즈", "수량", "비고"]
            ws.append(headers)

            # 헤더 스타일
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # 예시 데이터 추가
            example_row = [
                datetime.now().strftime('%Y-%m-%d'),
                "상품명 예시",
                "코드123",
                "빨강",
                "M",
                "15",
                "비고 예시"
            ]
            ws.append(example_row)

            # 컬럼 너비 조정
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 25
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 12
            ws.column_dimensions['E'].width = 10
            ws.column_dimensions['F'].width = 10
            ws.column_dimensions['G'].width = 30

            wb.save(filename)
            messagebox.showinfo("완료", f"출고 양식이 저장되었습니다:\n{filename}")

        except Exception as e:
            messagebox.showerror("오류", f"양식 저장 중 오류가 발생했습니다:\n{str(e)}")

    def import_order_excel(self):
        """엑셀 파일로 발주 일괄 등록"""
        file_path = filedialog.askopenfilename(
            title="엑셀 파일 선택",
            filetypes=[("Excel files", "*.xlsx *.xls")]
        )

        if not file_path:
            return

        try:
            import pandas as pd
            df = pd.read_excel(file_path)

            # 필수 컬럼 확인
            required_columns = ['발주일자', '상품명', '발주수량']
            missing_columns = [col for col in required_columns if col not in df.columns]
            if missing_columns:
                messagebox.showerror("오류", f"필수 컬럼이 없습니다: {', '.join(missing_columns)}\n\n필수 컬럼: {', '.join(required_columns)}")
                return

            imported_count = 0
            skipped_count = 0
            error_rows = []

            for idx, row in df.iterrows():
                try:
                    # 발주일자 처리
                    order_date_raw = row.get('발주일자')
                    if pd.isna(order_date_raw):
                        error_rows.append(f"행 {idx + 2}: 발주일자 없음")
                        skipped_count += 1
                        continue

                    if isinstance(order_date_raw, str):
                        order_date = order_date_raw
                    else:
                        order_date = order_date_raw.strftime('%Y-%m-%d')

                    # 상품명 처리
                    product_name = str(row.get('상품명', '')).strip()
                    if not product_name or product_name == 'nan':
                        error_rows.append(f"행 {idx + 2}: 상품명 없음")
                        skipped_count += 1
                        continue

                    # 상품 찾기 (상품명 또는 상품코드로)
                    product_code = str(row.get('상품코드', '')).strip() if '상품코드' in df.columns else ''
                    product = None

                    for p in self.products:
                        if p['name'] == product_name:
                            product = p
                            break
                        elif product_code and p.get('code', '') == product_code:
                            product = p
                            break

                    if not product:
                        error_rows.append(f"행 {idx + 2}: 상품 '{product_name}' 찾을 수 없음")
                        skipped_count += 1
                        continue

                    # 색상, 사이즈 처리
                    color = str(row.get('색상', '')).strip() if '색상' in df.columns and not pd.isna(row.get('색상')) else ''
                    size = str(row.get('사이즈', 'FREE')).strip() if '사이즈' in df.columns and not pd.isna(row.get('사이즈')) else 'FREE'

                    # 발주수량 처리
                    try:
                        quantity = int(float(row.get('발주수량', 0)))
                        if quantity <= 0:
                            error_rows.append(f"행 {idx + 2}: 잘못된 발주수량")
                            skipped_count += 1
                            continue
                    except (ValueError, TypeError):
                        error_rows.append(f"행 {idx + 2}: 잘못된 발주수량 형식")
                        skipped_count += 1
                        continue

                    # 매장 처리
                    store_id = None
                    if '발주매장' in df.columns and not pd.isna(row.get('발주매장')):
                        store_name = str(row.get('발주매장')).strip()
                        for s in self.stores:
                            if s['name'] == store_name:
                                store_id = s['id']
                                break

                    # 발주 생성
                    order = {
                        'product_id': product['id'],
                        'quantity': quantity,
                        'shipped_quantity': 0,
                        'date': order_date,
                        'color': color,
                        'size': size,
                        'status': 'pending',
                        'store_id': store_id
                    }

                    # DB에 저장
                    self.data_manager.add_order(order)
                    imported_count += 1

                except Exception as e:
                    error_rows.append(f"행 {idx + 2}: {str(e)}")
                    skipped_count += 1
                    continue

            # 데이터 저장
            if imported_count > 0:
                # self.data_manager.save_data()
                # ← 자동저장 제거됨
                self._refresh_data_shortcuts()
                self.refresh_orders_list()
                self.update_order_dates()

            # 결과 메시지
            result_msg = f"✅ 등록 완료: {imported_count}개"
            if skipped_count > 0:
                result_msg += f"\n⚠️ 건너뜀: {skipped_count}개"

            if error_rows:
                result_msg += "\n\n오류 내역:\n" + "\n".join(error_rows[:10])
                if len(error_rows) > 10:
                    result_msg += f"\n... 외 {len(error_rows) - 10}개"

            messagebox.showinfo("엑셀 발주 등록 완료", result_msg)

        except Exception as e:
            messagebox.showerror("오류", f"엑셀 파일 처리 중 오류가 발생했습니다:\n{str(e)}")

def main():
    root = tk.Tk()
    app = InventoryManagementSystem(root)
    root.mainloop()

if __name__ == "__main__":
    main()
